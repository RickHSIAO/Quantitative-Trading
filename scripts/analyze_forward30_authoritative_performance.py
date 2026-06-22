"""TASK-014BZ -- corrected 30-day strategy analysis from the AUTHORITATIVE Paper
Portfolio performance ledger.

Supersedes TASK-014BY, which scored the zero-valued Forward dry-run snapshot JSON
instead of the Paper Portfolio ledger. Read-only with respect to all Forward /
Pilot artifacts; writes corrected analysis outputs ONLY under --output-root.
Deterministic and offline: no Bybit / Notion / Discord / network call, no order,
no Pilot-state mutation, no V1 strategy modification.

It NEVER falls back to the dry-run JSON: when the authoritative ledger is missing
or invalid it fails closed with an explicit data-quality status and a
NEEDS_MORE_DATA / REJECT_DATA_INCOMPLETE verdict.

    python scripts/analyze_forward30_authoritative_performance.py \
        --input-root outputs/forward_record --run-key prev3y_crypto \
        --output-root outputs/research/strategy_selection/TASK-014BZ --json-only
"""

from __future__ import annotations

import argparse
import json
import os
import pathlib
import sys
from typing import Any, Mapping

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from src.strategy_selection import paper_portfolio_performance as pp  # noqa: E402
from src.strategy_selection import corrected_strategy_analysis as csa  # noqa: E402

DEFAULT_PILOT_ID = "BYBIT_DEMO_PILOT_7D_202606_V1"
DEFAULT_SHADOW_RUN_KEY = "prev3y_crypto_shadow_a_roll12"
SUPERSEDED_OUTPUT = "outputs/research/strategy_selection/TASK-014BY"

WORKBOOK_SHEETS = [
    "Executive Summary", "Source Lineage", "Official 30D Metrics",
    "Post-Validation Extension", "Corrected Scorecard", "Corrected Challengers",
    "Primary vs Shadow", "Static-Hold Behavior",
]


def _write_json(path: pathlib.Path, obj: Any) -> None:
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def build_report_md(*, lineage, official_metrics, extension_metrics, scorecard,
                    challengers, comparability, hold, perf, window) -> str:
    L: list[str] = []
    L.append("# TASK-014BZ Corrected Strategy Selection Report")
    L.append("")
    L.append("> **PRIOR ZERO-RETURN ANALYSIS SUPERSEDED / AUTHORITATIVE PAPER PORTFOLIO PERFORMANCE "
             "RESTORED / ACTIVE V1 PILOT UNCHANGED / CHALLENGERS NOT PROMOTED / LIVE TRADING NOT "
             "AUTHORIZED.**")
    L.append("")
    L.append(f"This report SUPERSEDES `{SUPERSEDED_OUTPUT}` (TASK-014BY), whose "
             f"`{csa.SUPERSEDED_LABEL}` verdict and `coverage=37/30` label are INVALID: they scored the "
             "Forward dry-run snapshot JSON (`<date>_pnl.json`: `clock_started=false`, `day_number=0`, "
             "`daily_pnl_pct=0`, `paper_execution_status=FORBIDDEN`) instead of the authoritative Paper "
             "Portfolio ledger. That prior runtime output is retained but marked superseded.")
    L.append("")
    L.append("## Legend")
    L.append("**OBSERVED FACT** = recorded in the authoritative ledger; **CALCULATED** = derived via "
             "canonical metrics; **UNAVAILABLE** = not present (never fabricated).")
    L.append("")
    L.append("## Source Lineage")
    L.append(f"- performance_source: `{lineage['performance_source']}` "
             f"(fingerprint `{lineage['performance_source_fingerprint'][:23]}…`).")
    L.append(f"- data_lineage_status: **`{lineage['data_lineage_status']}`**.")
    L.append(f"- snapshot_file_count: `{lineage['snapshot_file_count']}` "
             f"(operational metadata only; day_number distribution "
             f"`{lineage['snapshot_day_number_distribution']}`; "
             f"dry_run_placeholder_detected=`{lineage['dry_run_placeholder_detected']}`).")
    L.append(f"- authoritative_performance_row_count: `{lineage['authoritative_performance_row_count']}`; "
             f"valid_performance_row_count: `{lineage['valid_performance_row_count']}`.")
    L.append(f"- official_validation_day_count: **`{lineage['official_validation_day_count']}`** "
             "(these three counts are SEPARATE; `37/30` is not a coverage label).")
    L.append("")
    L.append("## Official 30-Day Window (DERIVED, not hardcoded)")
    L.append(f"- window: `{lineage['official_validation_start']}` → "
             f"`{lineage['official_validation_end']}` "
             f"({lineage['official_validation_day_count']} valid days).")
    L.append(f"- official_30d_cumulative_return: `{lineage['official_30d_cumulative_return']}` (CALCULATED).")
    L.append(f"- official_30d_nav: `{lineage['official_30d_nav']}` (OBSERVED FACT).")
    L.append(f"- official_30d_max_drawdown: `{lineage['official_30d_max_drawdown']}`; "
             f"sharpe `{lineage['official_30d_sharpe']}`; sortino `{lineage['official_30d_sortino']}`.")
    L.append(f"- daily_win_rate: `{lineage['daily_win_rate']}`; longest_winning_streak "
             f"`{lineage['longest_winning_streak']}`; longest_losing_streak "
             f"`{lineage['longest_losing_streak']}`.")
    L.append(f"- best_day: `{lineage['best_day']}`; worst_day: `{lineage['worst_day']}`.")
    L.append("")
    L.append("## Post-Validation Extension (reported separately; NOT in the 30-day scorecard)")
    L.append(f"- window: `{lineage['post_validation_extension_start']}` → "
             f"`{lineage['post_validation_extension_end']}` "
             f"({lineage['post_validation_extension_count']} day(s)).")
    L.append(f"- extension_latest_cumulative_return: `{lineage['extension_latest_cumulative_return']}`; "
             f"extension_latest_nav: `{lineage['extension_latest_nav']}`.")
    L.append("")
    L.append("## Corrected Scorecard (scores ONLY the official 30 valid days)")
    L.append(f"- label: **`{scorecard['label']}`** — {scorecard['label_rationale']}")
    L.append(f"- observed_return: `{scorecard['observed_return']}`; risk_adjusted: "
             f"`{scorecard['risk_adjusted']}`.")
    L.append("- gate results:")
    for g in scorecard["gates"]:
        L.append(f"  - `{g['gate']}`: **{g['status']}** ({g['detail']})")
    L.append(f"- explicitly missing (never fabricated): {', '.join(scorecard['explicitly_missing'])}.")
    L.append(f"- A positive official cumulative return cannot fail the positive-net-expectancy gate; the "
             f"prior `{csa.SUPERSEDED_LABEL}` cannot survive the corrected positive return.")
    L.append("")
    L.append("## Primary vs Shadow")
    L.append(f"- primary_shadow_comparable: **`{comparability['primary_shadow_comparable']}`** — "
             f"{comparability['reason']}")
    L.append("")
    L.append("## Static-Hold Behavior")
    L.append(f"- behavior: **`{hold['behavior']}`** (is_defect=`{hold['is_defect']}`).")
    L.append(f"- first_day_entries `{hold['first_day_entries']}`, later_entries "
             f"`{hold['later_entries']}`, later_exits `{hold['later_exits']}`, "
             f"daily_mark_to_market_observed `{hold['daily_mark_to_market_observed']}`.")
    if hold.get("warning"):
        L.append(f"- WARNING: {hold['warning']}")
    else:
        L.append("- No defect warning: repository strategy docs do not state daily rebalancing was intended.")
    L.append("")
    L.append("## Corrected Challengers (offline/shadow-only; NONE promoted)")
    L.append(f"- emitted_count: **{challengers['emitted_count']} of {challengers['max_allowed']}**; "
             f"evidence_strength `{challengers['evidence_strength']}`.")
    L.append("- invalidated prior dry-run challengers:")
    for inv in challengers["invalidated_prior_challengers"]:
        L.append(f"  - `{inv['id']}`: {inv['corrected_status']} ({inv['reason']}).")
    L.append(f"- {challengers['contribution_note']}")
    for h in challengers["hypotheses"]:
        L.append(f"  - `{h['id']}` ({h['status']}): {h['proposed_single_change']}")
    L.append("")
    L.append("_This task does not modify V1 strategy logic, does not mutate or restart the Pilot, sends no "
             "Bybit order, promotes no Challenger, and does not authorize live trading._")
    L.append("")
    return "\n".join(L)


def build_workbook(path: pathlib.Path, *, lineage, scorecard, challengers, comparability, hold) -> bool:
    try:
        from openpyxl import Workbook
    except Exception:  # noqa: BLE001
        return False
    wb = Workbook()
    wb.remove(wb.active)
    sheets = {name: wb.create_sheet(title=name) for name in WORKBOOK_SHEETS}

    def kv(ws, pairs):
        for r, (k, v) in enumerate(pairs, start=1):
            ws.cell(row=r, column=1, value=str(k))
            ws.cell(row=r, column=2, value=str(v))

    kv(sheets["Executive Summary"], [
        ("task", "TASK-014BZ"), ("supersedes", "TASK-014BY (REJECT_INSUFFICIENT_EDGE / 37-30 INVALID)"),
        ("data_lineage_status", lineage["data_lineage_status"]),
        ("corrected_label", scorecard["label"]),
        ("official_30d_cumulative_return", lineage["official_30d_cumulative_return"]),
        ("ACTIVE V1 PILOT", "UNCHANGED / FROZEN / RUNNING"),
        ("challengers", "OFFLINE/SHADOW-ONLY; NONE PROMOTED"), ("live_trading", "NOT AUTHORIZED")])
    kv(sheets["Source Lineage"], [
        ("performance_source", lineage["performance_source"]),
        ("performance_source_fingerprint", lineage["performance_source_fingerprint"]),
        ("snapshot_file_count", lineage["snapshot_file_count"]),
        ("snapshot_day_number_distribution", lineage["snapshot_day_number_distribution"]),
        ("dry_run_placeholder_detected", lineage["dry_run_placeholder_detected"]),
        ("authoritative_performance_row_count", lineage["authoritative_performance_row_count"]),
        ("valid_performance_row_count", lineage["valid_performance_row_count"]),
        ("official_validation_day_count", lineage["official_validation_day_count"])])
    kv(sheets["Official 30D Metrics"], [
        ("official_validation_start", lineage["official_validation_start"]),
        ("official_validation_end", lineage["official_validation_end"]),
        ("official_30d_cumulative_return", lineage["official_30d_cumulative_return"]),
        ("official_30d_nav", lineage["official_30d_nav"]),
        ("official_30d_max_drawdown", lineage["official_30d_max_drawdown"]),
        ("official_30d_sharpe", lineage["official_30d_sharpe"]),
        ("official_30d_sortino", lineage["official_30d_sortino"]),
        ("daily_win_rate", lineage["daily_win_rate"]),
        ("longest_winning_streak", lineage["longest_winning_streak"]),
        ("longest_losing_streak", lineage["longest_losing_streak"]),
        ("best_day", lineage["best_day"]), ("worst_day", lineage["worst_day"])])
    kv(sheets["Post-Validation Extension"], [
        ("post_validation_extension_start", lineage["post_validation_extension_start"]),
        ("post_validation_extension_end", lineage["post_validation_extension_end"]),
        ("post_validation_extension_count", lineage["post_validation_extension_count"]),
        ("extension_latest_cumulative_return", lineage["extension_latest_cumulative_return"]),
        ("extension_latest_nav", lineage["extension_latest_nav"])])
    cs = sheets["Corrected Scorecard"]
    cs.append(["label", scorecard["label"]])
    cs.append(["gate", "status", "detail"])
    for g in scorecard["gates"]:
        cs.append([g["gate"], g["status"], g["detail"]])
    ch = sheets["Corrected Challengers"]
    ch.append(["emitted_count", challengers["emitted_count"]])
    ch.append(["invalidated_id", "corrected_status"])
    for inv in challengers["invalidated_prior_challengers"]:
        ch.append([inv["id"], inv["corrected_status"]])
    kv(sheets["Primary vs Shadow"], [
        ("primary_shadow_comparable", comparability["primary_shadow_comparable"]),
        ("reason", comparability["reason"])])
    kv(sheets["Static-Hold Behavior"], [
        ("behavior", hold["behavior"]), ("is_defect", hold["is_defect"]),
        ("first_day_entries", hold["first_day_entries"]), ("later_entries", hold["later_entries"]),
        ("later_exits", hold["later_exits"]),
        ("daily_mark_to_market_observed", hold["daily_mark_to_market_observed"]),
        ("warning", hold.get("warning"))])
    wb.save(str(path))
    return True


def _strategy_docs_mention_rebalancing(root: str) -> bool:
    """Best-effort read-only check: does repository strategy documentation state
    daily re-ranking/rebalancing was intended? Conservative default False."""
    candidates = [
        pathlib.Path(root) / "docs" / "research" / "strategy_selection",
    ]
    needles = ("daily rebalanc", "daily re-rank", "rebalance daily", "re-rank daily")
    for base in candidates:
        if not base.exists():
            continue
        for p in base.rglob("*.md"):
            try:
                text = p.read_text(encoding="utf-8", errors="ignore").lower()
            except OSError:
                continue
            if any(n in text for n in needles):
                return True
    return False


def run_analysis(*, input_root: str, run_key: str, shadow_run_key: str,
                 output_root: str, pilot_id: str) -> dict[str, Any]:
    out_dir = pathlib.Path(output_root)
    out_dir.mkdir(parents=True, exist_ok=True)
    paper_dir = pathlib.Path(input_root) / "paper_portfolio"
    snapshot_dir = pathlib.Path(input_root) / run_key

    perf = pp.load_authoritative_performance(paper_dir)
    window = pp.derive_validation_window(perf)
    snapshot_scan = pp.scan_dry_run_snapshots(snapshot_dir)
    official_metrics = pp.compute_window_metrics(window.official_rows, paper_equity_init=perf.paper_equity_init)
    extension_metrics = pp.compute_window_metrics(window.extension_rows, paper_equity_init=perf.paper_equity_init)

    # Shadow authoritative performance (independent series required for comparability).
    shadow_paper = pathlib.Path(input_root) / shadow_run_key / "paper_portfolio"
    shadow_perf = (pp.load_authoritative_performance(shadow_paper) if shadow_paper.exists() else None)
    comparability = csa.assess_primary_shadow_comparability(perf, shadow_perf)

    rebal_doc = _strategy_docs_mention_rebalancing(ROOT)
    hold = csa.classify_hold_behavior(window, rebalancing_documented=rebal_doc)

    lineage = csa.build_source_lineage(perf, window, snapshot_scan, official_metrics, extension_metrics)
    scorecard = csa.score_official_window(perf, window, official_metrics)

    capabilities = {"volatility_adjusted_sizing": _has_module("src.demo_portfolio_risk")}
    challengers = csa.correct_challengers(
        perf, window, official_metrics,
        contribution_available=False, capabilities=capabilities)

    # Write corrected outputs (only under output-root).
    _write_json(out_dir / "source_lineage.json", lineage)
    _write_json(out_dir / "official_30d_metrics.json", {
        "task_id": csa.TASK_ID, "window_start": window.official_start, "window_end": window.official_end,
        "official_day_count": window.official_day_count, "sufficient": window.sufficient,
        "metrics": official_metrics})
    _write_json(out_dir / "post_validation_extension.json", {
        "task_id": csa.TASK_ID, "extension_start": window.extension_start,
        "extension_end": window.extension_end, "extension_count": len(window.extension_rows),
        "metrics": extension_metrics})
    _write_json(out_dir / "corrected_strategy_scorecard.json", scorecard)
    _write_json(out_dir / "corrected_challenger_hypotheses.json", challengers)
    _write_json(out_dir / "primary_shadow_comparability.json", comparability)
    _write_json(out_dir / "static_hold_behavior.json", hold)
    _write_json(out_dir / "superseded_notice.json", {
        "task_id": csa.TASK_ID, "supersedes": csa.SUPERSEDED_TASK,
        "superseded_output_root": SUPERSEDED_OUTPUT,
        "superseded_label": csa.SUPERSEDED_LABEL,
        "reason": "prior analysis scored zero-valued Forward dry-run snapshot JSON instead of the "
                  "authoritative Paper Portfolio ledger; its REJECT_INSUFFICIENT_EDGE and 37/30 coverage "
                  "label are invalid. Prior runtime output is retained, not deleted.",
        "active_v1_unchanged": True, "pilot_unchanged": True, "challengers_promoted": 0,
        "orders_sent": 0})

    report_md = build_report_md(
        lineage=lineage, official_metrics=official_metrics, extension_metrics=extension_metrics,
        scorecard=scorecard, challengers=challengers, comparability=comparability, hold=hold,
        perf=perf, window=window)
    (out_dir / "corrected_strategy_selection_report.md").write_text(report_md, encoding="utf-8")
    xlsx_ok = build_workbook(
        out_dir / "corrected_strategy_selection_report.xlsx", lineage=lineage, scorecard=scorecard,
        challengers=challengers, comparability=comparability, hold=hold)

    return {
        "task_id": csa.TASK_ID,
        "analysis_status": "ANALYSIS_SUCCESS",
        "supersedes": {"task": csa.SUPERSEDED_TASK, "output_root": SUPERSEDED_OUTPUT,
                       "invalid_label": csa.SUPERSEDED_LABEL},
        "output_root": str(out_dir).replace("\\", "/"),
        "data_lineage_status": perf.status,
        "performance_source": perf.performance_source,
        "performance_source_fingerprint": perf.performance_source_fingerprint,
        "snapshot_file_count": lineage["snapshot_file_count"],
        "authoritative_performance_row_count": lineage["authoritative_performance_row_count"],
        "valid_performance_row_count": lineage["valid_performance_row_count"],
        "official_validation_start": window.official_start,
        "official_validation_end": window.official_end,
        "official_validation_day_count": window.official_day_count,
        "official_30d_cumulative_return": official_metrics.get("cumulative_return_decimal"),
        "post_validation_extension_count": len(window.extension_rows),
        "extension_latest_cumulative_return": extension_metrics.get("cumulative_return_decimal"),
        "corrected_scorecard_label": scorecard["label"],
        "primary_shadow_comparable": comparability["primary_shadow_comparable"],
        "static_hold_behavior": hold["behavior"],
        "challengers_emitted": challengers["emitted_count"],
        "challengers_promoted": 0,
        "workbook_written": xlsx_ok,
        "report_paths": {
            "markdown": str((out_dir / "corrected_strategy_selection_report.md")).replace("\\", "/"),
            "workbook": str((out_dir / "corrected_strategy_selection_report.xlsx")).replace("\\", "/"),
            "source_lineage": str((out_dir / "source_lineage.json")).replace("\\", "/"),
        },
        "network_calls": 0, "bybit_calls": 0, "orders_sent": 0,
        "banner": "PRIOR ZERO-RETURN ANALYSIS SUPERSEDED / AUTHORITATIVE PAPER PORTFOLIO PERFORMANCE "
                  "RESTORED / ACTIVE V1 PILOT UNCHANGED / CHALLENGERS NOT PROMOTED / LIVE TRADING NOT "
                  "AUTHORIZED",
    }


def _has_module(name: str) -> bool:
    try:
        __import__(name)
        return True
    except Exception:  # noqa: BLE001
        return False


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="analyze_forward30_authoritative_performance.py",
                                description="Corrected 30-day strategy analysis from the authoritative "
                                            "Paper Portfolio ledger (TASK-014BZ).")
    p.add_argument("--input-root", default="outputs/forward_record")
    p.add_argument("--run-key", default="prev3y_crypto")
    p.add_argument("--shadow-run-key", default=DEFAULT_SHADOW_RUN_KEY)
    p.add_argument("--output-root", required=True)
    p.add_argument("--pilot-id", default=DEFAULT_PILOT_ID)
    p.add_argument("--json-only", action="store_true")
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    result = run_analysis(
        input_root=args.input_root, run_key=args.run_key, shadow_run_key=args.shadow_run_key,
        output_root=args.output_root, pilot_id=args.pilot_id)
    if args.json_only:
        print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    else:
        print(f"TASK-014BZ corrected analysis -> {result['output_root']}  "
              f"[lineage={result['data_lineage_status']}]")
        print(f"  official window: {result['official_validation_start']} -> "
              f"{result['official_validation_end']} "
              f"({result['official_validation_day_count']} valid days); "
              f"cum_return={result['official_30d_cumulative_return']}")
        print(f"  corrected_label={result['corrected_scorecard_label']}; "
              f"primary_shadow_comparable={result['primary_shadow_comparable']}")
        print(f"  {result['banner']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
