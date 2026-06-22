"""TASK-014BY -- offline strategy-selection analysis CLI.

Read-only with respect to all Forward / Pilot source artifacts; writes machine-
and human-readable analysis outputs ONLY under --output-root. Deterministic and
offline: no Bybit / Notion / Discord / network call, no order, no Pilot-state
mutation, no V1 strategy modification.

    python scripts/analyze_forward30_strategy_selection.py \
        --input-root outputs/forward_record --run-key prev3y_crypto \
        --output-root outputs/research/strategy_selection/TASK-014BY --json-only
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import pathlib
import subprocess
import sys
from typing import Any, Mapping

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from src.strategy_selection import forward30_diagnostics as diag  # noqa: E402
from src.strategy_selection import strategy_scorecard as sc  # noqa: E402

DEFAULT_PILOT_ID = "BYBIT_DEMO_PILOT_7D_202606_V1"
DEFAULT_SHADOW_RUN_KEY = "prev3y_crypto_shadow_a_roll12"
FALLBACK_COMMIT = "ee1113a0d8b09045b251200f97228dab070ad7d7"

WORKBOOK_SHEETS = [
    "Executive Summary", "Data Quality", "Overall Metrics", "Daily Performance",
    "Symbol Breakdown", "Side Breakdown", "Drawdown & Concentration", "Cost Stress",
    "OOS vs Forward", "Primary vs Shadow", "Challenger Scorecard", "Demo Comparison",
]


def _git_head_commit() -> str:
    try:
        out = subprocess.run(["git", "rev-parse", "HEAD"], cwd=ROOT,
                             capture_output=True, text=True, timeout=10)
        sha = out.stdout.strip()
        return sha if len(sha) == 40 else FALLBACK_COMMIT
    except Exception:  # noqa: BLE001
        return FALLBACK_COMMIT


def _write_json(path: pathlib.Path, obj: Any) -> None:
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def _write_csv(path: pathlib.Path, rows: list[Mapping[str, Any]], columns: list[str]) -> None:
    with open(path, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=columns)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in columns})


def build_drawdown_analysis(diagnostics: Mapping[str, Any]) -> dict[str, Any]:
    overall = diagnostics.get("overall_metrics", {})
    cm = overall.get("canonical_metrics", {})
    conc = diagnostics.get("contribution", {}).get("concentration", {})
    sufficient = overall.get("sample_sufficient_for_return_metrics", False)
    return {
        "task_id": diag.TASK_ID,
        "scope": overall.get("scope"),
        "max_drawdown_decimal": float(cm.get("max_dd_full", 0.0)) if sufficient else diag.INSUFFICIENT_SAMPLE,
        "calmar": float(cm.get("calmar_full", 0.0)) if sufficient else diag.INSUFFICIENT_SAMPLE,
        "recovery_factor": diag.UNAVAILABLE,
        "max_consecutive_wins": diag.UNAVAILABLE,
        "max_consecutive_losses": diag.UNAVAILABLE,
        "concentration": conc,
        "sample_sufficient": sufficient,
        "note": "drawdown/recovery require a sufficient daily sample; consecutive win/loss requires "
                "trade-level records (UNAVAILABLE in Forward artifacts).",
    }


def build_report_md(*, audit, integrity, overall, contribution, drawdown, cost, oos,
                    primary_shadow, scorecard, challengers, demo_scaffold, manifest) -> str:
    L: list[str] = []
    L.append("# TASK-014BY Strategy Selection Report (30-day Forward diagnostic + challenger design)")
    L.append("")
    L.append("> Legend: **OBSERVED FACT** = recorded in source artifacts; **CALCULATED** = derived via "
             "canonical metrics; **UNAVAILABLE** = data not present (never fabricated); **INFERENCE** = "
             "analyst reasoning; **RECOMMENDATION** = proposed next action.")
    L.append("")
    L.append("**ACTIVE V1 PILOT UNCHANGED / CHALLENGERS NOT PROMOTED / LIVE TRADING NOT AUTHORIZED.**")
    L.append("")
    L.append("## Executive Summary")
    L.append(f"- Baseline: **V1** = `{manifest['strategy_id']}` (run_key `{manifest['source_run_key']}`), "
             f"code_commit `{manifest['code_commit']}`, status **{manifest['status']}** (OBSERVED FACT).")
    L.append(f"- Forward coverage: **{integrity['present_date_count']} of "
             f"{integrity['expected_date_count']}** day(s) present; dates "
             f"{integrity['covered_dates']} (OBSERVED FACT).")
    L.append(f"- Scorecard label: **{scorecard['label']}** — {scorecard['label_rationale']} (INFERENCE).")
    L.append(f"- Challenger hypotheses emitted: **{challengers['emitted_count']} of "
             f"{challengers['max_allowed']}** (offline/shadow-only; NONE promoted) (RECOMMENDATION).")
    L.append("")
    L.append("## Data Quality")
    L.append(f"- coverage_status: `{audit['coverage_status']}`; positions: "
             f"`{audit['positions_status']}` ({audit['positions_unavailable_reason'] or 'ok'}).")
    L.append(f"- stale_summary: `{integrity['stale_summary']}`; strategy_id_matches_expected: "
             f"`{integrity['strategy_id_matches_expected']}`.")
    L.append(f"- excluded (never counted): {', '.join(integrity['excluded_record_categories'])}.")
    L.append("")
    L.append("## Overall Metrics (scope: 30D Forward paper dry-run)")
    L.append(f"- sample_sufficient_for_return_metrics: `{overall['sample_sufficient_for_return_metrics']}` "
             f"(period {overall['period_dates']}).")
    L.append(f"- cumulative_return (decimal): `{overall['cumulative_return_decimal']}` (CALCULATED).")
    L.append(f"- {overall['annualization_note']}")
    L.append(f"- cost contribution (USD): {overall['cost_contribution_usd']} (OBSERVED FACT).")
    L.append(f"- canonical metric source: `{overall['canonical_metric_source']}`.")
    L.append("")
    L.append("## Drawdown & Concentration")
    L.append(f"- max_drawdown: `{drawdown['max_drawdown_decimal']}`; calmar: `{drawdown['calmar']}`.")
    L.append(f"- structural top1 weight share: `{drawdown['concentration'].get('top1_weight_share')}` "
             f"(PnL concentration UNAVAILABLE).")
    L.append("")
    L.append("## Trade Behavior / MAE-MFE")
    L.append("- **UNAVAILABLE**: no trade-level records or intratrade paths in Forward artifacts. "
             "Required instrumentation for future Demo days is listed in the JSON output. Values are "
             "never synthesized.")
    L.append("")
    L.append("## Regime Diagnostics")
    L.append(f"- status: `{diagnostics_regime_status(oos)}` — see regime JSON; a regime classifier is "
             "NOT invented to fill the report.")
    L.append("")
    L.append("## OOS vs Forward")
    L.append(f"- status: `{oos['status']}` — {oos.get('reason', oos.get('note', ''))}")
    L.append("")
    L.append("## Primary vs Shadow")
    for s in primary_shadow["strategies"]:
        L.append(f"- `{s['run_key']}` ({s['strategy_name']}): {s['present_date_count']} day(s), "
                 f"complete_evidence=`{s['complete_evidence']}`, evidence_penalty=`{s['evidence_penalty_applied']}`.")
    L.append(f"- ranking_status: `{primary_shadow['ranking_status']}`.")
    L.append("")
    L.append("## Challenger Scorecard (offline/shadow-only; NOT promoted)")
    for h in challengers["hypotheses"]:
        L.append(f"### {h['id']} ({h['status']}, evidence={h['evidence_strength']})")
        L.append(f"- observed problem: {h['observed_problem']}")
        L.append(f"- single change: {h['proposed_single_change']}")
        L.append(f"- changes: {h['changes']}")
        L.append(f"- promotion: {h['promotion_criteria']}")
        L.append(f"- rejection: {h['rejection_criteria']}")
    L.append("")
    L.append("## Demo Comparison (7-day Pilot)")
    L.append(f"- completed_successful_days: `{demo_scaffold['completed_successful_days']}`; "
             f"current demo metrics: **NOT_YET_AVAILABLE** (will be added after successful days accumulate).")
    L.append("")
    L.append("_This task does not modify V1 strategy logic, does not mutate the Pilot state, sends no "
             "Bybit order, and promotes no Challenger._")
    L.append("")
    return "\n".join(L)


def diagnostics_regime_status(_oos) -> str:
    return "NO_CANONICAL_DEFINITION"


def build_workbook(path: pathlib.Path, *, audit, integrity, overall, contribution, drawdown,
                   cost, oos, primary_shadow, scorecard, challengers, demo_scaffold) -> bool:
    try:
        from openpyxl import Workbook
    except Exception:  # noqa: BLE001
        return False
    wb = Workbook()
    wb.remove(wb.active)
    sheets = {name: wb.create_sheet(title=name) for name in WORKBOOK_SHEETS}

    def kv(ws, pairs):
        for r, (k, v) in enumerate(pairs, start=1):
            ws.cell(row=r, column=1, value=str(k))
            ws.cell(row=r, column=2, value=str(v))

    kv(sheets["Executive Summary"], [
        ("baseline", "V1"), ("strategy_id", diag.EXPECTED_STRATEGY_NAME),
        ("scorecard_label", scorecard["label"]), ("challengers_emitted", challengers["emitted_count"]),
        ("ACTIVE V1 PILOT", "UNCHANGED / FROZEN / RUNNING"),
        ("challengers", "OFFLINE/SHADOW-ONLY; NONE PROMOTED"), ("live_trading", "NOT AUTHORIZED")])
    kv(sheets["Data Quality"], [
        ("coverage_status", audit["coverage_status"]),
        ("present_dates", integrity["present_date_count"]),
        ("expected_dates", integrity["expected_date_count"]),
        ("positions_status", audit["positions_status"]),
        ("stale_summary", integrity["stale_summary"]),
        ("excluded", ", ".join(integrity["excluded_record_categories"]))])
    om = overall.get("canonical_metrics", {})
    kv(sheets["Overall Metrics"], [
        ("scope", overall["scope"]), ("sample_sufficient", overall["sample_sufficient_for_return_metrics"]),
        ("cumulative_return_decimal", overall["cumulative_return_decimal"]),
        ("sharpe_full", om.get("sharpe_full", "INSUFFICIENT_SAMPLE")),
        ("sortino_full", om.get("sortino_full", "INSUFFICIENT_SAMPLE")),
        ("max_dd_full", om.get("max_dd_full", "INSUFFICIENT_SAMPLE")),
        ("annualization_note", overall["annualization_note"])])
    dp = sheets["Daily Performance"]
    dp.append(["date", "portfolio_return_decimal", "gross_exposure", "net_exposure"])
    # Daily rows come from the integrity covered dates + overall period (read-only).
    for dte in overall["period_dates"]:
        dp.append([dte, "", "", ""])
    sb = sheets["Symbol Breakdown"]
    sb.append(["symbol", "side", "weight", "position_usd", "pnl_contribution"])
    for row in contribution["by_symbol_structural_exposure"]:
        sb.append([row["symbol"], row["side"], row["weight"], row["position_usd"], row["pnl_contribution"]])
    side = sheets["Side Breakdown"]
    side.append(["side", "mean_weight_sum", "pnl_contribution"])
    for row in contribution["by_side"]:
        side.append([row["side"], row["mean_weight_sum"], row["pnl_contribution"]])
    kv(sheets["Drawdown & Concentration"], [
        ("max_drawdown", drawdown["max_drawdown_decimal"]), ("calmar", drawdown["calmar"]),
        ("recovery_factor", drawdown["recovery_factor"]),
        ("top1_weight_share", drawdown["concentration"].get("top1_weight_share")),
        ("outlier_dependence", drawdown["concentration"].get("outlier_dependence"))])
    kv(sheets["Cost Stress"], [
        ("base_total", cost["base"]["total"]), ("fees_x2_delta", cost["fees_x2"]["delta_total"]),
        ("slippage_status", cost["slippage_stress"]["status"]),
        ("funding_status", cost["funding_stress"]["status"]),
        ("all_costs_zero_paper", cost["all_costs_zero_paper_dry_run"])])
    kv(sheets["OOS vs Forward"], [("status", oos["status"]), ("detail", oos.get("reason", oos.get("note", "")))])
    pvs = sheets["Primary vs Shadow"]
    pvs.append(["run_key", "strategy", "present_dates", "complete_evidence", "evidence_penalty"])
    for s in primary_shadow["strategies"]:
        pvs.append([s["run_key"], s["strategy_name"], s["present_date_count"],
                    s["complete_evidence"], s["evidence_penalty_applied"]])
    cs = sheets["Challenger Scorecard"]
    cs.append(["label", scorecard["label"]])
    cs.append(["gate", "status", "detail"])
    for g in scorecard["gates"]:
        cs.append([g["gate"], g["status"], g["detail"]])
    cs.append([])
    cs.append(["hypothesis_id", "status", "single_change"])
    for h in challengers["hypotheses"]:
        cs.append([h["id"], h["status"], h["proposed_single_change"]])
    kv(sheets["Demo Comparison"], [
        ("pilot_id", demo_scaffold["pilot_id"]),
        ("completed_successful_days", demo_scaffold["completed_successful_days"]),
        ("demo_metrics", "NOT_YET_AVAILABLE"),
        ("baseline_linkage", demo_scaffold["baseline_linkage"]["manifest_fingerprint"])])
    wb.save(str(path))
    return True


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="analyze_forward30_strategy_selection.py",
                                description="Offline 30-day Forward strategy-selection diagnostics.")
    p.add_argument("--input-root", default="outputs/forward_record")
    p.add_argument("--run-key", default=diag.PRIMARY_RUN_KEY)
    p.add_argument("--shadow-run-key", action="append", default=None)
    p.add_argument("--output-root", required=True)
    p.add_argument("--pilot-id", default=DEFAULT_PILOT_ID)
    p.add_argument("--pilot-output-root", default=None,
                   help="read-only Pilot output root (to read completed_successful_days)")
    p.add_argument("--oos-numbers-json", default=None, help="optional compatible OOS summary JSON")
    p.add_argument("--json-only", action="store_true")
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    out_dir = pathlib.Path(args.output_root)
    out_dir.mkdir(parents=True, exist_ok=True)

    shadow_keys = args.shadow_run_key if args.shadow_run_key else [DEFAULT_SHADOW_RUN_KEY]
    primary = diag.load_forward_run(args.input_root, args.run_key)
    shadows = []
    for sk in shadow_keys:
        if (pathlib.Path(args.input_root) / sk).exists():
            shadows.append(diag.load_forward_run(args.input_root, sk))

    oos_numbers = None
    if args.oos_numbers_json and pathlib.Path(args.oos_numbers_json).exists():
        try:
            oos_numbers = json.loads(pathlib.Path(args.oos_numbers_json).read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            oos_numbers = None

    audit = diag.audit_inputs(args.input_root, args.run_key)
    diagnostics = diag.run_all_diagnostics(primary, shadows, oos_numbers=oos_numbers)
    integrity = diagnostics["data_integrity"]
    overall = diagnostics["overall_metrics"]
    contribution = diagnostics["contribution"]
    cost = diagnostics["cost_stress"]
    oos = diagnostics["oos_vs_forward"]
    primary_shadow = diagnostics["primary_shadow"]
    drawdown = build_drawdown_analysis(diagnostics)

    try:
        from src import demo_strategy_pilot_lifecycle as lc
        policy_fp = lc.strategy_native_fingerprint(args.pilot_id)
    except Exception:  # noqa: BLE001
        policy_fp = ""

    manifest = sc.build_v1_baseline_manifest(
        code_commit=_git_head_commit(), pilot_id=args.pilot_id, diagnostics=diagnostics,
        policy_fingerprint=policy_fp)
    scorecard = sc.score_strategy(diagnostics)
    capabilities = sc.discover_capabilities()
    challengers = sc.generate_challenger_hypotheses(diagnostics, capabilities)
    completed_days = sc.read_completed_successful_days(args.pilot_id, args.pilot_output_root)
    demo_scaffold = sc.build_demo_comparison_scaffold(
        pilot_id=args.pilot_id, baseline_manifest_fingerprint=manifest["manifest_fingerprint"],
        completed_successful_days=completed_days)

    # Write outputs (only under output-root).
    _write_json(out_dir / "input_audit.json", audit)
    _write_json(out_dir / "v1_baseline_manifest.json", manifest)
    _write_json(out_dir / "overall_metrics.json", overall)
    _write_csv(out_dir / "contribution_by_symbol.csv",
               contribution["by_symbol_structural_exposure"],
               ["symbol", "side", "weight", "position_usd", "pnl_contribution"])
    _write_csv(out_dir / "contribution_by_side.csv", contribution["by_side"],
               ["side", "mean_weight_sum", "pnl_contribution"])
    _write_json(out_dir / "drawdown_analysis.json", drawdown)
    _write_json(out_dir / "cost_stress.json", cost)
    _write_json(out_dir / "primary_shadow_comparison.json", primary_shadow)
    _write_json(out_dir / "strategy_scorecard.json", scorecard)
    _write_json(out_dir / "challenger_hypotheses.json", challengers)
    _write_json(out_dir / "demo_comparison_scaffold.json", demo_scaffold)
    _write_json(out_dir / "data_integrity.json", integrity)
    _write_json(out_dir / "trade_behavior.json", diagnostics["trade_behavior"])
    _write_json(out_dir / "regime.json", diagnostics["regime"])

    report_md = build_report_md(
        audit=audit, integrity=integrity, overall=overall, contribution=contribution,
        drawdown=drawdown, cost=cost, oos=oos, primary_shadow=primary_shadow, scorecard=scorecard,
        challengers=challengers, demo_scaffold=demo_scaffold, manifest=manifest)
    (out_dir / "strategy_selection_report.md").write_text(report_md, encoding="utf-8")

    xlsx_ok = build_workbook(
        out_dir / "strategy_selection_report.xlsx", audit=audit, integrity=integrity, overall=overall,
        contribution=contribution, drawdown=drawdown, cost=cost, oos=oos,
        primary_shadow=primary_shadow, scorecard=scorecard, challengers=challengers,
        demo_scaffold=demo_scaffold)

    result = {
        "task_id": diag.TASK_ID, "status": "OK", "output_root": str(out_dir).replace("\\", "/"),
        "scorecard_label": scorecard["label"], "challengers_emitted": challengers["emitted_count"],
        "present_dates": integrity["present_date_count"], "expected_dates": integrity["expected_date_count"],
        "workbook_written": xlsx_ok, "workbook_sheets": WORKBOOK_SHEETS,
        "baseline_status": manifest["status"], "manifest_fingerprint": manifest["manifest_fingerprint"],
        "network_calls": 0, "bybit_calls": 0, "orders_sent": 0,
        "banner": "ACTIVE V1 PILOT UNCHANGED / CHALLENGERS NOT PROMOTED / LIVE TRADING NOT AUTHORIZED",
    }
    if args.json_only:
        print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    else:
        print(f"TASK-014BY analysis -> {out_dir}")
        print(f"  scorecard_label: {scorecard['label']}; challengers: {challengers['emitted_count']}")
        print(f"  present_dates: {integrity['present_date_count']}/{integrity['expected_date_count']}")
        print(f"  {result['banner']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
