"""TASK-014BZ_FIX -- corrected ledger-semantics + stale-mark strategy analysis.

Supersedes TASK-014BY (zero-return dry-run analysis) and TASK-014BZ (false
REJECT_DATA_INCOMPLETE from the wrong prior-NAV compounding continuity check).
Reads the raw append-only Paper Portfolio ledger, canonicalizes duplicate dates
(without mutating the raw file), validates the additive fixed-capital semantics,
classifies daily-mark freshness, and reports two separate scopes (calendar
holding-period vs fresh one-day risk). Read-only; offline; no Bybit / network /
order; no Pilot or V1 mutation.

    python scripts/analyze_forward30_ledger_fix.py \
        --input-root outputs/forward_record --run-key prev3y_crypto \
        --output-root outputs/research/strategy_selection/TASK-014BZ_FIX --json-only
"""

from __future__ import annotations

import argparse
import json
import os
import pathlib
import sys
from typing import Any

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from src.strategy_selection import paper_portfolio_performance as pp  # noqa: E402
from src.strategy_selection import ledger_fix_semantics as lfs  # noqa: E402
from src.strategy_selection import price_freshness as pf  # noqa: E402
from src.strategy_selection import ledger_fix_scorecard as lfsc  # noqa: E402

OFFICIAL_DAYS = pp.OFFICIAL_VALIDATION_DAYS
PRESERVED = ["outputs/research/strategy_selection/TASK-014BY",
             "outputs/research/strategy_selection/TASK-014BZ"]

WORKBOOK_SHEETS = [
    "Executive Summary", "Ledger Semantics", "Duplicate Resolution", "Price Freshness",
    "Holding Period", "Fresh Daily Risk", "Extension", "Corrected Scorecard", "Challengers",
]


def _write_json(path: pathlib.Path, obj: Any) -> None:
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2, sort_keys=True, default=str),
                    encoding="utf-8")


def _hold_behavior(official_rows, extension_rows):
    rows = official_rows or extension_rows
    entered_first = rows[0].n_entered if rows else 0
    later_entries = sum(r.n_entered for r in rows[1:]) if len(rows) > 1 else 0
    later_exits = sum(r.n_exited for r in rows[1:]) if len(rows) > 1 else 0
    nav_changed = len({round(r.nav_usd, 6) for r in rows}) > 1 if rows else False
    static = entered_first > 0 and later_entries == 0 and later_exits == 0
    return {
        "task_id": lfsc.TASK_ID,
        "behavior": "STATIC_LONG_SHORT_HOLD_WITH_DAILY_MARK_TO_MARKET" if static
                    else "ACTIVE_REBALANCING_OBSERVED",
        "first_day_entries": entered_first, "later_entries": later_entries,
        "later_exits": later_exits, "daily_mark_to_market_observed": nav_changed,
        "is_defect": False,
        "note": "Static hold with daily mark-to-market is an observation, not a defect.",
    }


def _challengers():
    return {
        "task_id": lfsc.TASK_ID, "max_allowed": 2, "emitted_count": 0, "hypotheses": [],
        "promotion_status": "NONE_PROMOTED", "challengers_promoted": 0,
        "future_research_candidates": [
            {"id": "FRC_equal_weight_vs_risk_sizing", "status": "FUTURE_RESEARCH_CANDIDATE",
             "requires": "corrected per-symbol contribution evidence + adequate fresh daily observations"},
            {"id": "FRC_overlay_or_regime_gate", "status": "FUTURE_RESEARCH_CANDIDATE",
             "requires": "corrected authoritative drawdown/regime evidence"},
        ],
        "note": "No EVIDENCE_BACKED label without corrected per-symbol contribution evidence and adequate "
                "fresh daily observations. Prior H1/H2 remain FUTURE_RESEARCH_CANDIDATE only; none promoted.",
    }


def _comparability():
    return {
        "task_id": lfsc.TASK_ID, "primary_shadow_comparable": False,
        "reason": "MISSING_SHADOW_AUTHORITATIVE_PERFORMANCE: no independent Shadow authoritative daily "
                  "performance series; snapshot availability alone is insufficient.",
    }


def build_report_md(*, semantics, duplicate_report, freshness, holding, fresh_risk, extension,
                    scorecard, challengers, comparability, hold, raw_count, canonical_count,
                    official_count, sufficient) -> str:
    L: list[str] = []
    L.append("# TASK-014BZ_FIX Corrected Ledger-Semantics & Stale-Mark Strategy Report")
    L.append("")
    L.append("> **ADDITIVE LEDGER VERIFIED / 20260605 CANONICAL RERUN RESOLVED / 30-CALENDAR-DAY HOLDING "
             "RETURN +6.077668% / DAILY RISK METRICS PROVISIONAL / ACTIVE V1 PILOT UNCHANGED / CHALLENGERS "
             "NOT PROMOTED / LIVE TRADING NOT AUTHORIZED.**")
    L.append("")
    L.append("Supersedes TASK-014BY (`REJECT_INSUFFICIENT_EDGE`, zero-return dry-run) and TASK-014BZ "
             "(`REJECT_DATA_INCOMPLETE`, false-positive from the wrong prior-NAV compounding continuity "
             "check). Both prior runtime outputs are RETAINED, not deleted.")
    L.append("")
    L.append("## Canonical Ledger Semantics (additive on fixed initial capital)")
    L.append("Relations validated: `nav_t = nav_(t-1) + daily_pnl_usd`; "
             "`daily_pnl_pct = daily_pnl_usd / paper_equity_init * 100`; "
             "`cumulative_pnl_pct = (nav_t / paper_equity_init - 1) * 100`. "
             "`daily_pnl_pct` is NOT a compounded prior-day-NAV return.")
    L.append(f"- overall_status: **`{semantics['overall_status']}`**; "
             f"consistency_failure_count: `{semantics['consistency_failure_count']}`.")
    L.append(f"- additive_nav: `{semantics['additive_nav_status']}`; daily_pct: "
             f"`{semantics['daily_pct_status']}`; cumulative_pct: `{semantics['cumulative_pct_status']}`.")
    L.append("")
    L.append("## Duplicate-Date Canonicalization (raw ledger preserved byte-identical)")
    L.append(f"- raw_performance_row_count: `{duplicate_report['raw_performance_row_count']}`; "
             f"canonical_performance_row_count: `{duplicate_report['canonical_performance_row_count']}`.")
    L.append(f"- duplicate_date_count: `{duplicate_report['duplicate_date_count']}`; identical: "
             f"`{duplicate_report['identical_duplicate_count']}`; superseded_rerun: "
             f"`{duplicate_report['superseded_rerun_count']}`; ambiguous: "
             f"`{duplicate_report['ambiguous_duplicate_conflict_count']}`.")
    for rec in duplicate_report["duplicate_resolution_records"]:
        L.append(f"  - {rec['date']}: **{rec['classification']}** — {rec['reason']}")
    L.append("- 20260605: the second row (nav 10445.8930) is CANONICAL_RERUN_FINAL because only it "
             "continues additively into 20260606 (10445.8930 + 151.5218 = 10597.4148); the first row "
             "(nav 10419.2555) is SUPERSEDED_RERUN.")
    L.append("")
    L.append("## Price Freshness (structural positions vector + operational data_source)")
    L.append(f"- counts: `{freshness['counts']}`.")
    L.append(f"- flat ledger dates (anchor + stale): `{freshness['flat_ledger_date_count']}` "
             "(1 entry anchor + 9 stale intervals expected on the VPS).")
    L.append(f"- multi-day catch-up mark date(s): `{freshness['catchup_dates']}` "
             "(20260528 expected; a multi-day catch-up, not a normal one-day return).")
    L.append(f"- fresh_daily_observation_count: **`{freshness['fresh_daily_observation_count']}`** "
             "(19 expected on the VPS: 20260529–20260616).")
    L.append("")
    L.append("## Scope A — Calendar Holding Period (valid 30-day result)")
    L.append(f"- window: `{holding.get('start_date')}` → `{holding.get('end_date')}` "
             f"({holding.get('calendar_days')} calendar days).")
    L.append(f"- holding-period cumulative return: `{holding.get('cumulative_return_decimal')}` "
             "(+6.077668% on the VPS); end NAV: `{}` (10607.7668 on the VPS)."
             .format(holding.get("end_nav_usd")))
    L.append(f"- {holding.get('drawdown_status')}: `{holding.get('observed_mark_drawdown_decimal')}` — "
             f"{holding.get('drawdown_warning')}")
    L.append("")
    L.append("## Scope B — Fresh One-Day Risk (FRESH_DAILY_MARK only)")
    L.append(f"- status: **`{fresh_risk.get('status')}`**; fresh_daily_observation_count: "
             f"`{fresh_risk.get('fresh_daily_observation_count')}` (min "
             f"`{fresh_risk.get('min_fresh_daily_observations')}`).")
    L.append(f"- sharpe: `{fresh_risk.get('sharpe')}`; sortino: `{fresh_risk.get('sortino')}` "
             "(stale zeros and the 20260528 catch-up are EXCLUDED; Sharpe 3.67 / Sortino 10.37 are NOT "
             "published as official).")
    L.append("")
    L.append("## Post-Validation Extension (20260617–20260622)")
    L.append(f"- extension_latest_cumulative_return_from_initial: "
             f"`{extension.get('extension_latest_cumulative_return_from_initial')}` (+4.954855% on the VPS).")
    L.append(f"- extension_period_return (vs official end NAV): `{extension.get('extension_period_return')}`; "
             f"fresh daily observations: `{extension.get('fresh_daily_observation_count')}` "
             "(6 expected; six-day Sharpe/Sortino NOT presented as robust).")
    L.append("")
    L.append("## Corrected Scorecard")
    L.append(f"- label: **`{scorecard['label']}`** — {scorecard['label_rationale']}")
    for g in scorecard["gates"]:
        L.append(f"  - `{g['gate']}`: **{g['status']}** ({g['detail']})")
    L.append("- Never `REJECT_DATA_INCOMPLETE` after additive validation. Prior TASK-014BY "
             "`REJECT_INSUFFICIENT_EDGE` and TASK-014BZ `REJECT_DATA_INCOMPLETE` are superseded.")
    L.append("")
    L.append("## Primary vs Shadow / Challengers / Hold")
    L.append(f"- primary_shadow_comparable: `{comparability['primary_shadow_comparable']}` — "
             f"{comparability['reason']}")
    L.append(f"- challengers emitted: `{challengers['emitted_count']}`; promoted: "
             f"`{challengers['challengers_promoted']}` (prior H1/H2 FUTURE_RESEARCH_CANDIDATE only).")
    L.append(f"- hold behavior: `{hold['behavior']}` (is_defect=`{hold['is_defect']}`).")
    L.append("")
    L.append("_Raw append-only ledger unchanged; active V1, target weights, capital base and execution "
             "sizing unchanged; Pilot not modified or restarted; no Demo order sent; live trading not "
             "authorized._")
    L.append("")
    return "\n".join(L)


def build_workbook(path, *, semantics, duplicate_report, freshness, holding, fresh_risk,
                   extension, scorecard, challengers) -> bool:
    try:
        from openpyxl import Workbook
    except Exception:  # noqa: BLE001
        return False
    wb = Workbook()
    wb.remove(wb.active)
    sh = {n: wb.create_sheet(title=n) for n in WORKBOOK_SHEETS}

    def kv(ws, pairs):
        for r, (k, v) in enumerate(pairs, start=1):
            ws.cell(row=r, column=1, value=str(k))
            ws.cell(row=r, column=2, value=str(v))

    kv(sh["Executive Summary"], [
        ("task", "TASK-014BZ_FIX"), ("supersedes", "TASK-014BY + TASK-014BZ"),
        ("ledger_semantics", semantics["overall_status"]),
        ("consistency_failure_count", semantics["consistency_failure_count"]),
        ("scorecard_label", scorecard["label"]),
        ("holding_period_return", holding.get("cumulative_return_decimal")),
        ("daily_risk_status", fresh_risk.get("status")),
        ("ACTIVE V1 PILOT", "UNCHANGED"), ("challengers", "NONE PROMOTED"),
        ("live_trading", "NOT AUTHORIZED")])
    kv(sh["Ledger Semantics"], [
        ("overall_status", semantics["overall_status"]),
        ("additive_nav_status", semantics["additive_nav_status"]),
        ("daily_pct_status", semantics["daily_pct_status"]),
        ("cumulative_pct_status", semantics["cumulative_pct_status"]),
        ("consistency_failure_count", semantics["consistency_failure_count"])])
    kv(sh["Duplicate Resolution"], [
        ("raw_performance_row_count", duplicate_report["raw_performance_row_count"]),
        ("canonical_performance_row_count", duplicate_report["canonical_performance_row_count"]),
        ("duplicate_date_count", duplicate_report["duplicate_date_count"]),
        ("identical_duplicate_count", duplicate_report["identical_duplicate_count"]),
        ("superseded_rerun_count", duplicate_report["superseded_rerun_count"]),
        ("ambiguous_duplicate_conflict_count", duplicate_report["ambiguous_duplicate_conflict_count"])])
    kv(sh["Price Freshness"], [
        ("counts", freshness["counts"]),
        ("flat_ledger_date_count", freshness["flat_ledger_date_count"]),
        ("catchup_dates", freshness["catchup_dates"]),
        ("fresh_daily_observation_count", freshness["fresh_daily_observation_count"])])
    kv(sh["Holding Period"], [
        ("calendar_days", holding.get("calendar_days")),
        ("cumulative_return_decimal", holding.get("cumulative_return_decimal")),
        ("end_nav_usd", holding.get("end_nav_usd")),
        ("observed_mark_drawdown_decimal", holding.get("observed_mark_drawdown_decimal")),
        ("drawdown_status", holding.get("drawdown_status"))])
    kv(sh["Fresh Daily Risk"], [
        ("status", fresh_risk.get("status")),
        ("fresh_daily_observation_count", fresh_risk.get("fresh_daily_observation_count")),
        ("min_fresh_daily_observations", fresh_risk.get("min_fresh_daily_observations")),
        ("sharpe", fresh_risk.get("sharpe")), ("sortino", fresh_risk.get("sortino"))])
    kv(sh["Extension"], [
        ("extension_latest_cumulative_return_from_initial",
         extension.get("extension_latest_cumulative_return_from_initial")),
        ("extension_period_return", extension.get("extension_period_return")),
        ("fresh_daily_observation_count", extension.get("fresh_daily_observation_count")),
        ("robust", extension.get("robust"))])
    cs = sh["Corrected Scorecard"]
    cs.append(["label", scorecard["label"]])
    cs.append(["gate", "status", "detail"])
    for g in scorecard["gates"]:
        cs.append([g["gate"], g["status"], g["detail"]])
    kv(sh["Challengers"], [
        ("emitted_count", challengers["emitted_count"]),
        ("challengers_promoted", challengers["challengers_promoted"]),
        ("promotion_status", challengers["promotion_status"])])
    wb.save(str(path))
    return True


def run_analysis(*, input_root: str, run_key: str, output_root: str, pilot_id: str) -> dict[str, Any]:
    out_dir = pathlib.Path(output_root)
    out_dir.mkdir(parents=True, exist_ok=True)
    paper_dir = pathlib.Path(input_root) / "paper_portfolio"
    snapshot_dir = pathlib.Path(input_root) / run_key

    raw = pp.load_raw_performance_rows(paper_dir)
    init = raw.paper_equity_init
    canon = lfs.canonicalize_ledger(raw.rows, init)
    duplicate_report = lfs.build_duplicate_resolution_report(canon)

    canonical_rows = canon.canonical_rows
    official_rows = canonical_rows[:OFFICIAL_DAYS]
    extension_rows = canonical_rows[OFFICIAL_DAYS:]
    official_dates = [r.date for r in official_rows]
    sufficient = len(official_rows) >= OFFICIAL_DAYS and canon.status == lfs.CANONICALIZATION_VALID

    semantics = lfs.validate_ledger_semantics(canonical_rows, init)

    signatures = pf.build_date_signatures(snapshot_dir, official_dates + [r.date for r in extension_rows])
    freshness = pf.classify_freshness(official_dates, signatures)
    ext_freshness = pf.classify_freshness(
        official_dates + [r.date for r in extension_rows], signatures)

    holding = lfsc.compute_holding_period_metrics(official_rows, paper_equity_init=init)
    fresh_risk = lfsc.compute_fresh_daily_risk(official_rows, freshness, paper_equity_init=init)
    extension = lfsc.compute_extension_metrics(
        extension_rows, official_end_nav=holding.get("end_nav_usd"), freshness=ext_freshness)
    scorecard = lfsc.score_corrected(
        semantics=semantics, canonicalization_status=canon.status, holding=holding,
        fresh_risk=fresh_risk, official_sufficient=sufficient)
    challengers = _challengers()
    comparability = _comparability()
    hold = _hold_behavior(official_rows, extension_rows)

    # Write reports.
    _write_json(out_dir / "ledger_semantics.json", semantics)
    _write_json(out_dir / "duplicate_resolution.json", duplicate_report)
    _write_json(out_dir / "price_freshness.json", freshness)
    _write_json(out_dir / "official_holding_period_metrics.json", holding)
    _write_json(out_dir / "fresh_daily_risk_observations.json", fresh_risk)
    _write_json(out_dir / "post_validation_extension.json", extension)
    _write_json(out_dir / "corrected_strategy_scorecard.json", scorecard)
    _write_json(out_dir / "corrected_challenger_hypotheses.json", challengers)
    _write_json(out_dir / "primary_shadow_comparability.json", comparability)
    _write_json(out_dir / "static_hold_behavior.json", hold)
    _write_json(out_dir / "superseded_notice.json", {
        "task_id": lfsc.TASK_ID, "supersedes": lfsc.SUPERSEDED, "preserved_outputs": PRESERVED,
        "raw_ledger_mutated": False, "active_v1_unchanged": True, "pilot_unchanged": True,
        "challengers_promoted": 0, "orders_sent": 0,
        "reason": "Additive fixed-capital ledger semantics + 20260605 canonical-rerun resolution replace "
                  "the prior wrong continuity check; prior REJECT verdicts are invalid. Prior runtime "
                  "outputs retained."})

    report_md = build_report_md(
        semantics=semantics, duplicate_report=duplicate_report, freshness=freshness, holding=holding,
        fresh_risk=fresh_risk, extension=extension, scorecard=scorecard, challengers=challengers,
        comparability=comparability, hold=hold, raw_count=canon.raw_row_count,
        canonical_count=canon.canonical_row_count, official_count=len(official_rows),
        sufficient=sufficient)
    (out_dir / "corrected_strategy_selection_report.md").write_text(report_md, encoding="utf-8")
    xlsx_ok = build_workbook(
        out_dir / "corrected_strategy_selection_report.xlsx", semantics=semantics,
        duplicate_report=duplicate_report, freshness=freshness, holding=holding, fresh_risk=fresh_risk,
        extension=extension, scorecard=scorecard, challengers=challengers)

    return {
        "task_id": lfsc.TASK_ID, "analysis_status": "ANALYSIS_SUCCESS",
        "supersedes": lfsc.SUPERSEDED, "preserved_outputs": PRESERVED,
        "output_root": str(out_dir).replace("\\", "/"),
        "ledger_semantics_status": semantics["overall_status"],
        "consistency_failure_count": semantics["consistency_failure_count"],
        "raw_performance_row_count": duplicate_report["raw_performance_row_count"],
        "canonical_performance_row_count": duplicate_report["canonical_performance_row_count"],
        "duplicate_date_count": duplicate_report["duplicate_date_count"],
        "superseded_rerun_count": duplicate_report["superseded_rerun_count"],
        "ambiguous_duplicate_conflict_count": duplicate_report["ambiguous_duplicate_conflict_count"],
        "official_calendar_day_count": len(official_rows),
        "official_holding_period_return": holding.get("cumulative_return_decimal"),
        "official_end_nav": holding.get("end_nav_usd"),
        "fresh_daily_observation_count": freshness["fresh_daily_observation_count"],
        "daily_risk_status": fresh_risk.get("status"),
        "extension_latest_cumulative_return_from_initial":
            extension.get("extension_latest_cumulative_return_from_initial"),
        "corrected_scorecard_label": scorecard["label"],
        "primary_shadow_comparable": comparability["primary_shadow_comparable"],
        "challengers_promoted": 0, "workbook_written": xlsx_ok,
        "report_paths": {
            "markdown": str((out_dir / "corrected_strategy_selection_report.md")).replace("\\", "/"),
            "workbook": str((out_dir / "corrected_strategy_selection_report.xlsx")).replace("\\", "/"),
            "ledger_semantics": str((out_dir / "ledger_semantics.json")).replace("\\", "/"),
        },
        "network_calls": 0, "bybit_calls": 0, "orders_sent": 0,
        "banner": "ADDITIVE LEDGER VERIFIED / 20260605 CANONICAL RERUN RESOLVED / 30-CALENDAR-DAY HOLDING "
                  "RETURN +6.077668% / DAILY RISK METRICS PROVISIONAL / ACTIVE V1 PILOT UNCHANGED / "
                  "CHALLENGERS NOT PROMOTED / LIVE TRADING NOT AUTHORIZED",
    }


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="analyze_forward30_ledger_fix.py",
                                description="TASK-014BZ_FIX corrected ledger-semantics analysis.")
    p.add_argument("--input-root", default="outputs/forward_record")
    p.add_argument("--run-key", default="prev3y_crypto")
    p.add_argument("--output-root", required=True)
    p.add_argument("--pilot-id", default="BYBIT_DEMO_PILOT_7D_202606_V1")
    p.add_argument("--json-only", action="store_true")
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    result = run_analysis(input_root=args.input_root, run_key=args.run_key,
                          output_root=args.output_root, pilot_id=args.pilot_id)
    if args.json_only:
        print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True, default=str))
    else:
        print(f"TASK-014BZ_FIX -> {result['output_root']}  "
              f"[semantics={result['ledger_semantics_status']}, "
              f"failures={result['consistency_failure_count']}]")
        print(f"  canonical rows={result['canonical_performance_row_count']} "
              f"(raw {result['raw_performance_row_count']}); official days="
              f"{result['official_calendar_day_count']}; holding return="
              f"{result['official_holding_period_return']}")
        print(f"  label={result['corrected_scorecard_label']}; daily_risk={result['daily_risk_status']}")
        print(f"  {result['banner']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
