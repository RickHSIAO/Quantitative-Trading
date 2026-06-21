"""TASK-014BQ -- build a real .xlsx workbook for the demo strategy pilot.

Strictly OFFLINE. Uses ``openpyxl`` (never LibreOffice). Reads the append-only
pilot store and produces a real workbook:

    outputs/demo_trading/pilot/<pilot_id>/demo_strategy_pilot_results.xlsx
    outputs/demo_trading/pilot/<pilot_id>/snapshots/demo_strategy_pilot_results_<YYYYMMDD>.xlsx

Sheets (deterministic order): Pilot Summary, Daily Performance, Trades,
Execution Quality, Forward Comparison, Audit Log.

No network, no orders, no scheduler, no secrets.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any, Mapping, Sequence

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from src.demo_strategy_pilot_store import PilotStore  # noqa: E402


WORKBOOK_FILENAME = "demo_strategy_pilot_results.xlsx"
SHEET_ORDER = [
    "Pilot Summary",
    "Daily Performance",
    "Trades",
    "Execution Quality",
    "Forward Comparison",
    "Audit Log",
]

MONEY_FMT = "#,##0.00######"
PCT_FMT = "0.00%"
QTY_FMT = "0.########"

# column spec: (header, source_key, type)   type in {text,int,num,money,pct,bool,list}
DAILY_COLUMNS = [
    ("Date", "date", "text"),
    ("Pilot Day", "pilot_day", "int"),
    ("Runner Status", "runner_status", "text"),
    ("Signal Count", "signal_count", "int"),
    ("Order Count", "order_count", "int"),
    ("Filled Count", "filled_count", "int"),
    ("Closed Trade Count", "closed_trade_count", "int"),
    ("Realized PnL USDT", "realized_pnl_usdt", "money"),
    ("Trading Fees USDT", "trading_fees_usdt", "money"),
    ("Funding PnL USDT", "funding_pnl_usdt", "money"),
    ("Daily Net PnL USDT", "daily_net_pnl_usdt", "money"),
    ("Cumulative Net PnL USDT", "cumulative_net_pnl_usdt", "money"),
    ("Daily Return %", "daily_return_pct", "pct"),
    ("Cumulative Return %", "cumulative_return_pct", "pct"),
    ("Max Drawdown %", "max_drawdown_pct", "pct"),
    ("Position Symbol", "current_position_symbol", "text"),
    ("Position Side", "current_position_side", "text"),
    ("Position Qty", "current_position_qty", "num"),
    ("Notion Sync", "notion_sync_status", "text"),
    ("Excel Export", "excel_export_status", "text"),
    ("Discord Notify", "discord_notify_status", "text"),
    ("Alerts Triggered", "alerts_triggered", "list"),
    ("Notes", "notes", "text"),
]

TRADE_COLUMNS = [
    ("Pilot ID", "pilot_id", "text"),
    ("Trade ID", "trade_id", "text"),
    ("Signal ID", "signal_id", "text"),
    ("Symbol", "symbol", "text"),
    ("Side", "side", "text"),
    ("Entry Order ID", "entry_order_id", "text"),
    ("Exit Order ID", "exit_order_id", "text"),
    ("Entry orderLinkId", "entry_order_link_id", "text"),
    ("Exit orderLinkId", "exit_order_link_id", "text"),
    ("Entry Time UTC", "entry_time_utc", "text"),
    ("Exit Time UTC", "exit_time_utc", "text"),
    ("Requested Qty", "requested_qty", "num"),
    ("Executed Qty", "executed_qty", "num"),
    ("Entry Price", "entry_price", "money"),
    ("Exit Price", "exit_price", "money"),
    ("Entry Fee", "entry_fee", "money"),
    ("Exit Fee", "exit_fee", "money"),
    ("Funding PnL", "funding_pnl", "money"),
    ("Gross PnL", "gross_pnl", "money"),
    ("Net PnL", "net_pnl", "money"),
    ("Slippage Entry bps", "slippage_entry_bps", "num"),
    ("Slippage Exit bps", "slippage_exit_bps", "num"),
    ("Final Status", "final_status", "text"),
    ("Included In Performance", "included_in_performance", "bool"),
]

EXECUTION_QUALITY_COLUMNS = [
    ("Trade ID", "trade_id", "text"),
    ("Symbol", "symbol", "text"),
    ("Side", "side", "text"),
    ("Requested Qty", "requested_qty", "num"),
    ("Executed Qty", "executed_qty", "num"),
    ("Entry Price", "entry_price", "money"),
    ("Exit Price", "exit_price", "money"),
    ("Slippage Entry bps", "slippage_entry_bps", "num"),
    ("Slippage Exit bps", "slippage_exit_bps", "num"),
    ("Entry Fee", "entry_fee", "money"),
    ("Exit Fee", "exit_fee", "money"),
    ("Net PnL", "net_pnl", "money"),
]

AUDIT_COLUMNS = [
    ("Timestamp UTC", "timestamp_utc", "text"),
    ("Pilot ID", "pilot_id", "text"),
    ("Event Type", "event_type", "text"),
    ("Component", "component", "text"),
    ("Status", "status", "text"),
    ("Message", "message", "text"),
    ("Reference ID", "reference_id", "text"),
]

FORWARD_COMPARISON_DEFAULT_COLUMNS = ["Metric", "Pilot Value", "Forward Value", "Delta"]


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _set_cell(ws, row: int, col: int, value: Any, ctype: str) -> None:
    cell = ws.cell(row=row, column=col)
    if ctype in ("money", "num"):
        d = _to_decimal(value)
        if d is not None:
            cell.value = float(d)
            cell.number_format = MONEY_FMT if ctype == "money" else QTY_FMT
        return
    if ctype == "pct":
        d = _to_decimal(value)
        if d is not None:
            # Stored as a percentage value (e.g. 1.23 == 1.23%); Excel percent
            # cells use the fraction, so divide by 100.
            cell.value = float(d) / 100.0
            cell.number_format = PCT_FMT
        return
    if ctype == "int":
        try:
            cell.value = int(value)
        except (TypeError, ValueError):
            cell.value = 0
        return
    if ctype == "bool":
        cell.value = bool(value)
        return
    if ctype == "list":
        seq = value if isinstance(value, (list, tuple)) else ([] if value in (None, "") else [value])
        cell.value = ", ".join(str(x) for x in seq)
        return
    cell.value = "" if value is None else str(value)


def _write_table(ws, columns: Sequence[tuple[str, str, str]], rows: Sequence[Mapping[str, Any]]) -> None:
    headers = [c[0] for c in columns]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(j)].width = max(12, min(40, len(h) + 4))
    for i, row in enumerate(rows, start=2):
        for j, (_, key, ctype) in enumerate(columns, start=1):
            _set_cell(ws, i, j, row.get(key), ctype)
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    last_row = max(1, len(rows) + 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"


def _write_summary(ws, store: PilotStore, daily, trades, audit) -> None:
    ws.cell(row=1, column=1, value="Field").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Value").font = Font(bold=True)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 48
    config = store.read_config() or {}
    summary = store.read_latest_summary() or {}
    pairs: list[tuple[str, Any]] = [
        ("Pilot ID", store.pilot_id),
        ("Strategy Name", config.get("strategy_name", "")),
        ("Environment", config.get("environment", "")),
        ("Start Date", config.get("start_date", "")),
        ("Maximum Calendar Days", config.get("maximum_calendar_days", "")),
        ("Minimum Closed Trades", config.get("minimum_closed_trades", "")),
        ("Initial Equity USDT", config.get("initial_equity_usdt", "")),
        ("Total Daily Records", len(daily)),
        ("Total Trades", len(trades)),
        ("Total Audit Events", len(audit)),
    ]
    for k in sorted(summary.keys()):
        pairs.append((f"summary.{k}", summary[k]))
    for i, (k, v) in enumerate(pairs, start=2):
        ws.cell(row=i, column=1, value=str(k))
        ws.cell(row=i, column=2, value="" if v is None else (v if isinstance(v, (int, float)) else str(v)))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{max(1, len(pairs) + 1)}"


def _write_forward_comparison(ws, comparison: Any) -> None:
    rows: list[Mapping[str, Any]] = []
    columns = list(FORWARD_COMPARISON_DEFAULT_COLUMNS)
    if isinstance(comparison, Mapping) and comparison.get("columns"):
        columns = [str(c) for c in comparison["columns"]]
        rows = list(comparison.get("rows", []))
    elif isinstance(comparison, list) and comparison:
        keys: list[str] = []
        for r in comparison:
            for k in r:
                if k not in keys:
                    keys.append(k)
        columns = keys
        rows = comparison
    for j, h in enumerate(columns, start=1):
        c = ws.cell(row=1, column=j, value=str(h))
        c.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(j)].width = 22
    for i, row in enumerate(rows, start=2):
        for j, key in enumerate(columns, start=1):
            val = row.get(key) if isinstance(row, Mapping) else None
            ws.cell(row=i, column=j, value="" if val is None else val)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(rows) + 1)}"


def build_workbook(
    pilot_id: str,
    output_root: str | os.PathLike[str] | None = None,
    *,
    comparison: Any = None,
    snapshot_date: str | None = None,
) -> dict[str, str]:
    """Build the pilot workbook (atomic) plus a dated snapshot. Returns paths."""
    store = PilotStore(pilot_id, output_root)
    daily = store.read_daily()
    trades = store.read_trades()
    audit = store.read_audit()

    wb = Workbook()
    # Remove default sheet; create in deterministic order.
    wb.remove(wb.active)
    sheets = {name: wb.create_sheet(title=name) for name in SHEET_ORDER}

    _write_summary(sheets["Pilot Summary"], store, daily, trades, audit)
    _write_table(sheets["Daily Performance"], DAILY_COLUMNS, daily)
    _write_table(sheets["Trades"], TRADE_COLUMNS, trades)
    _write_table(sheets["Execution Quality"], EXECUTION_QUALITY_COLUMNS, trades)
    _write_forward_comparison(sheets["Forward Comparison"], comparison)
    _write_table(sheets["Audit Log"], AUDIT_COLUMNS, audit)

    store._ensure_dir()
    final_path = store.dir / WORKBOOK_FILENAME
    tmp_path = store.dir / (WORKBOOK_FILENAME + ".tmp")
    wb.save(str(tmp_path))
    os.replace(tmp_path, final_path)

    stamp = snapshot_date or datetime.now(timezone.utc).strftime("%Y%m%d")
    snap_dir = store.dir / "snapshots"
    snap_dir.mkdir(parents=True, exist_ok=True)
    snap_path = snap_dir / f"demo_strategy_pilot_results_{stamp}.xlsx"
    wb.save(str(snap_path))

    return {"workbook": str(final_path), "snapshot": str(snap_path)}


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(prog="build_demo_strategy_pilot_workbook.py",
                                description="Build the demo strategy pilot .xlsx workbook (offline).")
    p.add_argument("--pilot-id", required=True)
    p.add_argument("--output-root", default=None,
                   help="optional output root (defaults to canonical outputs/demo_trading/pilot)")
    p.add_argument("--comparison-json", default=None,
                   help="optional sanitized Forward Comparison fixture JSON path")
    p.add_argument("--snapshot-date", default=None, help="override snapshot YYYYMMDD (testing)")
    args = p.parse_args(argv)

    comparison = None
    if args.comparison_json:
        with open(args.comparison_json, "r", encoding="utf-8") as fh:
            comparison = json.load(fh)

    paths = build_workbook(args.pilot_id, args.output_root, comparison=comparison,
                           snapshot_date=args.snapshot_date)
    print(json.dumps(paths, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
