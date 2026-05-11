from __future__ import annotations

import json
import threading
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from src.database import get_connection


_WRITE_LOCK = threading.Lock()
TABLE_NAME = "bybit_live_orders"

EXCEL_COLUMNS = [
    "id",
    "recorded_at",
    "environment",
    "action",
    "symbol",
    "bybit_symbol",
    "side",
    "direction",
    "quantity",
    "price",
    "stop_loss",
    "take_profit",
    "strategy",
    "score",
    "signal_date",
    "reason",
    "pnl",
    "fee",
    "balance_usdt",
    "order_id",
    "order_link_id",
    "ret_code",
    "ret_msg",
]

EXCEL_COLUMN_NAMES = {
    "id": "ID",
    "recorded_at": "記錄時間",
    "environment": "環境",
    "action": "動作",
    "symbol": "交易標的",
    "bybit_symbol": "Bybit代號",
    "side": "買賣",
    "direction": "方向",
    "quantity": "數量",
    "price": "成交價",
    "stop_loss": "停損價",
    "take_profit": "停利價",
    "strategy": "策略",
    "score": "分數",
    "signal_date": "訊號日期",
    "reason": "原因",
    "pnl": "損益(USDT)",
    "fee": "手續費(USDT)",
    "balance_usdt": "帳戶餘額(USDT)",
    "order_id": "Bybit訂單ID",
    "order_link_id": "自訂訂單ID",
    "ret_code": "回傳碼",
    "ret_msg": "回傳訊息",
}

ACTION_LABELS = {
    "ENTRY": "進場",
    "EXIT": "出場",
}

SIDE_LABELS = {
    "Buy": "買入",
    "Sell": "賣出",
}

DIRECTION_LABELS = {
    1: "做多",
    -1: "做空",
}

ENVIRONMENT_LABELS = {
    "demo": "Demo模擬",
    "testnet": "測試網",
    "live": "正式實盤",
}

STRATEGY_LABELS = {
    "bb": "布林均值回歸",
    "vp": "成交量剖面",
    "trend": "趨勢策略",
    "combined": "綜合策略",
    "unknown": "未知",
}

REASON_LABELS = {
    "SL": "停損",
    "TP": "停利",
    "FLIP": "訊號反轉",
    "SOFT": "軟停損",
    "MAXHOLD": "最長持倉",
    "BB-TGT": "布林獲利目標",
    "BB-MID": "布林中線出場",
    "BB-RSI": "RSI出場",
    "REMOTE_CLOSED": "交易所端已平倉",
    "REMOTE_CLOSED_SL": "交易所端止損平倉",
}

RET_MSG_LABELS = {
    "OK": "成功",
    "success": "成功",
    "remote position closed": "交易所端已平倉",
    "backfilled from Bybit execution": "由Bybit成交紀錄補登",
    "backfilled from Bybit execution and closed PnL": "由Bybit成交與已實現損益補登",
}

DAY_ROW_FILLS = (
    PatternFill("solid", fgColor="EAF4FF"),
    PatternFill("solid", fgColor="FFF4D6"),
)


def _now_local() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _json_dumps(value: Any) -> str:
    try:
        return json.dumps(value or {}, ensure_ascii=False, sort_keys=True)
    except TypeError:
        return json.dumps({"repr": repr(value)}, ensure_ascii=False, sort_keys=True)


def _symbol_to_bybit(symbol: str) -> str:
    symbol = str(symbol or "").strip()
    if symbol.startswith("BYBIT:") and symbol.endswith(".P"):
        return symbol[6:-2]
    return symbol.replace("-USD", "USDT")


def _environment_label() -> str:
    if getattr(config, "BYBIT_TESTNET", False):
        return "testnet"
    if getattr(config, "BYBIT_DEMO", False):
        return "demo"
    return "live"


def _response_result(response: dict[str, Any] | None) -> dict[str, Any]:
    if not isinstance(response, dict):
        return {}
    result = response.get("result") or {}
    return result if isinstance(result, dict) else {}


def _response_value(response: dict[str, Any] | None, key: str) -> Any:
    if not isinstance(response, dict):
        return None
    result = _response_result(response)
    return result.get(key) or response.get(key)


def _side_for(action: str, direction: int | None) -> str:
    if direction not in (1, -1):
        return ""
    normalized = str(action or "").upper()
    if normalized == "ENTRY":
        return "Buy" if direction == 1 else "Sell"
    if normalized == "EXIT":
        return "Sell" if direction == 1 else "Buy"
    return ""


def _excel_path(path: str | None = None) -> Path:
    if path:
        return Path(path)
    configured = getattr(config, "BYBIT_LIVE_ORDER_XLSX", None)
    if configured:
        return Path(configured)
    return Path(getattr(config, "OUTPUT_DIR", "output")) / "Bybit_Live_Orders.xlsx"


def _map_direction(value: Any) -> str:
    try:
        return DIRECTION_LABELS.get(int(value), str(value or ""))
    except (TypeError, ValueError):
        return str(value or "")


def _map_reason(value: Any) -> str:
    text = str(value or "")
    return REASON_LABELS.get(text, text)


def _map_strategy(value: Any) -> str:
    text = str(value or "")
    return STRATEGY_LABELS.get(text, text)


def _map_ret_msg(value: Any) -> str:
    text = str(value or "")
    return RET_MSG_LABELS.get(text, text)


def _prepare_excel_display(df: pd.DataFrame) -> pd.DataFrame:
    display_cols = [c for c in EXCEL_COLUMNS if c in df.columns]
    display = df[display_cols].copy()
    if display.empty:
        return display.rename(columns=EXCEL_COLUMN_NAMES)

    if "environment" in display:
        display["environment"] = display["environment"].map(
            lambda v: ENVIRONMENT_LABELS.get(str(v or ""), str(v or ""))
        )
    if "action" in display:
        display["action"] = display["action"].map(
            lambda v: ACTION_LABELS.get(str(v or "").upper(), str(v or ""))
        )
    if "side" in display:
        display["side"] = display["side"].map(
            lambda v: SIDE_LABELS.get(str(v or ""), str(v or ""))
        )
    if "direction" in display:
        display["direction"] = display["direction"].map(_map_direction)
    if "strategy" in display:
        display["strategy"] = display["strategy"].map(_map_strategy)
    if "reason" in display:
        display["reason"] = display["reason"].map(_map_reason)
    if "ret_msg" in display:
        display["ret_msg"] = display["ret_msg"].map(_map_ret_msg)

    return display.rename(columns=EXCEL_COLUMN_NAMES)


def _recorded_date_key(value: Any) -> str:
    if value in (None, ""):
        return ""
    try:
        return pd.Timestamp(value).date().isoformat()
    except Exception:
        text = str(value)
        return text[:10] if len(text) >= 10 else text


def _recorded_date_keys(df: pd.DataFrame) -> list[str]:
    if "recorded_at" not in df.columns or df.empty:
        return []
    return [_recorded_date_key(value) for value in df["recorded_at"].tolist()]


def _apply_daily_row_fills(ws, date_keys: list[str]) -> None:
    date_to_fill_idx: dict[str, int] = {}
    for row_idx, date_key in enumerate(date_keys, start=2):
        if row_idx > ws.max_row:
            break
        if date_key not in date_to_fill_idx:
            date_to_fill_idx[date_key] = len(date_to_fill_idx) % len(DAY_ROW_FILLS)
        fill = DAY_ROW_FILLS[date_to_fill_idx[date_key]]
        for cell in ws[row_idx]:
            cell.fill = fill


def ensure_bybit_live_order_ledger() -> None:
    conn = get_connection()
    with _WRITE_LOCK, conn:
        conn.executescript(
            f"""
            CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                recorded_at   TEXT NOT NULL,
                environment   TEXT,
                action        TEXT NOT NULL,
                symbol        TEXT NOT NULL,
                bybit_symbol  TEXT,
                side          TEXT,
                direction     INTEGER,
                quantity      REAL,
                price         REAL,
                stop_loss     REAL,
                take_profit   REAL,
                strategy      TEXT,
                score         INTEGER,
                signal_date   TEXT,
                reason        TEXT,
                pnl           REAL,
                fee           REAL,
                balance_usdt  REAL,
                order_id      TEXT,
                order_link_id TEXT,
                ret_code      INTEGER,
                ret_msg       TEXT,
                raw_response  TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_bybit_live_orders_recorded_at
                ON {TABLE_NAME}(recorded_at);
            CREATE INDEX IF NOT EXISTS idx_bybit_live_orders_symbol
                ON {TABLE_NAME}(symbol);
            CREATE INDEX IF NOT EXISTS idx_bybit_live_orders_order_id
                ON {TABLE_NAME}(order_id);
            """
        )


def load_bybit_live_orders(limit: int | None = None) -> pd.DataFrame:
    ensure_bybit_live_order_ledger()
    sql = f"SELECT * FROM {TABLE_NAME} ORDER BY recorded_at ASC, id ASC"
    params: tuple[Any, ...] = ()
    if limit is not None:
        sql = f"SELECT * FROM {TABLE_NAME} ORDER BY recorded_at DESC, id DESC LIMIT ?"
        params = (int(limit),)
    with get_connection() as conn:
        df = pd.read_sql_query(sql, conn, params=params)
    if limit is not None and not df.empty:
        df = df.sort_values("id").reset_index(drop=True)
    return df


def export_bybit_live_orders_to_excel(path: str | None = None) -> str:
    ensure_bybit_live_order_ledger()
    output_path = _excel_path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df = load_bybit_live_orders()
    if df.empty:
        df = pd.DataFrame(columns=EXCEL_COLUMNS)

    date_keys = _recorded_date_keys(df)
    display = _prepare_excel_display(df)

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            orders_sheet = "Bybit進出場"
            summary_sheet = "彙總"
            display.to_excel(writer, sheet_name=orders_sheet, index=False)
            wb = writer.book
            ws = writer.sheets[orders_sheet]
            header_fill = PatternFill("solid", fgColor="1F3864")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            _apply_daily_row_fills(ws, date_keys)
            ws.freeze_panes = "A2"
            if ws.max_column and ws.max_row:
                ws.auto_filter.ref = (
                    f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                )
            for col in ws.columns:
                width = max(len(str(cell.value or "")) for cell in col) + 2
                ws.column_dimensions[col[0].column_letter].width = min(max(width, 10), 42)

            if not display.empty:
                summary = (
                    display.groupby(["交易標的", "動作"], dropna=False)
                    .agg({
                        "ID": "count",
                        "數量": "sum",
                        "損益(USDT)": "sum",
                        "手續費(USDT)": "sum",
                    })
                    .reset_index()
                    .rename(columns={
                        "ID": "筆數",
                        "損益(USDT)": "損益合計(USDT)",
                        "手續費(USDT)": "手續費合計(USDT)",
                    })
                )
            else:
                summary = pd.DataFrame(
                    columns=[
                        "交易標的",
                        "動作",
                        "筆數",
                        "數量",
                        "損益合計(USDT)",
                        "手續費合計(USDT)",
                    ]
                )
            summary.to_excel(writer, sheet_name=summary_sheet, index=False)
            ws_sum = writer.sheets[summary_sheet]
            for cell in ws_sum[1]:
                cell.fill = header_fill
                cell.font = header_font
            ws_sum.freeze_panes = "A2"
            for col in ws_sum.columns:
                width = max(len(str(cell.value or "")) for cell in col) + 2
                ws_sum.column_dimensions[col[0].column_letter].width = min(max(width, 10), 32)
    except PermissionError:
        print(f"[WARN] Cannot update live order Excel; close it first: {output_path}")

    return str(output_path)


def record_bybit_order(
    *,
    action: str,
    symbol: str,
    direction: int | None = None,
    quantity: Any = None,
    price: Any = None,
    stop_loss: Any = None,
    take_profit: Any = None,
    strategy: str = "",
    score: int | None = None,
    signal_date: str = "",
    reason: str = "",
    response: dict[str, Any] | None = None,
    pnl: Any = None,
    fee: Any = None,
    balance_usdt: Any = None,
    environment: str | None = None,
    recorded_at: str | None = None,
    export_excel: bool = True,
) -> int:
    ensure_bybit_live_order_ledger()

    direction_i = _to_int(direction)
    response = response or {}
    row = {
        "recorded_at": recorded_at or _now_local(),
        "environment": environment or _environment_label(),
        "action": str(action or "").upper(),
        "symbol": str(symbol or ""),
        "bybit_symbol": _symbol_to_bybit(symbol),
        "side": _side_for(action, direction_i),
        "direction": direction_i,
        "quantity": _to_float(quantity),
        "price": _to_float(price),
        "stop_loss": _to_float(stop_loss),
        "take_profit": _to_float(take_profit),
        "strategy": strategy or "",
        "score": _to_int(score),
        "signal_date": signal_date or "",
        "reason": reason or "",
        "pnl": _to_float(pnl),
        "fee": _to_float(fee),
        "balance_usdt": _to_float(balance_usdt),
        "order_id": _response_value(response, "orderId") or "",
        "order_link_id": _response_value(response, "orderLinkId") or "",
        "ret_code": _to_int(response.get("retCode")) if isinstance(response, dict) else None,
        "ret_msg": str(response.get("retMsg", "")) if isinstance(response, dict) else "",
        "raw_response": _json_dumps(response),
    }

    cols = list(row.keys())
    placeholders = ",".join("?" for _ in cols)
    sql = (
        f"INSERT INTO {TABLE_NAME} ({','.join(cols)}) "
        f"VALUES ({placeholders})"
    )
    conn = get_connection()
    with _WRITE_LOCK, conn:
        cur = conn.execute(sql, tuple(row[c] for c in cols))
        row_id = int(cur.lastrowid)

    if export_excel:
        export_bybit_live_orders_to_excel()
    return row_id
