"""
Excel 回測報告產生器。

工作表結構：
  📊 Summary          — 完整績效摘要 + 資金曲線圖
  📈 Monthly P&L      — 月度損益矩陣（資產類型 × 月份，熱力圖配色）
  🔍 Strategy Stats   — 三大策略並排比較
  📋 Asset Stats      — 各資產類型統計
  YYYY-QN  (×20)      — 各季交易明細（進出場點位、原因、盈虧、R倍數）
  📋 All Trades       — 全部交易彙總
"""
from __future__ import annotations
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
import config
from src.backtester import Trade

# ─── 樣式 ────────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill('solid', fgColor='1F3864')
HDR_FONT   = Font(color='FFFFFF', bold=True, size=10)
SUB_FILL   = PatternFill('solid', fgColor='2E75B6')
SUB_FONT   = Font(color='FFFFFF', bold=True, size=10)
WIN_FILL   = PatternFill('solid', fgColor='C6EFCE')
LOSS_FILL  = PatternFill('solid', fgColor='FFC7CE')
NEUT_FILL  = PatternFill('solid', fgColor='FFEB9C')
SECT_FILL  = PatternFill('solid', fgColor='D9E1F2')
TITLE_FONT = Font(bold=True, size=14, color='1F3864')
BOLD       = Font(bold=True)
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center')


# ─── 工具 ────────────────────────────────────────────────────────────────────
def _style_row(ws, row: int, fill=None, font=None, align=CENTER):
    for cell in ws[row]:
        if fill:  cell.fill      = fill
        if font:  cell.font      = font
        if align: cell.alignment = align
        cell.border = THIN


def _auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        w = min_w
        for cell in col:
            try: w = max(w, len(str(cell.value or '')))
            except: pass
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, max_w)


def _add_autofilter(ws, header_row: int):
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
    )


# ─── Trade → DataFrame ────────────────────────────────────────────────────────
TRADE_COLS = [
    '進場日期', '出場日期', '持倉天數',
    '代號', '資產類型', '策略',
    '方向', '進場原因',
    '進場價格', '止損點位', '止盈點位', '出場價格',
    '出場原因',
    '風險金額(USD)', 'R倍數',
    'P&L (USD)', '報酬率(%)',
    'MAE(%)', 'MFE(%)',
]


def trades_to_df(trades: list[Trade]) -> pd.DataFrame:
    rows = []
    for t in trades:
        if t.exit_date is None:
            continue
        rows.append({
            '進場日期':    t.entry_date,
            '出場日期':    t.exit_date,
            '持倉天數':    t.holding_days or 0,
            '代號':        t.symbol,
            '資產類型':    t.asset_type or '',
            '策略':        _strat_label(t.strategy),
            '方向':        '多 ▲' if t.direction == 1 else '空 ▼',
            '進場原因':    t.entry_reason or '',
            '進場價格':    round(t.entry_price, 6),
            '止損點位':    round(t.stop_loss, 6),
            '止盈點位':    round(t.take_profit, 6),
            '出場價格':    round(t.exit_price, 6) if t.exit_price else None,
            '出場原因':    t.exit_reason or '',
            '風險金額(USD)': round(t.risk_usd, 2) if t.risk_usd else None,
            'R倍數':       round(t.r_multiple, 2) if t.r_multiple is not None else None,
            'P&L (USD)':  round(t.pnl, 2) if t.pnl is not None else None,
            '報酬率(%)':   round(t.return_pct, 2) if t.return_pct is not None else None,
            'MAE(%)':      t.mae,
            'MFE(%)':      t.mfe,
        })
    df = pd.DataFrame(rows, columns=TRADE_COLS)
    if not df.empty:
        df['進場日期'] = pd.to_datetime(df['進場日期'])
        df['出場日期'] = pd.to_datetime(df['出場日期'])
        df = df.sort_values('出場日期').reset_index(drop=True)
    return df


def _strat_label(s: str) -> str:
    return {'trend': '趨勢動能', 'vp': 'Volume Profile', 'bb': '布林回歸',
            'combined': '多策略確認'}.get(s, s)


MONTH_ZH = ['一月','二月','三月','四月','五月','六月',
            '七月','八月','九月','十月','十一月','十二月']

MONTH_HDR_FILL = PatternFill('solid', fgColor='2E75B6')
MONTH_HDR_FONT = Font(color='FFFFFF', bold=True, size=11)
MONTH_TOT_FILL = PatternFill('solid', fgColor='D9E1F2')


def _write_data_rows(ws, display: pd.DataFrame, pnl_idx: int):
    """寫入交易資料列（綠獲利 / 紅虧損），回傳最後一列列號。"""
    for _, r in display.iterrows():
        ws.append([r[c] for c in TRADE_COLS])
        cur = ws.max_row
        try:
            fill = WIN_FILL if float(ws.cell(cur, pnl_idx).value or 0) > 0 else LOSS_FILL
        except (TypeError, ValueError):
            fill = NEUT_FILL
        for c in range(1, len(TRADE_COLS) + 1):
            cell = ws.cell(cur, c)
            cell.fill      = fill
            cell.alignment = LEFT if c in (8, 13) else CENTER
            cell.border    = THIN
    return ws.max_row


def _write_month_total_row(ws, month_pnl: float, month_trades: int,
                           month_wins: int):
    """月份小計列。"""
    wr = month_wins / month_trades * 100 if month_trades else 0
    label = (f"小計：{month_trades} 筆 | 勝率 {wr:.0f}% | "
             f"損益 ${month_pnl:+,.2f}")
    ws.append([label] + [''] * (len(TRADE_COLS) - 1))
    cur = ws.max_row
    ws.merge_cells(f'A{cur}:{get_column_letter(len(TRADE_COLS))}{cur}')
    cell = ws.cell(cur, 1)
    cell.fill      = WIN_FILL if month_pnl >= 0 else LOSS_FILL
    cell.font      = Font(bold=True, size=10)
    cell.alignment = LEFT
    cell.border    = THIN


# ─── 寫入年度工作表（內含月份分隔）────────────────────────────────────────────
def _write_year_sheet(wb: Workbook, year: int, df: pd.DataFrame,
                      year_start_cap: float = 0.0):
    """
    每個年度一張工作表，內部依月份分組。
    每月開頭有藍色月份標題列，結尾有月份小計列。
    """
    ws = wb.create_sheet(title=str(year))
    n_cols = len(TRADE_COLS)

    # ── 年度標題 ──────────────────────────────────────────────────────────
    ws.cell(1, 1, f'{year} 年交易明細').font = TITLE_FONT
    ws.append([])

    # ── 年度彙總 ──────────────────────────────────────────────────────────
    pnl_data = df['P&L (USD)'].dropna()
    year_pnl = pnl_data.sum()
    year_ret  = year_pnl / year_start_cap * 100 if year_start_cap else 0.0
    ret_str   = f"年度報酬率：{year_ret:+.2f}%"

    summary_vals = [
        f"全年交易：{len(df)} 筆",
        f"獲利：{(pnl_data > 0).sum()} 筆",
        f"虧損：{(pnl_data <= 0).sum()} 筆",
        f"勝率：{(pnl_data > 0).mean()*100:.1f}%" if len(pnl_data) else '勝率：—',
        f"全年損益：${year_pnl:+,.2f}",
        ret_str,
    ]
    for i, s in enumerate(summary_vals):
        cell = ws.cell(3, 1 + i, s)
        cell.font = Font(bold=True, color='1F3864')

    # 年度報酬率用顏色強調
    ret_cell = ws.cell(3, len(summary_vals))
    ret_cell.font = Font(bold=True, color=('375623' if year_ret >= 0 else '9C0006'), size=11)
    ws.append([])

    pnl_idx = TRADE_COLS.index('P&L (USD)') + 1
    r_idx   = TRADE_COLS.index('R倍數') + 1
    first_hdr_row = None

    # ── 依月份分組 ────────────────────────────────────────────────────────
    df_w = df.copy()
    df_w['_month'] = df_w['出場日期'].dt.month

    for month_num in range(1, 13):
        m_df = df_w[df_w['_month'] == month_num].drop(columns=['_month'])
        if m_df.empty:
            continue

        # 月份標題列
        month_label = f"{MONTH_ZH[month_num-1]}（{year}/{month_num:02d}）"
        ws.append([month_label] + [''] * (n_cols - 1))
        cur = ws.max_row
        ws.merge_cells(f'A{cur}:{get_column_letter(n_cols)}{cur}')
        ws.cell(cur, 1).fill      = MONTH_HDR_FILL
        ws.cell(cur, 1).font      = MONTH_HDR_FONT
        ws.cell(cur, 1).alignment = LEFT
        ws.cell(cur, 1).border    = THIN

        # 欄位標題（第一次才固定；之後省略，靠 freeze 參照）
        ws.append(TRADE_COLS)
        hdr_row = ws.max_row
        _style_row(ws, hdr_row, fill=HDR_FILL, font=HDR_FONT)
        if first_hdr_row is None:
            first_hdr_row = hdr_row

        # 交易資料列
        display = m_df.copy()
        for col in ['進場日期', '出場日期']:
            if pd.api.types.is_datetime64_any_dtype(display[col]):
                display[col] = display[col].dt.strftime('%Y-%m-%d')
        _write_data_rows(ws, display, pnl_idx)

        # 月份小計
        m_pnl   = m_df['P&L (USD)'].dropna().sum()
        m_wins  = (m_df['P&L (USD)'] > 0).sum()
        _write_month_total_row(ws, m_pnl, len(m_df), int(m_wins))

        ws.append([])  # 空行間隔

    # R倍數條件格式
    if first_hdr_row:
        r_col_letter = get_column_letter(r_idx)
        data_rng = f"{r_col_letter}{first_hdr_row+1}:{r_col_letter}{ws.max_row}"
        ws.conditional_formatting.add(data_rng,
            CellIsRule(operator='greaterThan', formula=['0'],
                       fill=PatternFill('solid', fgColor='375623')))
        ws.conditional_formatting.add(data_rng,
            CellIsRule(operator='lessThan', formula=['0'],
                       fill=PatternFill('solid', fgColor='9C0006')))
        ws.freeze_panes = ws.cell(first_hdr_row + 1, 1)

    _auto_width(ws)
    return ws


def _write_trade_sheet(wb: Workbook, name: str, df: pd.DataFrame, title: str = ''):
    """All Trades 用的單一平面工作表（無月份分組）。"""
    ws = wb.create_sheet(title=name[:31])
    if title:
        ws.cell(1, 1, title).font = TITLE_FONT
        ws.append([])

    if df.empty:
        ws.cell(ws.max_row + 1, 1, '無交易記錄')
        return ws

    display = df.copy()
    for col in ['進場日期', '出場日期']:
        if pd.api.types.is_datetime64_any_dtype(display.get(col, pd.Series())):
            display[col] = display[col].dt.strftime('%Y-%m-%d')

    ws.append(TRADE_COLS)
    _style_row(ws, ws.max_row, fill=HDR_FILL, font=HDR_FONT)
    hdr_row = ws.max_row
    pnl_idx = TRADE_COLS.index('P&L (USD)') + 1
    r_idx   = TRADE_COLS.index('R倍數') + 1

    _write_data_rows(ws, display, pnl_idx)

    r_col_letter = get_column_letter(r_idx)
    data_rng = f"{r_col_letter}{hdr_row+1}:{r_col_letter}{ws.max_row}"
    ws.conditional_formatting.add(data_rng,
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=PatternFill('solid', fgColor='375623')))
    ws.conditional_formatting.add(data_rng,
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill('solid', fgColor='9C0006')))

    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    _add_autofilter(ws, hdr_row)
    _auto_width(ws)
    return ws


# ─── Monthly P&L Matrix ───────────────────────────────────────────────────────
def _write_monthly_pnl(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet(title='📈 Monthly P&L')
    ws.cell(1, 1, '月度損益矩陣 (USD)').font = TITLE_FONT
    ws.cell(2, 1, '▲ 綠色 = 獲利月份  ▼ 紅色 = 虧損月份  數值 = 當月合計損益').font = Font(italic=True, color='595959')

    if df.empty or 'P&L (USD)' not in df.columns:
        ws.cell(4, 1, '無交易資料')
        return ws

    df_w = df.copy()
    df_w['Month'] = df_w['出場日期'].dt.to_period('M').astype(str)

    # Pivot: 月 × 資產類型
    pivot = df_w.pivot_table(
        values='P&L (USD)', index='Month',
        columns='資產類型', aggfunc='sum', fill_value=0
    )
    pivot['合計'] = pivot.sum(axis=1)
    pivot = pivot.reset_index()

    # Header row
    header = ['月份'] + list(pivot.columns[1:])
    ws.append([])
    ws.append(header)
    hdr_row = ws.max_row
    _style_row(ws, hdr_row, fill=HDR_FILL, font=HDR_FONT)

    pnl_start_col = 2  # first numeric column (1-based)
    for _, row in pivot.iterrows():
        ws.append(list(row))
        cur = ws.max_row
        for c in range(pnl_start_col, len(header) + 1):
            cell = ws.cell(cur, c)
            try:
                val  = float(cell.value or 0)
                cell.fill      = WIN_FILL if val > 0 else (LOSS_FILL if val < 0 else NEUT_FILL)
                cell.value     = round(val, 2)
                cell.number_format = '#,##0.00'
            except (TypeError, ValueError):
                pass
            cell.alignment = CENTER
            cell.border    = THIN
        ws.cell(cur, 1).alignment = CENTER
        ws.cell(cur, 1).border    = THIN

    # 最後一行：年度合計
    ws.append(['全期合計'] + [round(pivot[c].sum(), 2) for c in pivot.columns[1:]])
    _style_row(ws, ws.max_row, fill=SECT_FILL, font=BOLD)

    # 色階規則（整個數字區塊）
    data_range = (f"B{hdr_row+1}:{get_column_letter(len(header))}{ws.max_row-1}")
    ws.conditional_formatting.add(data_range, ColorScaleRule(
        start_type='min', start_color='F8696B',
        mid_type='num',   mid_value=0,   mid_color='FFFFFF',
        end_type='max',   end_color='63BE7B',
    ))

    # 月度柱狀圖（合計）
    n_months = len(pivot)
    chart = BarChart()
    chart.type   = 'col'
    chart.title  = '月度合計損益'
    chart.height = 12
    chart.width  = 22
    chart.y_axis.title = 'P&L (USD)'

    total_col = len(header)  # last column = 合計
    data_ref  = Reference(ws, min_col=total_col, min_row=hdr_row,
                          max_row=hdr_row + n_months)
    cat_ref   = Reference(ws, min_col=1, min_row=hdr_row+1, max_row=hdr_row+n_months)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    ws.add_chart(chart, f'A{hdr_row + n_months + 4}')

    _auto_width(ws)
    return ws


# ─── Strategy Stats ────────────────────────────────────────────────────────────
def _write_strategy_stats(wb: Workbook, df: pd.DataFrame, metrics: dict):
    ws = wb.create_sheet(title='🔍 Strategy Stats')
    ws.cell(1, 1, '策略效益比較').font = TITLE_FONT

    by_strat = metrics.get('by_strategy', {})
    strat_order = [('trend', '趨勢動能 (Supertrend+EMA200)'),
                   ('vp',    'Volume Profile (POC支撐阻力)'),
                   ('bb',    '布林均值回歸 (BB+RSI+ATR)')]

    header = ['策略', '交易次數', '勝率', '平均損益(USD)',
              '平均R倍數', '合計損益(USD)', '佔總損益%']
    ws.append([])
    ws.append(header)
    _style_row(ws, ws.max_row, fill=HDR_FILL, font=HDR_FONT)

    total_pnl = metrics.get('total_pnl', 1) or 1
    for key, label in strat_order:
        stat = by_strat.get(key)
        if not stat:
            ws.append([label, 0, '—', '—', '—', '—', '—'])
            continue
        pct = stat['total_pnl'] / total_pnl * 100 if total_pnl else 0
        row = [label, stat['trades'],
               f"{stat['win_rate']*100:.1f}%",
               f"${stat['avg_pnl']:+,.2f}",
               f"{stat.get('avg_r', 0):+.2f}R",
               f"${stat['total_pnl']:+,.2f}",
               f"{pct:.1f}%"]
        ws.append(row)
        cur = ws.max_row
        fill = WIN_FILL if stat['total_pnl'] > 0 else LOSS_FILL
        for c in range(1, len(header)+1):
            ws.cell(cur, c).fill      = fill
            ws.cell(cur, c).alignment = CENTER
            ws.cell(cur, c).border    = THIN

    # 出場原因分布
    ws.append([])
    ws.append(['出場原因', '次數', '佔比'])
    _style_row(ws, ws.max_row, fill=SUB_FILL, font=SUB_FONT)
    exit_dist = metrics.get('exit_distribution', {})
    total_ex  = sum(exit_dist.values()) or 1
    for reason, cnt in exit_dist.items():
        ws.append([reason, cnt, f"{cnt/total_ex*100:.1f}%"])
        for c in range(1, 4):
            ws.cell(ws.max_row, c).alignment = CENTER
            ws.cell(ws.max_row, c).border    = THIN

    # 多空勝率
    ws.append([])
    ws.append(['方向', '勝率'])
    _style_row(ws, ws.max_row, fill=SUB_FILL, font=SUB_FONT)
    ws.append(['多單 (Long)',  f"{metrics.get('win_rate_long', 0)*100:.1f}%"])
    ws.append(['空單 (Short)', f"{metrics.get('win_rate_short', 0)*100:.1f}%"])
    for r in [ws.max_row-1, ws.max_row]:
        for c in range(1, 3):
            ws.cell(r, c).alignment = CENTER
            ws.cell(r, c).border    = THIN

    _auto_width(ws)
    return ws


# ─── Asset Stats ──────────────────────────────────────────────────────────────
def _write_asset_stats(wb: Workbook, df: pd.DataFrame, metrics: dict):
    ws = wb.create_sheet(title='📋 Asset Stats')
    ws.cell(1, 1, '資產類型統計').font = TITLE_FONT

    by_type = metrics.get('by_asset_type', {})
    header  = ['資產類型', '交易次數', '勝率', '合計損益(USD)', '佔總損益%']
    ws.append([])
    ws.append(header)
    _style_row(ws, ws.max_row, fill=HDR_FILL, font=HDR_FONT)

    total_pnl = metrics.get('total_pnl', 1) or 1
    order = ['US Stock', 'TW Stock', 'Crypto', 'Commodity']
    shown = set()
    for atype in order + [k for k in by_type if k not in order]:
        stat = by_type.get(atype)
        if not stat or atype in shown:
            continue
        shown.add(atype)
        pct = stat['total_pnl'] / total_pnl * 100 if total_pnl else 0
        ws.append([atype, stat['trades'],
                   f"{stat['win_rate']*100:.1f}%",
                   f"${stat['total_pnl']:+,.2f}",
                   f"{pct:.1f}%"])
        fill = WIN_FILL if stat['total_pnl'] > 0 else LOSS_FILL
        for c in range(1, len(header)+1):
            ws.cell(ws.max_row, c).fill      = fill
            ws.cell(ws.max_row, c).alignment = CENTER
            ws.cell(ws.max_row, c).border    = THIN

    # 各標的 Top 10 / Bottom 10
    if not df.empty and 'P&L (USD)' in df.columns:
        sym_pnl = (df.groupby('代號')['P&L (USD)']
                     .sum().sort_values(ascending=False).reset_index())

        for title, data in [('Top 10 獲利標的', sym_pnl.head(10)),
                             ('Bottom 10 虧損標的', sym_pnl.tail(10))]:
            ws.append([])
            ws.append([title, '合計損益(USD)'])
            _style_row(ws, ws.max_row, fill=SUB_FILL, font=SUB_FONT)
            for _, r in data.iterrows():
                ws.append([r['代號'], f"${r['P&L (USD)']:+,.2f}"])
                fill = WIN_FILL if r['P&L (USD)'] > 0 else LOSS_FILL
                for c in range(1, 3):
                    ws.cell(ws.max_row, c).fill      = fill
                    ws.cell(ws.max_row, c).alignment = CENTER
                    ws.cell(ws.max_row, c).border    = THIN

    _auto_width(ws)
    return ws


# ─── Summary ─────────────────────────────────────────────────────────────────
def _write_summary(wb: Workbook, metrics: dict, equity_curve: list[dict]):
    ws = wb.create_sheet(title='📊 Summary', index=0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22

    ws.cell(1, 1, '量化交易回測報告').font = Font(bold=True, size=18, color='1F3864')
    ws.cell(2, 1,
        f'策略：趨勢動能 · Volume Profile · 布林均值回歸  |  '
        f'RR = 1:{int(config.RISK_REWARD_RATIO)}  |  Kelly = 1/{int(1/config.KELLY_FRACTION)}'
    ).font = Font(italic=True, size=10, color='595959')
    ws.append([])

    def section(title_text: str):
        ws.append([title_text])
        _style_row(ws, ws.max_row, fill=SUB_FILL, font=SUB_FONT)

    def kv(label, value, win=None):
        ws.append([label, value])
        r = ws.max_row
        ws.cell(r, 1).font      = BOLD
        ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 2).alignment = CENTER
        ws.cell(r, 1).border    = THIN
        ws.cell(r, 2).border    = THIN
        if win is True:
            ws.cell(r, 2).fill = WIN_FILL
        elif win is False:
            ws.cell(r, 2).fill = LOSS_FILL

    # ── 資金績效 ──────────────────────────────────────────────────────────
    section('💰 資金績效')
    kv('初始資金',          f"${metrics.get('initial_capital', 0):>,.2f}")
    kv('最終資金',          f"${metrics.get('final_capital', 0):>,.2f}")
    total_pnl = metrics.get('total_pnl', 0)
    kv('總損益',            f"${total_pnl:>+,.2f}", win=(total_pnl >= 0))
    kv('總報酬率',          f"{metrics.get('total_return_pct', 0):>+.2f}%",
       win=(metrics.get('total_return_pct', 0) >= 0))
    kv('年化報酬率',        f"{metrics.get('annual_return_pct', 0):>+.2f}%",
       win=(metrics.get('annual_return_pct', 0) >= 0))

    # ── 交易統計 ──────────────────────────────────────────────────────────
    ws.append([])
    section('📊 交易統計')
    kv('總交易次數',        metrics.get('total_trades', 0))
    kv('整體勝率',          f"{metrics.get('win_rate', 0)*100:.1f}%",
       win=(metrics.get('win_rate', 0) >= 0.5))
    kv('多單勝率',          f"{metrics.get('win_rate_long', 0)*100:.1f}%")
    kv('空單勝率',          f"{metrics.get('win_rate_short', 0)*100:.1f}%")
    kv('平均獲利',          f"${metrics.get('avg_win', 0):>,.2f}")
    kv('平均虧損',          f"${metrics.get('avg_loss', 0):>,.2f}")
    kv('期望值 (每筆)',      f"${metrics.get('expectancy', 0):>+,.2f}",
       win=(metrics.get('expectancy', 0) >= 0))
    kv('盈虧比 (PF)',        f"{metrics.get('profit_factor', 0):.3f}",
       win=(metrics.get('profit_factor', 0) >= 1.0))
    kv('平均 R 倍數',        f"{metrics.get('avg_r_multiple', 0):>+.3f}R",
       win=(metrics.get('avg_r_multiple', 0) >= 0))
    kv('最大連勝',          metrics.get('max_consec_wins', 0))
    kv('最大連敗',          metrics.get('max_consec_losses', 0))
    kv('平均持倉天數',       f"{metrics.get('avg_holding_days', 0):.1f} 天")
    kv('最佳單筆',          f"${metrics.get('best_trade', 0):>+,.2f}", win=True)
    kv('最差單筆',          f"${metrics.get('worst_trade', 0):>+,.2f}", win=False)

    # ── 風險指標 ──────────────────────────────────────────────────────────
    ws.append([])
    section('⚠️ 風險指標')
    kv('夏普比率',          f"{metrics.get('sharpe_ratio', 0):.3f}",
       win=(metrics.get('sharpe_ratio', 0) >= 1.0))
    kv('卡瑪比率',          f"{metrics.get('calmar_ratio', 0):.3f}",
       win=(metrics.get('calmar_ratio', 0) >= 1.0))
    kv('恢復因子',          f"{metrics.get('recovery_factor', 0):.3f}",
       win=(metrics.get('recovery_factor', 0) >= 1.0))
    kv('最大回撤 (%)',       f"{metrics.get('max_drawdown_pct', 0):.2f}%", win=False)
    kv('最大回撤 (USD)',     f"${metrics.get('max_drawdown_usd', 0):>,.2f}", win=False)

    # ── 資金曲線圖 ────────────────────────────────────────────────────────
    if equity_curve:
        eq_df    = pd.DataFrame(equity_curve)
        eq_start = ws.max_row + 3
        ws.cell(eq_start, 4, 'Date').font  = BOLD
        ws.cell(eq_start, 5, 'Equity').font = BOLD
        for i, rec in enumerate(eq_df.itertuples()):
            ws.cell(eq_start + 1 + i, 4, rec.date)
            ws.cell(eq_start + 1 + i, 5, rec.capital)

        chart = LineChart()
        chart.title  = 'Equity Curve'
        chart.style  = 10
        chart.height = 16
        chart.width  = 32
        chart.y_axis.title = 'Capital (USD)'
        chart.x_axis.title = 'Date'
        data_ref = Reference(ws, min_col=5, min_row=eq_start,
                             max_row=eq_start + len(eq_df))
        chart.add_data(data_ref, titles_from_data=True)
        ws.add_chart(chart, 'D4')

    return ws


# ─── Per-Symbol Stats ────────────────────────────────────────────────────────
SYM_COLS = [
    '代號', '資產類型',
    '買進次數', '賣出次數', '做多次數', '做空次數',
    '勝率(%)', '多單勝率(%)', '空單勝率(%)',
    '平均獲利(USD)', '平均虧損(USD)',
    '最大獲利(USD)', '最大虧損(USD)',
    '合計損益(USD)', '平均持倉天數', '平均R倍數',
]


def _write_symbol_stats(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet(title='Per Symbol Stats')
    ws.cell(1, 1, '各標的交易統計').font = TITLE_FONT
    ws.cell(2, 1, '依合計損益由高到低排序').font = Font(italic=True, color='595959')
    ws.append([])

    if df.empty or 'P&L (USD)' not in df.columns:
        ws.cell(4, 1, '無交易資料')
        return ws

    rows = []
    for sym, grp in df.groupby('代號'):
        pnl   = grp['P&L (USD)'].dropna()
        wins  = pnl[pnl > 0]
        losses = pnl[pnl <= 0]
        longs  = grp[grp['方向'].str.contains('多', na=False)]
        shorts = grp[grp['方向'].str.contains('空', na=False)]
        l_wins = longs['P&L (USD)'].dropna()
        s_wins = shorts['P&L (USD)'].dropna()

        hold = grp['持倉天數'].dropna()
        r_m  = grp['R倍數'].dropna()

        rows.append({
            '代號':         sym,
            '資產類型':     grp['資產類型'].iloc[0] if len(grp) else '',
            '買進次數':     int((grp['方向'].str.contains('多', na=False)).sum()),
            '賣出次數':     int((grp['方向'].str.contains('空', na=False)).sum()),
            '做多次數':     len(longs),
            '做空次數':     len(shorts),
            '勝率(%)':      round(len(wins) / len(pnl) * 100, 1) if len(pnl) else 0,
            '多單勝率(%)':  round((l_wins > 0).sum() / len(longs) * 100, 1) if len(longs) else 0,
            '空單勝率(%)':  round((s_wins > 0).sum() / len(shorts) * 100, 1) if len(shorts) else 0,
            '平均獲利(USD)': round(wins.mean(), 2)    if len(wins)   else 0,
            '平均虧損(USD)': round(losses.mean(), 2)  if len(losses) else 0,
            '最大獲利(USD)': round(pnl.max(), 2)      if len(pnl)    else 0,
            '最大虧損(USD)': round(pnl.min(), 2)      if len(pnl)    else 0,
            '合計損益(USD)': round(pnl.sum(), 2),
            '平均持倉天數':  round(hold.mean(), 1)    if len(hold)   else 0,
            '平均R倍數':     round(r_m.mean(), 2)     if len(r_m)    else 0,
        })

    result = pd.DataFrame(rows, columns=SYM_COLS)
    result = result.sort_values('合計損益(USD)', ascending=False).reset_index(drop=True)

    # Header
    ws.append(SYM_COLS)
    hdr_row = ws.max_row
    _style_row(ws, hdr_row, fill=HDR_FILL, font=HDR_FONT)

    pnl_col  = SYM_COLS.index('合計損益(USD)') + 1
    win_col  = SYM_COLS.index('勝率(%)') + 1

    for _, r in result.iterrows():
        ws.append([r[c] for c in SYM_COLS])
        cur = ws.max_row
        try:
            fill = WIN_FILL if float(r['合計損益(USD)'] or 0) > 0 else LOSS_FILL
        except (TypeError, ValueError):
            fill = NEUT_FILL
        for c in range(1, len(SYM_COLS) + 1):
            cell = ws.cell(cur, c)
            cell.fill      = fill
            cell.alignment = CENTER
            cell.border    = THIN

    # 合計損益色階
    pnl_letter = get_column_letter(pnl_col)
    ws.conditional_formatting.add(
        f"{pnl_letter}{hdr_row+1}:{pnl_letter}{ws.max_row}",
        ColorScaleRule(
            start_type='min', start_color='F8696B',
            mid_type='num',   mid_value=0,   mid_color='FFFFFF',
            end_type='max',   end_color='63BE7B',
        )
    )

    # 勝率色階
    wr_letter = get_column_letter(win_col)
    ws.conditional_formatting.add(
        f"{wr_letter}{hdr_row+1}:{wr_letter}{ws.max_row}",
        ColorScaleRule(
            start_type='num', start_value=0,  start_color='F8696B',
            mid_type='num',   mid_value=50,   mid_color='FFEB9C',
            end_type='num',   end_value=100,  end_color='63BE7B',
        )
    )

    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    _add_autofilter(ws, hdr_row)
    _auto_width(ws)
    return ws


# ─── 主入口 ───────────────────────────────────────────────────────────────────
def generate_excel_report(trades: list[Trade],
                          metrics: dict,
                          equity_curve: list[dict],
                          output_path: str | None = None) -> str:
    if output_path is None:
        Path(config.OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
        output_path = str(Path(config.OUTPUT_DIR) / config.OUTPUT_FILENAME)

    df = trades_to_df(trades)
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 1. Summary
    _write_summary(wb, metrics, equity_curve)

    # 2. Monthly P&L matrix
    _write_monthly_pnl(wb, df)

    # 3. Strategy stats
    _write_strategy_stats(wb, df, metrics)

    # 4. Asset stats
    _write_asset_stats(wb, df, metrics)

    # 5. 年度工作表（每年一張，內分月份）
    if not df.empty:
        df_w = df.copy()
        df_w['_year'] = df_w['出場日期'].dt.year

        # 從 equity_curve 算出每年初的資金作為報酬率基礎
        year_start_caps: dict[int, float] = {}
        if equity_curve:
            eq_df = pd.DataFrame(equity_curve)
            eq_df['_eq_year'] = pd.to_datetime(eq_df['date']).dt.year
            initial_cap = metrics.get('initial_capital', config.INITIAL_CAPITAL)
            for year in sorted(df_w['_year'].unique()):
                prev = eq_df[eq_df['_eq_year'] < year]
                year_start_caps[year] = (
                    float(prev.iloc[-1]['capital']) if not prev.empty else initial_cap
                )

        for year in sorted(df_w['_year'].unique()):
            y_df = df_w[df_w['_year'] == year].drop(columns=['_year'])
            _write_year_sheet(wb, int(year), y_df,
                              year_start_cap=year_start_caps.get(year, 0.0))

        # 6. All trades（平面總覽）
        _write_trade_sheet(wb, '📋 All Trades',
                           df_w.drop(columns=['_year']),
                           title='全部交易記錄')

    # 7. Per-symbol stats
    if not df.empty:
        _write_symbol_stats(wb, df)

    wb.save(output_path)
    print(f'\n[OK] 回測報告已儲存：{output_path}')
    return output_path
