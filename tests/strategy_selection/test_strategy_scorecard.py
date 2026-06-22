"""TASK-014BY -- offline tests for the scorecard, baseline manifest, challengers,
and demo scaffold. Deterministic; temp roots only; no network/Bybit/order."""

from __future__ import annotations

import importlib
import json
import pathlib
import sys

import pytest

ROOT = pathlib.Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.strategy_selection import forward30_diagnostics as diag
from src.strategy_selection import strategy_scorecard as sc

from tests.strategy_selection.test_forward30_diagnostics import make_run, days

cli = importlib.import_module("scripts.analyze_forward30_strategy_selection")


def diagnostics_for(tmp_path, n, *, returns=None, costs=None, positions=None,
                    strategy=diag.EXPECTED_STRATEGY_NAME, oos=None):
    make_run(tmp_path, "prev3y_crypto", days(n), returns=returns, costs=costs,
             positions=positions, strategy=strategy)
    run = diag.load_forward_run(tmp_path, "prev3y_crypto")
    return diag.run_all_diagnostics(run, [], oos_numbers=oos)


# --- Scorecard labels ------------------------------------------------------


def test_scorecard_needs_more_data_on_short_sample(tmp_path):
    d = diagnostics_for(tmp_path, 2)
    score = sc.score_strategy(d)
    assert score["label"] == sc.NEEDS_MORE_DATA
    assert score["ranks_by_return_only"] is False


def test_scorecard_reject_data_incomplete_on_strategy_mismatch(tmp_path):
    d = diagnostics_for(tmp_path, 30, strategy="wrong_strategy")
    score = sc.score_strategy(d)
    assert score["label"] == sc.REJECT_DATA_INCOMPLETE


def test_scorecard_keep_baseline_full_sample(tmp_path):
    d = diagnostics_for(tmp_path, 30, returns=[0.1] * 30,
                        costs={"fee_cost_usd": 0.1}, oos={"sharpe": 0.9})
    score = sc.score_strategy(d)
    assert score["label"] == sc.KEEP_BASELINE


def test_scorecard_reject_excessive_drawdown(tmp_path):
    # 30 days with a deep mid-sample crash -> max_dd > 20%.
    rets = [0.1] * 10 + [-30.0] + [0.1] * 19
    d = diagnostics_for(tmp_path, 30, returns=rets, costs={"fee_cost_usd": 0.1}, oos={"sharpe": 0.9})
    score = sc.score_strategy(d)
    assert score["label"] == sc.REJECT_EXCESSIVE_DRAWDOWN


def test_scorecard_reject_concentrated(tmp_path):
    positions = [{"symbol": "DOMUSDT", "side": "long", "weight": 0.9, "position_usd": 9000.0}] + \
                [{"symbol": f"S{i}USDT", "side": "long", "weight": 0.001, "position_usd": 10.0}
                 for i in range(5)]
    d = diagnostics_for(tmp_path, 30, returns=[0.1] * 30, costs={"fee_cost_usd": 0.1},
                        positions=positions, oos={"sharpe": 0.9})
    score = sc.score_strategy(d)
    assert score["label"] == sc.REJECT_CONCENTRATED_RESULT


def test_scorecard_deterministic(tmp_path):
    d = diagnostics_for(tmp_path, 2)
    s1 = sc.score_strategy(d)
    s2 = sc.score_strategy(d)
    assert json.dumps(s1, sort_keys=True) == json.dumps(s2, sort_keys=True)


# --- V1 baseline manifest --------------------------------------------------


def test_manifest_identity_stable_and_pending_on_partial(tmp_path):
    d = diagnostics_for(tmp_path, 2)
    m1 = sc.build_v1_baseline_manifest(code_commit="c" * 40, pilot_id="P", diagnostics=d)
    m2 = sc.build_v1_baseline_manifest(code_commit="c" * 40, pilot_id="P", diagnostics=d)
    assert m1["manifest_fingerprint"] == m2["manifest_fingerprint"]
    # Incomplete local snapshot -> truthful PENDING status, not FROZEN_ACTIVE_BASELINE.
    assert m1["status"] == sc.FROZEN_BASELINE_IDENTITY_PENDING
    assert m1["baseline_identity"]["exclusions"] == list(sc.EXCLUSIONS)
    assert m1["local_evidence_snapshot"]["local_snapshot_status"] == diag.PARTIAL
    assert m1["local_evidence_snapshot"]["authoritative"] is False


def test_manifest_full_coverage_is_frozen_active(tmp_path):
    d = diagnostics_for(tmp_path, 30, returns=[0.1] * 30)
    m = sc.build_v1_baseline_manifest(code_commit="c" * 40, pilot_id="P", diagnostics=d)
    assert m["status"] == sc.FROZEN_ACTIVE_BASELINE
    assert m["local_evidence_snapshot"]["authoritative"] is True


def test_incomplete_manifest_cannot_claim_full30_finalization(tmp_path):
    d = diagnostics_for(tmp_path, 2)
    m = sc.build_v1_baseline_manifest(code_commit="c" * 40, pilot_id="P", diagnostics=d)
    assert m["status"] != sc.FROZEN_ACTIVE_BASELINE
    assert m["local_evidence_snapshot"]["present_date_count"] == 2
    assert m["local_evidence_snapshot"]["expected_date_count"] == 30


def test_manifest_fingerprint_changes_with_commit(tmp_path):
    d = diagnostics_for(tmp_path, 2)
    m1 = sc.build_v1_baseline_manifest(code_commit="a" * 40, pilot_id="P", diagnostics=d)
    m2 = sc.build_v1_baseline_manifest(code_commit="b" * 40, pilot_id="P", diagnostics=d)
    assert m1["manifest_fingerprint"] != m2["manifest_fingerprint"]


def test_committed_v1_manifest_identity_pending_full30():
    p = ROOT / "docs" / "research" / "strategy_selection" / "V1_BASELINE_MANIFEST.json"
    data = json.loads(p.read_text(encoding="utf-8"))
    # Committed manifest freezes IDENTITY but truthfully marks the full-30 evidence pending.
    assert data["status"] == sc.FROZEN_BASELINE_IDENTITY_PENDING
    assert data["baseline_identity"]["strategy_id"] == diag.EXPECTED_STRATEGY_NAME
    assert data["local_evidence_snapshot"]["local_snapshot_status"] == diag.PARTIAL
    assert data["local_evidence_snapshot"]["authoritative"] is False
    assert data["manifest_fingerprint"] == sc.manifest_fingerprint(data)


# --- Challenger hypotheses -------------------------------------------------


def test_insufficient_2_of_30_emits_zero_challengers(tmp_path):
    # Even with all capabilities confirmed, a 2/30 day-0 sample selects ZERO challengers.
    d = diagnostics_for(tmp_path, 2)
    caps = {"canonical_regime_gate": True, "overlay_gate_machinery": True,
            "volatility_adjusted_sizing": True, "existing_exit_rule_variants": True}
    ch = sc.generate_challenger_hypotheses(d, caps)
    assert ch["emitted_count"] == 0
    assert ch["evidence_strength"] == "INSUFFICIENT_SAMPLE_NO_SELECTION"
    assert ch["promotion_status"] == "NONE_PROMOTED"
    # Structural observations live under future_research_candidates, not as hypotheses.
    assert len(ch["future_research_candidates"]) >= 1


def test_full_sufficient_emits_at_most_two_evidence_backed(tmp_path):
    # 30 comparable dates + a comparable shadow -> challengers may be selected (<=2).
    make_run(tmp_path, "prev3y_crypto", days(30), returns=[0.1] * 30)
    make_run(tmp_path, "prev3y_crypto_shadow_a_roll12", days(30), returns=[0.1] * 30, strategy="shadow_x")
    primary = diag.load_forward_run(tmp_path, "prev3y_crypto")
    shadow = diag.load_forward_run(tmp_path, "prev3y_crypto_shadow_a_roll12")
    d = diag.run_all_diagnostics(primary, [shadow])
    caps = {"canonical_regime_gate": False, "overlay_gate_machinery": True,
            "volatility_adjusted_sizing": True, "existing_exit_rule_variants": True}
    ch = sc.generate_challenger_hypotheses(d, caps)
    assert ch["emitted_count"] <= sc.MAX_CHALLENGERS
    assert ch["evidence_gate"]["sufficient"] is True
    for h in ch["hypotheses"]:
        assert h["status"] == "EVIDENCE_BACKED"
        changed = [k for k, v in h["changes"].items() if v]
        assert len(changed) >= 1


def test_overlay_gate_not_labelled_regime_without_canonical_def(tmp_path):
    make_run(tmp_path, "prev3y_crypto", days(30), returns=[0.1] * 30)
    make_run(tmp_path, "prev3y_crypto_shadow_a_roll12", days(30), returns=[0.1] * 30, strategy="shadow_x")
    primary = diag.load_forward_run(tmp_path, "prev3y_crypto")
    shadow = diag.load_forward_run(tmp_path, "prev3y_crypto_shadow_a_roll12")
    d = diag.run_all_diagnostics(primary, [shadow])
    caps = {"canonical_regime_gate": False, "overlay_gate_machinery": True,
            "volatility_adjusted_sizing": False, "existing_exit_rule_variants": False}
    ch = sc.generate_challenger_hypotheses(d, caps)
    overlay = [h for h in ch["hypotheses"] if "overlay" in h["id"]]
    assert overlay and overlay[0]["changes"]["regime_filter"] is False


# --- Demo comparison scaffold ----------------------------------------------


def test_demo_scaffold_not_yet_available(tmp_path):
    scaffold = sc.build_demo_comparison_scaffold(
        pilot_id="P", baseline_manifest_fingerprint="sha256:abc", completed_successful_days="UNKNOWN")
    assert all(v == "NOT_YET_AVAILABLE" for v in scaffold["demo_metrics"].values())
    assert scaffold["baseline_linkage"]["baseline_name"] == "V1"
    assert scaffold["completed_successful_days"] == "UNKNOWN"


def test_read_completed_days_unknown_when_absent(tmp_path):
    val = sc.read_completed_successful_days("NO_SUCH_PILOT", str(tmp_path))
    assert val == "UNKNOWN"


# --- Comparison regression / CLI ------------------------------------------


def test_cli_end_to_end_writes_outputs_and_workbook(tmp_path):
    pytest.importorskip("openpyxl")
    make_run(tmp_path / "fwd", "prev3y_crypto", days(2))
    out = tmp_path / "out"
    rc = cli.main(["--input-root", str(tmp_path / "fwd"), "--run-key", "prev3y_crypto",
                   "--output-root", str(out), "--json-only"])
    assert rc == 0
    for f in ("input_audit.json", "v1_baseline_manifest.json", "overall_metrics.json",
              "contribution_by_symbol.csv", "contribution_by_side.csv", "drawdown_analysis.json",
              "cost_stress.json", "primary_shadow_comparison.json", "strategy_scorecard.json",
              "challenger_hypotheses.json", "demo_comparison_scaffold.json",
              "strategy_selection_report.md", "strategy_selection_report.xlsx"):
        assert (out / f).exists(), f
    from openpyxl import load_workbook
    wb = load_workbook(str(out / "strategy_selection_report.xlsx"))
    assert wb.sheetnames == cli.WORKBOOK_SHEETS
    assert len(wb.sheetnames) == 12


def test_cli_is_read_only_on_sources(tmp_path):
    fwd = tmp_path / "fwd"
    make_run(fwd, "prev3y_crypto", days(2))
    summary_path = fwd / "prev3y_crypto" / "forward_summary.json"
    before = summary_path.read_bytes()
    cli.main(["--input-root", str(fwd), "--run-key", "prev3y_crypto",
              "--output-root", str(tmp_path / "out"), "--json-only"])
    assert summary_path.read_bytes() == before  # source untouched


def test_incompatible_periods_not_ranked_comparable(tmp_path):
    make_run(tmp_path, "prev3y_crypto", days(2))
    make_run(tmp_path, "prev3y_crypto_shadow_a_roll12", days(1), strategy="shadow_x")
    primary = diag.load_forward_run(tmp_path, "prev3y_crypto")
    shadow = diag.load_forward_run(tmp_path, "prev3y_crypto_shadow_a_roll12")
    ps = diag.compute_primary_shadow(primary, [shadow])
    assert ps["ranking_status"] == "ALL_INSUFFICIENT_EVIDENCE"
