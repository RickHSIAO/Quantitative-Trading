"""Tests for TASK-014BQ demo round-trip closeout + pilot reporting foundation.

All tests are strictly offline. No network, no orders, no Notion/Discord HTTP.
"""

from __future__ import annotations

import dataclasses
import importlib
import json
import pathlib
import sys
from decimal import Decimal

import pytest

ROOT = pathlib.Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from src import demo_strategy_pilot_reporting as rep
from src import demo_strategy_pilot_store as store_mod
from src.demo_strategy_pilot_store import (
    DuplicateRecordError,
    MalformedStoreError,
    PilotStore,
)

wb_builder = importlib.import_module("scripts.build_demo_strategy_pilot_workbook")
notion = importlib.import_module("scripts.preview_demo_strategy_pilot_notion_payload")
discord = importlib.import_module("scripts.preview_demo_strategy_pilot_discord_summary")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def make_config():
    return rep.PilotConfig(pilot_id="pilot_fixture", start_date="2026-06-22",
                           strategy_name="prev3y_momentum", initial_equity_usdt=Decimal("10000"))


def make_daily(date="2026-06-22", day=1, **over):
    base = dict(date=date, pilot_day=day, runner_status="RUNNING", signal_count=3,
                order_count=1, filled_count=1, closed_trade_count=1,
                realized_pnl_usdt=Decimal("1.23"), trading_fees_usdt=Decimal("0.05"),
                funding_pnl_usdt=Decimal("0"), daily_net_pnl_usdt=Decimal("1.18"),
                cumulative_net_pnl_usdt=Decimal("1.18"), daily_return_pct=Decimal("0.0118"),
                cumulative_return_pct=Decimal("0.0118"), max_drawdown_pct=Decimal("0.5"),
                current_position_symbol="SOLUSDT", current_position_side="Buy",
                current_position_qty=Decimal("0.1"), alerts_triggered=("none",), notes="ok")
    base.update(over)
    return rep.PilotDailyRecord(**base)


def make_trade(trade_id="T1", **over):
    base = dict(pilot_id="pilot_fixture", trade_id=trade_id, signal_id="S1", symbol="SOLUSDT",
                side="Buy", entry_order_id="e1", exit_order_id="x1", entry_order_link_id="el1",
                exit_order_link_id="xl1", entry_time_utc="2026-06-22T10:00:00Z",
                exit_time_utc="2026-06-22T11:00:00Z", requested_qty=Decimal("0.1"),
                executed_qty=Decimal("0.1"), entry_price=Decimal("74.11"), exit_price=Decimal("73.8"),
                entry_fee=Decimal("0.00407605"), exit_fee=Decimal("0.004059"), funding_pnl=Decimal("0"),
                gross_pnl=Decimal("-0.031"), net_pnl=Decimal("-0.03913505"),
                slippage_entry_bps=Decimal("1.2"), slippage_exit_bps=Decimal("0.8"),
                final_status="CLOSED", included_in_performance=True)
    base.update(over)
    return rep.PilotTradeRecord(**base)


def make_audit(**over):
    base = dict(timestamp_utc="2026-06-22T00:00:00Z", pilot_id="pilot_fixture",
                event_type="DAILY_CLOSE", component="runner", status="OK", message="done")
    base.update(over)
    return rep.PilotAuditEvent(**base)


def seeded_store(tmp_path):
    st = PilotStore("pilot_fixture", tmp_path)
    st.write_config(make_config())
    st.append_daily(make_daily())
    st.append_trade(make_trade())
    st.append_audit(make_audit())
    return st


# ---------------------------------------------------------------------------
# 1-6. Closeout
# ---------------------------------------------------------------------------


def test_closeout_opening_identity():
    c = rep.build_round_trip_closeout()
    o = c["opening"]
    assert o["order_id"] == "77173918-71f6-4829-91c9-025bd8cd76fa"
    assert o["order_link_id"] == "BO1-4696d511edf11b50"
    assert o["side"] == "Buy" and o["reduce_only"] is False
    assert o["final_conclusion"] == "DEMO_ORDER_FILLED_VERIFIED"


def test_closeout_closing_identity():
    c = rep.build_round_trip_closeout()
    cl = c["closing"]
    assert cl["close_order_id"] == "4ae9e849-655c-4ac3-b830-d49d587c4f4c"
    assert cl["close_order_link_id"] == "BC1-566b8509e96b2def"
    assert cl["side"] == "Sell" and cl["reduce_only"] is True
    assert cl["final_conclusion"] == "DEMO_REDUCE_ONLY_CLOSE_FILLED_POSITION_ZERO_VERIFIED"


def test_position_zero_verified():
    c = rep.build_round_trip_closeout()
    assert c["safety"]["position_zero_verified"] is True
    assert c["closing"]["position_after"] == "0"


def test_no_short_position_created():
    c = rep.build_round_trip_closeout()
    assert c["safety"]["short_position_created"] is False
    assert c["closing"]["short_position_after"] is False


def test_estimated_net_pnl_decimal():
    pnl = rep.compute_round_trip_pnl(open_avg_price="74.11", close_avg_price="73.8",
                                     quantity="0.1", open_fee="0.00407605", close_fee="0.004059")
    assert pnl["gross_price_pnl"] == Decimal("-0.031")
    assert pnl["total_fees"] == Decimal("0.00813505")
    assert pnl["estimated_net_pnl_excluding_funding"] == Decimal("-0.03913505")
    assert isinstance(pnl["estimated_net_pnl_excluding_funding"], Decimal)


def test_validation_trade_excluded_from_strategy_metrics():
    c = rep.build_round_trip_closeout()
    cls = c["classification"]
    assert cls["trade_classification"] == "MANUAL_EXECUTION_PIPELINE_VALIDATION"
    assert cls["included_in_strategy_performance"] is False
    assert cls["included_in_pilot_performance"] is False


def test_committed_closeout_artifacts_exist_and_match():
    j = ROOT / "docs/research/review_packets/TASK-014BQ_DEMO_ROUND_TRIP_CLOSEOUT.json"
    m = ROOT / "docs/research/review_packets/TASK-014BQ_DEMO_ROUND_TRIP_CLOSEOUT.md"
    assert j.exists() and m.exists()
    data = json.loads(j.read_text(encoding="utf-8"))
    assert data["calculation"]["estimated_net_pnl_excluding_funding"] == "-0.03913505"
    assert "X-BAPI-SIGN" not in j.read_text(encoding="utf-8")
    assert "MANUAL_EXECUTION_PIPELINE_VALIDATION" in m.read_text(encoding="utf-8")


# ---------------------------------------------------------------------------
# 7-8. Dataclasses / Decimal serialization
# ---------------------------------------------------------------------------


def test_frozen_dataclasses():
    for c in (rep.PilotConfig, rep.PilotDailyRecord, rep.PilotTradeRecord, rep.PilotAuditEvent):
        assert dataclasses.is_dataclass(c)
    d = make_daily()
    with pytest.raises(dataclasses.FrozenInstanceError):
        d.signal_count = 9  # type: ignore[misc]


def test_decimal_serialized_as_strings():
    d = make_daily().to_dict()
    assert d["realized_pnl_usdt"] == "1.23" and isinstance(d["realized_pnl_usdt"], str)
    t = make_trade().to_dict()
    assert t["net_pnl"] == "-0.03913505" and isinstance(t["net_pnl"], str)
    cfg = make_config().to_dict()
    assert cfg["initial_equity_usdt"] == "10000"
    assert cfg["environment"] == "BYBIT_DEMO_ONLY" and cfg["maximum_calendar_days"] == 14
    assert cfg["excel_enabled"] is True


# ---------------------------------------------------------------------------
# 9-16. Store
# ---------------------------------------------------------------------------


def test_append_daily(tmp_path):
    st = PilotStore("p", tmp_path)
    st.append_daily(make_daily())
    rows = st.read_daily()
    assert len(rows) == 1 and rows[0]["date"] == "2026-06-22"


def test_duplicate_daily_date_refused(tmp_path):
    st = PilotStore("p", tmp_path)
    st.append_daily(make_daily())
    with pytest.raises(DuplicateRecordError):
        st.append_daily(make_daily())


def test_idempotent_daily_upsert(tmp_path):
    st = PilotStore("p", tmp_path)
    st.append_daily(make_daily(signal_count=3))
    st.upsert_daily(make_daily(signal_count=9))
    rows = st.read_daily()
    assert len(rows) == 1 and rows[0]["signal_count"] == 9
    st.upsert_daily(make_daily(date="2026-06-23", day=2))
    assert len(st.read_daily()) == 2


def test_append_trade(tmp_path):
    st = PilotStore("p", tmp_path)
    st.append_trade(make_trade())
    assert len(st.read_trades()) == 1


def test_duplicate_trade_id_refused(tmp_path):
    st = PilotStore("p", tmp_path)
    st.append_trade(make_trade())
    with pytest.raises(DuplicateRecordError):
        st.append_trade(make_trade())


def test_append_audit_event(tmp_path):
    st = PilotStore("p", tmp_path)
    st.append_audit(make_audit())
    st.append_audit(make_audit(message="second"))
    assert len(st.read_audit()) == 2


def test_malformed_jsonl_refused(tmp_path):
    st = PilotStore("p", tmp_path)
    st.append_daily(make_daily())
    with open(st.daily_path, "a", encoding="utf-8") as fh:
        fh.write("{not valid json\n")
    with pytest.raises(MalformedStoreError):
        st.read_daily()


def test_atomic_latest_summary_write(tmp_path):
    st = PilotStore("p", tmp_path)
    st.write_latest_summary({"cumulative_net_pnl_usdt": "1.18", "pilot_day": 1})
    assert not (st.dir / (store_mod.LATEST_SUMMARY_FILENAME + ".tmp")).exists()
    assert st.read_latest_summary()["pilot_day"] == 1


def test_config_round_trip(tmp_path):
    st = PilotStore("p", tmp_path)
    st.write_config(make_config())
    cfg = st.read_config()
    assert cfg["strategy_name"] == "prev3y_momentum" and cfg["environment"] == "BYBIT_DEMO_ONLY"


# ---------------------------------------------------------------------------
# 17-26. Workbook
# ---------------------------------------------------------------------------


def test_valid_empty_workbook(tmp_path):
    paths = wb_builder.build_workbook("empty_pilot", tmp_path, snapshot_date="20260621")
    wb = load_workbook(paths["workbook"])
    assert wb.sheetnames == wb_builder.SHEET_ORDER
    # Each table sheet has headers in row 1.
    assert wb["Daily Performance"]["A1"].value == "Date"


def test_required_six_sheets_in_order(tmp_path):
    paths = wb_builder.build_workbook("p", tmp_path, snapshot_date="20260621")
    wb = load_workbook(paths["workbook"])
    assert wb.sheetnames == [
        "Pilot Summary", "Daily Performance", "Trades",
        "Execution Quality", "Forward Comparison", "Audit Log",
    ]


def test_workbook_reopens_with_openpyxl(tmp_path):
    paths = wb_builder.build_workbook("p", tmp_path, snapshot_date="20260621")
    wb = load_workbook(paths["workbook"])  # must not raise
    assert wb is not None


def test_headers_exist_on_every_sheet(tmp_path):
    paths = wb_builder.build_workbook("p", tmp_path, snapshot_date="20260621")
    wb = load_workbook(paths["workbook"])
    for name in wb.sheetnames:
        assert wb[name]["A1"].value not in (None, "")


def test_daily_trade_audit_rows_exported(tmp_path):
    seeded_store(tmp_path)
    paths = wb_builder.build_workbook("pilot_fixture", tmp_path, snapshot_date="20260621")
    wb = load_workbook(paths["workbook"])
    assert wb["Daily Performance"]["A2"].value == "2026-06-22"
    assert wb["Trades"]["B2"].value == "T1"
    assert wb["Audit Log"]["A2"].value == "2026-06-22T00:00:00Z"


def test_percentage_cells_numeric(tmp_path):
    seeded_store(tmp_path)
    paths = wb_builder.build_workbook("pilot_fixture", tmp_path, snapshot_date="20260621")
    wb = load_workbook(paths["workbook"])
    dp = wb["Daily Performance"]
    # "Daily Return %" is column 13.
    cell = dp.cell(row=2, column=13)
    assert cell.data_type == "n"
    assert "%" in cell.number_format


def test_monetary_cells_numeric(tmp_path):
    seeded_store(tmp_path)
    paths = wb_builder.build_workbook("pilot_fixture", tmp_path, snapshot_date="20260621")
    wb = load_workbook(paths["workbook"])
    dp = wb["Daily Performance"]
    # "Realized PnL USDT" is column 8.
    cell = dp.cell(row=2, column=8)
    assert cell.data_type == "n"
    assert float(cell.value) == pytest.approx(1.23)


def test_snapshot_workbook_generated(tmp_path):
    paths = wb_builder.build_workbook("p", tmp_path, snapshot_date="20260621")
    snap = pathlib.Path(paths["snapshot"])
    assert snap.exists() and snap.name == "demo_strategy_pilot_results_20260621.xlsx"
    assert load_workbook(str(snap)) is not None


def test_freeze_and_filter_set(tmp_path):
    seeded_store(tmp_path)
    paths = wb_builder.build_workbook("pilot_fixture", tmp_path, snapshot_date="20260621")
    wb = load_workbook(paths["workbook"])
    dp = wb["Daily Performance"]
    assert dp.freeze_panes == "A2"
    assert dp.auto_filter.ref is not None


def test_forward_comparison_accepts_fixture(tmp_path):
    seeded_store(tmp_path)
    comp = {"columns": ["Metric", "Pilot", "Forward"],
            "rows": [{"Metric": "net_pnl", "Pilot": "1.18", "Forward": "2.00"}]}
    paths = wb_builder.build_workbook("pilot_fixture", tmp_path, comparison=comp, snapshot_date="20260621")
    wb = load_workbook(paths["workbook"])
    fc = wb["Forward Comparison"]
    assert fc["A1"].value == "Metric" and fc["A2"].value == "net_pnl"


# ---------------------------------------------------------------------------
# 27-32. Notion / Discord previews
# ---------------------------------------------------------------------------


def test_notion_payload_exact_fields(tmp_path):
    st = seeded_store(tmp_path)
    payload = notion.build_notion_payload("pilot_fixture", st.read_daily()[-1])
    props = payload["properties"]
    for field in ("Date", "Pilot ID", "Pilot Day", "Runner Status", "Signal Count",
                  "Order Count", "Filled Count", "Closed Trade Count", "Realized PnL USDT",
                  "Trading Fees USDT", "Funding PnL USDT", "Daily Net PnL USDT",
                  "Cumulative Net PnL USDT", "Daily Return %", "Cumulative Return %",
                  "Max Drawdown %", "Current Position", "Alerts Triggered",
                  "Excel Export Status", "Notion Sync Status", "Discord Notify Status", "Notes"):
        assert field in props


def test_notion_idempotent_key(tmp_path):
    st = seeded_store(tmp_path)
    payload = notion.build_notion_payload("pilot_fixture", st.read_daily()[-1])
    assert payload["idempotency_key"] == "pilot_fixture:2026-06-22"
    assert payload["operation"] == "upsert"


def test_notion_preview_no_http_and_no_token():
    src = pathlib.Path(notion.__file__).read_text(encoding="utf-8")
    assert "import requests" not in src and "urllib" not in src
    assert "os.environ" not in src and "getenv" not in src and "NOTION_TOKEN" not in src
    payload = notion.build_notion_payload("p", None)
    assert payload["no_http_performed"] is True and payload["notion_token_read"] is False


def test_discord_preview_chinese_fields(tmp_path):
    st = seeded_store(tmp_path)
    text = discord.build_discord_summary("pilot_fixture", st.read_daily()[-1])
    assert "Bybit Demo 策略試運行日報" in text
    assert "Pilot ID：pilot_fixture" in text
    assert "已實現 PnL" in text and "最大回撤" in text and "目前持倉" in text


def test_discord_preview_no_http():
    src = pathlib.Path(discord.__file__).read_text(encoding="utf-8")
    assert "import requests" not in src and "urllib" not in src
    assert "webhook" not in src.lower() or "no webhook" in src.lower()


# ---------------------------------------------------------------------------
# 33-36. Safety scans
# ---------------------------------------------------------------------------


NEW_FILES = [
    ROOT / "src/demo_strategy_pilot_reporting.py",
    ROOT / "src/demo_strategy_pilot_store.py",
    ROOT / "scripts/build_demo_strategy_pilot_workbook.py",
    ROOT / "scripts/preview_demo_strategy_pilot_notion_payload.py",
    ROOT / "scripts/preview_demo_strategy_pilot_discord_summary.py",
]


def _import_lines(text):
    return [ln.strip() for ln in text.splitlines() if ln.strip().startswith(("import ", "from "))]


def test_no_bybit_or_network_import():
    for f in NEW_FILES:
        for ln in _import_lines(f.read_text(encoding="utf-8")):
            assert "executors.bybit" not in ln
            assert "import requests" not in ln
            assert "pybit" not in ln


def test_no_live_executor_import():
    for f in NEW_FILES:
        text = f.read_text(encoding="utf-8")
        assert "BybitExecutor" not in text
        for ln in _import_lines(text):
            assert "src.risk" not in ln
            assert not ln.startswith("import main") and not ln.startswith("from main ")


def test_no_scheduler():
    for f in NEW_FILES:
        for ln in _import_lines(f.read_text(encoding="utf-8")):
            low = ln.lower()
            assert "apscheduler" not in low and "crontab" not in low
            assert not low.startswith("import schedule") and not low.startswith("import sched")


def test_no_order_endpoint_string_in_new_modules():
    for f in NEW_FILES:
        text = f.read_text(encoding="utf-8")
        assert "/v5/order" not in text
        assert "api-demo.bybit.com" not in text
        assert "order/create" not in text


# ---------------------------------------------------------------------------
# 37-40. Docs / outputs location
# ---------------------------------------------------------------------------


def test_documentation_updated():
    for rel in ("README.md", "docs/research/commands/NEXT_ACTION.md",
                "docs/research/commands/COMMAND_LOG.md"):
        text = (ROOT / rel).read_text(encoding="utf-8")
        assert "TASK-014BQ" in text


def test_outputs_path_is_under_gitignored_outputs():
    # Canonical runtime root lives under outputs/ (kept out of Git).
    rel = store_mod.CANONICAL_PILOT_ROOT.resolve().relative_to(rep.PROJECT_ROOT.resolve())
    assert str(rel).replace("\\", "/") == "outputs/demo_trading/pilot"
    gi = (ROOT / ".gitignore").read_text(encoding="utf-8") if (ROOT / ".gitignore").exists() else ""
    assert "outputs" in gi
