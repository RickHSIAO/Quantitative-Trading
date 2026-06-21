"""Tests for TASK-014BR demo strategy pilot daily runner (DRY-RUN orchestration).

All tests are strictly offline: no Bybit/Notion/Discord network, no orders.
Notion/Discord use injected fake transports. Runtime state uses temp roots.
"""

from __future__ import annotations

import dataclasses
import importlib
import json
import pathlib
import sys
from decimal import Decimal

import pytest

ROOT = pathlib.Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from src import demo_strategy_pilot_daily_journal as jr
from src import demo_strategy_pilot_daily_runner as rr
from src import demo_strategy_pilot_discord_notify as dn
from src import demo_strategy_pilot_notion_sync as ns
from src.demo_strategy_pilot_reporting import PilotConfig
from src.demo_strategy_pilot_store import PilotStore

cli = importlib.import_module("scripts.run_demo_strategy_pilot_daily")

PILOT = "P1"
DATE = "2026-06-22"


# ---------------------------------------------------------------------------
# Fakes
# ---------------------------------------------------------------------------


class FakeNotionTransport:
    def __init__(self, *, existing_page=None, raise_on=None):
        self.existing_page = existing_page
        self.raise_on = raise_on
        self.query_calls = 0
        self.upsert_calls = 0
        self.seen_headers = []

    def query(self, *, database_id, idempotency_key, headers, properties=None):
        self.query_calls += 1
        self.seen_headers.append(headers)
        if self.raise_on == "query":
            raise RuntimeError("notion query failed")
        return {"page_id": self.existing_page} if self.existing_page else {}

    def upsert(self, *, database_id, page_id, properties, headers):
        self.upsert_calls += 1
        if self.raise_on == "upsert":
            raise RuntimeError("notion upsert failed")
        return {"page_id": page_id or "new-page"}


class FakeDiscordTransport:
    def __init__(self, *, raise_on=False):
        self.raise_on = raise_on
        self.posts = []

    def post(self, *, webhook_url, content):
        self.posts.append((webhook_url, content))
        if self.raise_on:
            raise RuntimeError("discord post failed")
        return True


def cfg(**over):
    base = dict(pilot_id=PILOT, start_date="2026-06-22", strategy_name=rr.EXPECTED_STRATEGY_NAME,
                initial_equity_usdt=Decimal("10000"), notion_enabled=True, excel_enabled=True,
                discord_enabled=True)
    base.update(over)
    return PilotConfig(**base)


def strat(**over):
    base = dict(data_date="2026-06-22", data_status="OK",
                signals=[{"symbol": "SOLUSDT", "side": "long"}])
    base.update(over)
    return base


def run(tmp_path, *, mode="dry_run", strategy_result=None, config=None, notion_sync=None,
        discord_notify=None, allow_notion=False, allow_discord=False, date=DATE):
    return rr.run_daily(
        mode=mode, pilot_id=PILOT, date=date, config=config or cfg(),
        strategy_result=strat() if strategy_result is None else strategy_result,
        output_root=str(tmp_path), notion_sync=notion_sync, discord_notify=discord_notify,
        allow_notion_network=allow_notion, allow_discord_network=allow_discord,
        snapshot_date="20260622",
    )


# ---------------------------------------------------------------------------
# 1-11. Identity, modes, plan, signals
# ---------------------------------------------------------------------------


def test_task_identity():
    assert rr.TASK_ID == "TASK-014BR"
    assert rr.IDENTITY == "DEMO-STRATEGY-PILOT-DAILY-RUNNER-DRY-RUN"


def test_default_mode_is_offline_plan():
    assert cli.build_parser().get_default("mode") == "plan"


def test_no_order_execution_mode_exists():
    assert "execute" not in rr.MODES and "execute_once" not in rr.MODES
    opts = set()
    for a in cli.build_parser()._actions:
        opts.update(a.option_strings)
    for forbidden in ("--execute", "--send-order", "--allow-bybit-order", "--live",
                      "--scheduler", "--retry-order", "--reset-journal", "--force-rerun",
                      "--qty", "--symbol", "--endpoint"):
        assert forbidden not in opts


def test_execution_authorization_always_false():
    assert rr.ORDER_EXECUTION_AUTHORIZED is False
    plan = rr.build_plan(pilot_id=PILOT, date=DATE, config=cfg(), strategy_result=strat(), runner_mode="plan")
    assert plan.order_execution_authorized is False


def test_exact_non_authorization_reason():
    assert rr.REASON_NOT_AUTHORIZED == "TASK-014BR_IS_DRY_RUN_REPORTING_WIRING_ONLY"
    plan = rr.build_plan(pilot_id=PILOT, date=DATE, config=cfg(), strategy_result=strat(), runner_mode="plan")
    assert plan.reason_execution_not_authorized == rr.REASON_NOT_AUTHORIZED


def test_existing_30day_strategy_identifier_reused():
    assert rr.EXPECTED_STRATEGY_NAME == "prev3y_crypto_combined_paper_safe_variant"
    assert rr.resolve_strategy_name() == rr.EXPECTED_STRATEGY_NAME
    assert rr.resolve_strategy_name({"strategy": rr.EXPECTED_STRATEGY_NAME}) == rr.EXPECTED_STRATEGY_NAME


def test_ambiguous_strategy_fails_closed():
    with pytest.raises(rr.StrategyAmbiguousError):
        rr.resolve_strategy_name({"strategy": "prev3y_crypto_shadow_a_roll12"})
    with pytest.raises(rr.StrategyAmbiguousError):
        rr.resolve_strategy_name({"strategy": "some_other_strategy"})


def test_signal_normalization_deterministic():
    s = strat(signals=[{"symbol": "solusdt", "side": "Buy"}, {"symbol": "BTCUSDT", "side": "short"}])
    a = rr.normalize_signals(s)
    b = rr.normalize_signals(s)
    assert a == b
    assert a[0]["symbol"] == "BTCUSDT" and a[1]["symbol"] == "SOLUSDT"
    assert a[1]["side"] == "long"


def test_protected_symbols_blocked():
    actions = rr.classify_actions(rr.normalize_signals(strat(signals=[{"symbol": "ENAUSDT", "side": "long"}])))
    assert actions[0]["eligibility"] == rr.PROTECTED_BLOCKED
    assert actions[0]["executable"] is False


def test_invalid_signals_blocked():
    actions = rr.classify_actions(rr.normalize_signals(strat(signals=[{"symbol": "", "side": "weird"}])))
    assert actions[0]["eligibility"] == rr.INVALID_BLOCKED


def test_proposed_actions_never_increment_order_count(tmp_path):
    result = run(tmp_path, strategy_result=strat(signals=[{"symbol": "SOLUSDT", "side": "long"},
                                                          {"symbol": "BTCUSDT", "side": "long"}]))
    assert result.daily_record["order_count"] == 0
    assert len(result.plan["proposed_actions"]) == 2


# ---------------------------------------------------------------------------
# 12-16. Daily record + idempotency + conflict
# ---------------------------------------------------------------------------


def test_dry_run_appends_exactly_one_daily_record(tmp_path):
    run(tmp_path)
    store = PilotStore(PILOT, str(tmp_path))
    assert len(store.read_daily()) == 1


def test_dry_run_appends_no_trade_record(tmp_path):
    run(tmp_path)
    store = PilotStore(PILOT, str(tmp_path))
    assert store.read_trades() == []


def test_identical_rerun_idempotent(tmp_path):
    run(tmp_path)
    r2 = run(tmp_path)
    assert r2.status == rr.STATUS_ALREADY_COMMITTED_IDEMPOTENT and r2.exit_code == rr.EXIT_OK
    assert len(PilotStore(PILOT, str(tmp_path)).read_daily()) == 1


def test_changed_input_fingerprint_conflicts(tmp_path):
    run(tmp_path)
    r2 = run(tmp_path, strategy_result=strat(data_date="2026-06-21"))  # changed source data date
    assert r2.status == rr.STATUS_DAILY_PLAN_CONFLICT and r2.exit_code == rr.EXIT_CONFLICT


def test_changed_plan_fingerprint_conflicts(tmp_path):
    run(tmp_path)
    r2 = run(tmp_path, strategy_result=strat(signals=[{"symbol": "BTCUSDT", "side": "long"}]))
    assert r2.status == rr.STATUS_DAILY_PLAN_CONFLICT and r2.exit_code == rr.EXIT_CONFLICT


# ---------------------------------------------------------------------------
# 17-21. Journal path / states / atomic / malformed
# ---------------------------------------------------------------------------


def test_daily_journal_canonical_path(tmp_path):
    j = jr.DailyRunJournal(PILOT, DATE, str(tmp_path))
    assert j.dir == (tmp_path / PILOT / "daily_runs" / DATE).resolve()


def test_path_traversal_refused(tmp_path):
    for bad in ("../evil", "p/../../x", "a b", "."):
        with pytest.raises(jr.UnsafeJournalPathError):
            jr.DailyRunJournal(bad, DATE, str(tmp_path))
    for bad_date in ("2026-13-99", "../2026-01-01", "2026/06/22"):
        with pytest.raises(jr.UnsafeJournalPathError):
            jr.DailyRunJournal(PILOT, bad_date, str(tmp_path))


def test_journal_state_history_preserved(tmp_path):
    run(tmp_path)
    j = jr.DailyRunJournal(PILOT, DATE, str(tmp_path))
    states = [h["state"] for h in j.history()]
    assert states[0] == jr.RUN_INTENT_RECORDED
    assert jr.DAILY_RECORD_COMMITTED in states
    assert states[-1] == jr.RUN_COMPLETED
    assert len(states) >= 6


def test_atomic_journal_writes(tmp_path):
    run(tmp_path)
    j = jr.DailyRunJournal(PILOT, DATE, str(tmp_path))
    assert not (j.dir / (jr.RUN_JOURNAL_FILENAME + ".tmp")).exists()
    assert j.journal_path.exists()


def test_malformed_journal_fails_closed(tmp_path):
    run(tmp_path)
    j = jr.DailyRunJournal(PILOT, DATE, str(tmp_path))
    j.journal_path.write_text("{not json", encoding="utf-8")
    with pytest.raises(jr.MalformedJournalError):
        j.read()


# ---------------------------------------------------------------------------
# 22-26. Excel
# ---------------------------------------------------------------------------


def test_excel_latest_workbook_built(tmp_path):
    run(tmp_path)
    assert (tmp_path / PILOT / "demo_strategy_pilot_results.xlsx").exists()


def test_excel_dated_snapshot_built(tmp_path):
    run(tmp_path)
    assert (tmp_path / PILOT / "snapshots" / "demo_strategy_pilot_results_20260622.xlsx").exists()


def test_workbook_reopens(tmp_path):
    run(tmp_path)
    wb = load_workbook(tmp_path / PILOT / "demo_strategy_pilot_results.xlsx")
    assert wb.sheetnames[0] == "Pilot Summary"


def test_daily_row_exported_exactly_once(tmp_path):
    run(tmp_path)
    wb = load_workbook(tmp_path / PILOT / "demo_strategy_pilot_results.xlsx")
    dp = wb["Daily Performance"]
    dates = [c.value for c in dp["A"]][1:]
    assert [d for d in dates if d] == [DATE]


def test_excel_failure_does_not_duplicate_daily_record(tmp_path):
    def boom(*a, **k):
        raise RuntimeError("xlsx failed")
    result = rr.run_daily(mode="dry_run", pilot_id=PILOT, date=DATE, config=cfg(),
                          strategy_result=strat(), output_root=str(tmp_path),
                          workbook_builder=boom, snapshot_date="20260622")
    assert result.excel["status"] == "FAIL"
    assert result.status == rr.STATUS_PARTIAL_OUTPUT_FAILURE and result.exit_code == rr.EXIT_PARTIAL_OUTPUT
    assert len(PilotStore(PILOT, str(tmp_path)).read_daily()) == 1


# ---------------------------------------------------------------------------
# 27-32. Notion
# ---------------------------------------------------------------------------


def test_notion_preview_exact_required_fields(tmp_path):
    result = run(tmp_path)
    j = jr.DailyRunJournal(PILOT, DATE, str(tmp_path))
    payload = json.loads((j.dir / jr.NOTION_PAYLOAD_FILENAME).read_text(encoding="utf-8"))
    for field in ("Date", "Pilot ID", "Pilot Day", "Runner Status", "Signal Count", "Order Count",
                  "Filled Count", "Closed Trade Count", "Realized PnL USDT", "Trading Fees USDT",
                  "Funding PnL USDT", "Daily Net PnL USDT", "Cumulative Net PnL USDT", "Daily Return %",
                  "Cumulative Return %", "Max Drawdown %", "Current Position", "Alerts Triggered",
                  "Excel Export Status", "Notion Sync Status", "Discord Notify Status", "Notes",
                  "Plan Fingerprint", "Input Fingerprint"):
        assert field in payload["properties"]


def test_notion_idempotency_key():
    payload = ns.build_notion_payload(PILOT, {"date": DATE})
    assert payload["idempotency_key"] == f"{PILOT}:{DATE}"


def test_notion_network_off_by_default(tmp_path):
    result = run(tmp_path)  # allow_notion=False
    assert result.notion["status"] == ns.SYNC_SKIPPED
    assert result.notion["network_attempted"] is False


def test_fake_notion_upsert_pass(tmp_path):
    fake = FakeNotionTransport()
    sync = ns.NotionDailySync(allow_network=True, transport=fake,
                              env={"NOTION_TOKEN": "tok", "NOTION_PILOT_DATABASE_ID": "db"})
    result = run(tmp_path, notion_sync=sync, allow_notion=True)
    assert result.notion["status"] == ns.SYNC_PASS
    assert fake.upsert_calls == 1


def test_fake_notion_failure_recorded_without_daily_rerun(tmp_path):
    fake = FakeNotionTransport(raise_on="upsert")
    sync = ns.NotionDailySync(allow_network=True, transport=fake,
                              env={"NOTION_TOKEN": "tok", "NOTION_PILOT_DATABASE_ID": "db"})
    result = run(tmp_path, notion_sync=sync, allow_notion=True)
    assert result.notion["status"] == ns.SYNC_FAIL
    assert result.status == rr.STATUS_PARTIAL_OUTPUT_FAILURE and result.exit_code == rr.EXIT_PARTIAL_OUTPUT
    assert len(PilotStore(PILOT, str(tmp_path)).read_daily()) == 1


def test_reconcile_retries_only_failed_or_skipped_notion(tmp_path):
    run(tmp_path)  # notion SKIPPED (network off)
    fake = FakeNotionTransport()
    sync = ns.NotionDailySync(allow_network=True, transport=fake,
                              env={"NOTION_TOKEN": "tok", "NOTION_PILOT_DATABASE_ID": "db"})
    result = rr.run_daily(mode="reconcile_outputs", pilot_id=PILOT, date=DATE, config=cfg(),
                          strategy_result=None, output_root=str(tmp_path), notion_sync=sync,
                          allow_notion_network=True, snapshot_date="20260622")
    assert result.notion["status"] == ns.SYNC_PASS
    assert fake.upsert_calls == 1
    # daily record untouched
    assert len(PilotStore(PILOT, str(tmp_path)).read_daily()) == 1


# ---------------------------------------------------------------------------
# 33-37. Discord
# ---------------------------------------------------------------------------


def test_discord_network_off_by_default(tmp_path):
    result = run(tmp_path)
    assert result.discord["status"] == dn.NOTIFY_SKIPPED and result.discord["network_attempted"] is False


def test_discord_summary_contains_dry_run_warning(tmp_path):
    run(tmp_path)
    j = jr.DailyRunJournal(PILOT, DATE, str(tmp_path))
    summary = (j.dir / jr.DISCORD_SUMMARY_FILENAME).read_text(encoding="utf-8")
    assert "DRY-RUN／尚未授權自動下單" in summary


def test_fake_discord_pass(tmp_path):
    fake = FakeDiscordTransport()
    notify = dn.DiscordDailyNotify(allow_network=True, transport=fake,
                                   env={"MONITOR_DISCORD_WEBHOOK_URL": "http://hook"})
    result = run(tmp_path, discord_notify=notify, allow_discord=True)
    assert result.discord["status"] == dn.NOTIFY_PASS and len(fake.posts) == 1


def test_fake_discord_failure_recorded_without_daily_rerun(tmp_path):
    fake = FakeDiscordTransport(raise_on=True)
    notify = dn.DiscordDailyNotify(allow_network=True, transport=fake,
                                   env={"MONITOR_DISCORD_WEBHOOK_URL": "http://hook"})
    result = run(tmp_path, discord_notify=notify, allow_discord=True)
    assert result.discord["status"] == dn.NOTIFY_FAIL
    assert result.exit_code == rr.EXIT_PARTIAL_OUTPUT
    assert len(PilotStore(PILOT, str(tmp_path)).read_daily()) == 1


def test_reconcile_retries_only_failed_or_skipped_discord(tmp_path):
    run(tmp_path)  # discord SKIPPED
    fake = FakeDiscordTransport()
    notify = dn.DiscordDailyNotify(allow_network=True, transport=fake,
                                   env={"MONITOR_DISCORD_WEBHOOK_URL": "http://hook"})
    result = rr.run_daily(mode="reconcile_outputs", pilot_id=PILOT, date=DATE, config=cfg(),
                          strategy_result=None, output_root=str(tmp_path), discord_notify=notify,
                          allow_discord_network=True, snapshot_date="20260622")
    assert result.discord["status"] == dn.NOTIFY_PASS and len(fake.posts) == 1


# ---------------------------------------------------------------------------
# 38-44. Secrets / imports / scheduler / PnL / BO-BP exclusion
# ---------------------------------------------------------------------------


NEW_FILES = [
    ROOT / "src/demo_strategy_pilot_daily_runner.py",
    ROOT / "src/demo_strategy_pilot_daily_journal.py",
    ROOT / "src/demo_strategy_pilot_notion_sync.py",
    ROOT / "src/demo_strategy_pilot_discord_notify.py",
    ROOT / "scripts/run_demo_strategy_pilot_daily.py",
]


def test_credentials_and_webhook_absent_from_reports_and_errors(tmp_path):
    # Notion failure with token present must not leak the token.
    fake = FakeNotionTransport(raise_on="query")
    sync = ns.NotionDailySync(allow_network=True, transport=fake,
                              env={"NOTION_TOKEN": "SECRET_TOKEN_VALUE", "NOTION_PILOT_DATABASE_ID": "db"})
    res = sync.upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert "SECRET_TOKEN_VALUE" not in json.dumps(res.to_dict())
    # Discord failure must not leak webhook URL.
    fd = FakeDiscordTransport(raise_on=True)
    notify = dn.DiscordDailyNotify(allow_network=True, transport=fd,
                                   env={"MONITOR_DISCORD_WEBHOOK_URL": "http://secret-hook"})
    dres = notify.notify("msg")
    assert "secret-hook" not in json.dumps(dres.to_dict())


def _import_lines(text):
    return [ln.strip() for ln in text.splitlines() if ln.strip().startswith(("import ", "from "))]


def test_no_bybit_order_endpoint_import_or_string():
    for f in NEW_FILES:
        text = f.read_text(encoding="utf-8")
        assert "/v5/order" not in text
        assert "order/create" not in text
        assert "api-demo.bybit.com" not in text
        for ln in _import_lines(text):
            assert "executors.bybit" not in ln


def test_no_live_executor_import():
    for f in NEW_FILES:
        text = f.read_text(encoding="utf-8")
        assert "BybitExecutor" not in text
        for ln in _import_lines(text):
            assert "src.risk" not in ln
            assert not ln.startswith("import main") and not ln.startswith("from main ")


def test_no_strategy_parameter_mutation():
    for f in NEW_FILES:
        text = f.read_text(encoding="utf-8")
        # The runner must not import or rewrite strategy/config modules.
        for ln in _import_lines(text):
            assert "STRATEGY_PROFILES" not in ln
            assert not ln.startswith("from src.strategies") and not ln.startswith("import src.strategies")


def test_no_scheduler_or_cron():
    for f in NEW_FILES:
        for ln in _import_lines(f.read_text(encoding="utf-8")):
            low = ln.lower()
            assert "apscheduler" not in low and "crontab" not in low
            assert not low.startswith("import schedule") and not low.startswith("import sched")


def test_pnl_remains_zero_without_real_trades(tmp_path):
    result = run(tmp_path)
    rec = result.daily_record
    assert rec["realized_pnl_usdt"] == "0" and rec["daily_net_pnl_usdt"] == "0"
    assert rec["cumulative_net_pnl_usdt"] == "0" and rec["closed_trade_count"] == 0


def test_manual_bo_bp_trade_excluded(tmp_path):
    # No trade records are created by the dry-run runner at all.
    run(tmp_path)
    assert PilotStore(PILOT, str(tmp_path)).read_trades() == []


# ---------------------------------------------------------------------------
# 45-47. CLI / exit codes
# ---------------------------------------------------------------------------


def test_json_only_output_is_valid_json(tmp_path, capsys):
    fixture = tmp_path / "sig.json"
    fixture.write_text(json.dumps(strat()), encoding="utf-8")
    rc = cli.main(["--mode", "plan", "--pilot-id", PILOT, "--date", DATE,
                   "--fixture", str(fixture), "--json-only"])
    out = capsys.readouterr().out
    payload = json.loads(out)
    assert payload["task_id"] == "TASK-014BR" and payload["order_execution_authorized"] is False
    assert rc == rr.EXIT_OK


def test_exit_code_partial_output_failure(tmp_path):
    fake = FakeDiscordTransport(raise_on=True)
    notify = dn.DiscordDailyNotify(allow_network=True, transport=fake,
                                   env={"MONITOR_DISCORD_WEBHOOK_URL": "http://hook"})
    result = run(tmp_path, discord_notify=notify, allow_discord=True)
    assert result.exit_code == rr.EXIT_PARTIAL_OUTPUT


def test_exit_code_plan_conflict(tmp_path):
    run(tmp_path)
    r2 = run(tmp_path, strategy_result=strat(signals=[{"symbol": "BTCUSDT", "side": "long"}]))
    assert r2.exit_code == rr.EXIT_CONFLICT


def test_cli_test_output_root_refused_outside_temp(capsys):
    rc = cli.main(["--mode", "dry_run", "--date", DATE, "--test-output-root", "/var/lib/production"])
    payload = json.loads(capsys.readouterr().out)
    assert payload["status"] == "SAFETY_REFUSAL" and rc == rr.EXIT_SAFETY


# ---------------------------------------------------------------------------
# 48-52. Docs / outputs / untracked
# ---------------------------------------------------------------------------


def test_documentation_updated():
    for rel in ("README.md", "docs/research/commands/NEXT_ACTION.md",
                "docs/research/commands/COMMAND_LOG.md"):
        text = (ROOT / rel).read_text(encoding="utf-8")
        assert "TASK-014BR" in text


def test_runtime_outputs_under_gitignored_outputs():
    from src.demo_strategy_pilot_store import CANONICAL_PILOT_ROOT
    rel = CANONICAL_PILOT_ROOT.resolve().relative_to(ROOT.resolve())
    assert str(rel).replace("\\", "/") == "outputs/demo_trading/pilot"


def test_protected_symbol_constant_unchanged():
    assert rr.PROTECTED_SYMBOLS == frozenset({"ENAUSDT", "TIAUSDT", "AIXBTUSDT", "POLYXUSDT", "EDUUSDT"})


def test_frozen_plan_dataclass():
    plan = rr.build_plan(pilot_id=PILOT, date=DATE, config=cfg(), strategy_result=strat(), runner_mode="plan")
    assert dataclasses.is_dataclass(plan)
    with pytest.raises(dataclasses.FrozenInstanceError):
        plan.signal_count = 99  # type: ignore[misc]


def test_plan_mode_creates_no_state_without_output_root():
    result = rr.run_daily(mode="plan", pilot_id=PILOT, date=DATE, config=cfg(),
                          strategy_result=strat(), output_root=None)
    assert result.status == rr.STATUS_PLAN_ONLY
    assert result.daily_record is None
