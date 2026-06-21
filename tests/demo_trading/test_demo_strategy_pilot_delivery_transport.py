"""Tests for TASK-014BU explicitly gated real delivery transports + reconcile
preview finalization.

All tests are strictly offline: fake HTTP objects, temp roots, monkeypatched
factories. No test performs real HTTP, contacts Bybit, or sends an order.
"""

from __future__ import annotations

import importlib
import json
import pathlib
import sys
from decimal import Decimal

import pytest

ROOT = pathlib.Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from src import demo_strategy_pilot_daily_journal as jr
from src import demo_strategy_pilot_daily_runner as rr
from src import demo_strategy_pilot_delivery_transport as dt
from src import demo_strategy_pilot_discord_notify as dn
from src import demo_strategy_pilot_notion_sync as ns
from src.demo_strategy_pilot_reporting import PilotConfig
from src.demo_strategy_pilot_store import PilotStore

cli = importlib.import_module("scripts.run_demo_strategy_pilot_daily")

DATE = "2026-06-21"
PILOT = "P1"
# Full compatible schema covering EVERY property the finalized Pilot payload
# sends (TASK-014BU blocker 2 requires validating the complete payload).
_FULL_PAYLOAD_PROPS = list(ns.build_notion_payload(PILOT, {"date": DATE})["properties"].keys())
COMPAT_SCHEMA = {p: {"type": "rich_text"} for p in _FULL_PAYLOAD_PROPS}
COMPAT_SCHEMA["Date"] = {"type": "date"}


# ---------------------------------------------------------------------------
# Fakes
# ---------------------------------------------------------------------------


class FakeNotionHttp:
    def __init__(self, *, schema=None, existing_page=None, results=None, fail_on=None):
        self.schema = schema if schema is not None else dict(COMPAT_SCHEMA)
        self.existing_page = existing_page
        self.results = results  # explicit list of {"id":...} for multi-match tests
        self.fail_on = fail_on
        self.calls = []          # (method, path)
        self.bodies = []         # (method, path, body)
        self.query_filter = None

    def request(self, method, path, token, body=None):
        self.calls.append((method, path))
        self.bodies.append((method, path, body))
        if self.fail_on and self.fail_on in path:
            import urllib.error
            raise urllib.error.URLError("boom")
        if method == "GET" and path.startswith("/databases/"):
            return {"properties": self.schema}
        if method == "POST" and path.endswith("/query"):
            self.query_filter = (body or {}).get("filter")
            if self.results is not None:
                return {"results": list(self.results)}
            return {"results": ([{"id": self.existing_page}] if self.existing_page else [])}
        if method == "POST" and path == "/pages":
            return {"id": "new-page"}
        if method == "PATCH" and path.startswith("/pages/"):
            return {"id": path.split("/")[-1]}
        return {}

    def wrote_page(self):
        return any((m == "POST" and p == "/pages") or m == "PATCH" for m, p in self.calls)


class FakeHttpResult:
    def __init__(self, status_code, text=""):
        self.status_code = status_code
        self.text = text


class FakeDiscordHttp:
    def __init__(self, status=204):
        self.status = status
        self.posts = []

    def post_json(self, url, payload, timeout_seconds):
        self.posts.append((url, payload))
        return FakeHttpResult(self.status, "" if 200 <= self.status < 300 else "error body")


def cfg(**over):
    base = dict(pilot_id=PILOT, start_date=DATE, strategy_name=rr.EXPECTED_STRATEGY_NAME,
                initial_equity_usdt=Decimal("10000"), notion_enabled=True, discord_enabled=True)
    base.update(over)
    return PilotConfig(**base)


def sr():
    return {"data_date": DATE, "data_status": "cache_fallback",
            "signals": [{"symbol": "SOLUSDT", "side": "long"}],
            "forward_summary": {"strategy": rr.EXPECTED_STRATEGY_NAME},
            "run_key": "prev3y_crypto", "market_data_date": DATE}


def dry_run(tmp_path, *, notion_sync=None, discord_notify=None, allow_n=False, allow_d=False):
    return rr.run_daily(mode="dry_run", pilot_id=PILOT, date=DATE, config=cfg(), strategy_result=sr(),
                        output_root=str(tmp_path), notion_sync=notion_sync, discord_notify=discord_notify,
                        allow_notion_network=allow_n, allow_discord_network=allow_d, snapshot_date="20260621")


def reconcile(tmp_path, *, notion_sync=None, discord_notify=None, allow_n=False, allow_d=False):
    return rr.run_daily(mode="reconcile_outputs", pilot_id=PILOT, date=DATE, config=cfg(),
                        strategy_result=None, output_root=str(tmp_path), notion_sync=notion_sync,
                        discord_notify=discord_notify, allow_notion_network=allow_n,
                        allow_discord_network=allow_d, snapshot_date="20260621")


def notion_adapter(http, *, env=None):
    t, _ = dt.build_notion_transport(allow_network=True, http=http,
                                     env=env or {"NOTION_TOKEN": "tok", "NOTION_PILOT_DATABASE_ID": "db"})
    return ns.NotionDailySync(allow_network=True, transport=t)


def discord_adapter(http):
    t, _ = dt.build_discord_transport(allow_network=True, http=http,
                                      env={"MONITOR_DISCORD_WEBHOOK_URL": "http://h"})
    return dn.DiscordDailyNotify(allow_network=True, transport=t)


# ---------------------------------------------------------------------------
# 1-8. CLI transport construction gating
# ---------------------------------------------------------------------------


def _spy(monkeypatch):
    calls = {"notion": [], "discord": []}
    monkeypatch.setattr(cli.dt, "build_notion_transport",
                        lambda **k: (calls["notion"].append(k.get("allow_network")), (None, "NETWORK_NOT_ALLOWED"))[1])
    monkeypatch.setattr(cli.dt, "build_discord_transport",
                        lambda **k: (calls["discord"].append(k.get("allow_network")), (None, "NETWORK_NOT_ALLOWED"))[1])
    monkeypatch.setattr(cli.fs, "load_primary_forward_strategy_result", lambda **k: _FakeSrc())
    return calls


class _FakeSrc:
    def to_strategy_result(self):
        return sr()


def test_default_cli_constructs_no_network_transport(monkeypatch, capsys):
    calls = _spy(monkeypatch)
    cli.main(["--mode", "plan", "--pilot-id", PILOT, "--date", DATE, "--json-only"])
    capsys.readouterr()
    assert calls["notion"] == [False] and calls["discord"] == [False]


def test_plan_mode_never_constructs_transport(monkeypatch, capsys):
    calls = _spy(monkeypatch)
    cli.main(["--mode", "plan", "--pilot-id", PILOT, "--date", DATE, "--json-only"])
    capsys.readouterr()
    assert calls["notion"] == [False] and calls["discord"] == [False]


def test_no_network_dry_run_never_constructs_transport(monkeypatch, capsys, tmp_path):
    calls = _spy(monkeypatch)
    cli.main(["--mode", "dry_run", "--pilot-id", PILOT, "--date", DATE,
              "--test-output-root", str(tmp_path / "tmp_t"), "--json-only"])
    capsys.readouterr()
    assert calls["notion"] == [False] and calls["discord"] == [False]


def test_reconcile_without_flags_constructs_no_transport(monkeypatch, capsys, tmp_path):
    calls = _spy(monkeypatch)
    cli.main(["--mode", "reconcile_outputs", "--pilot-id", PILOT, "--date", DATE,
              "--test-output-root", str(tmp_path / "tmp_t"), "--json-only"])
    capsys.readouterr()
    assert calls["notion"] == [False] and calls["discord"] == [False]


def test_notion_flag_constructs_notion_transport_only(monkeypatch, capsys, tmp_path):
    calls = _spy(monkeypatch)
    cli.main(["--mode", "reconcile_outputs", "--pilot-id", PILOT, "--date", DATE,
              "--test-output-root", str(tmp_path / "tmp_t"), "--allow-notion-network", "--json-only"])
    capsys.readouterr()
    assert calls["notion"] == [True] and calls["discord"] == [False]


def test_discord_flag_constructs_discord_transport_only(monkeypatch, capsys, tmp_path):
    calls = _spy(monkeypatch)
    cli.main(["--mode", "reconcile_outputs", "--pilot-id", PILOT, "--date", DATE,
              "--test-output-root", str(tmp_path / "tmp_t"), "--allow-discord-network", "--json-only"])
    capsys.readouterr()
    assert calls["notion"] == [False] and calls["discord"] == [True]


def test_both_flags_construct_both_transports(monkeypatch, capsys, tmp_path):
    calls = _spy(monkeypatch)
    cli.main(["--mode", "reconcile_outputs", "--pilot-id", PILOT, "--date", DATE,
              "--test-output-root", str(tmp_path / "tmp_t"),
              "--allow-notion-network", "--allow-discord-network", "--json-only"])
    capsys.readouterr()
    assert calls["notion"] == [True] and calls["discord"] == [True]


def test_credentials_read_only_after_flag():
    # allow=False short-circuits to NETWORK_NOT_ALLOWED before any cred check.
    assert dt.build_notion_transport(allow_network=False, env={"NOTION_TOKEN": "t"})[1] == dt.NETWORK_NOT_ALLOWED
    assert dt.build_discord_transport(allow_network=False, env={"MONITOR_DISCORD_WEBHOOK_URL": "w"})[1] == dt.NETWORK_NOT_ALLOWED


# ---------------------------------------------------------------------------
# 9-16. Credential / database selection / schema
# ---------------------------------------------------------------------------


def test_missing_notion_token_fails_safely():
    t, d = dt.build_notion_transport(allow_network=True, env={"NOTION_PILOT_DATABASE_ID": "db"})
    assert t is None and d == dt.CREDENTIAL_MISSING


def test_missing_notion_database_id_fails_safely():
    t, d = dt.build_notion_transport(allow_network=True, env={"NOTION_TOKEN": "tok"})
    assert t is None and d == dt.CREDENTIAL_MISSING


def test_missing_discord_webhook_fails_safely():
    t, d = dt.build_discord_transport(allow_network=True, env={})
    assert t is None and d == dt.CREDENTIAL_MISSING


def test_pilot_database_id_preferred():
    db, source = dt.select_notion_database({"NOTION_PILOT_DATABASE_ID": "pilot_db",
                                            "NOTION_FORWARD_VALIDATION_DATABASE_ID": "fwd_db"})
    assert db == "pilot_db" and source == "pilot"


def test_forward_validation_fallback_used_when_only_option():
    db, source = dt.select_notion_database({"NOTION_FORWARD_VALIDATION_DATABASE_ID": "fwd_db"})
    assert db == "fwd_db" and source == "forward_validation_fallback"


def test_incompatible_fallback_schema_refuses_before_write():
    http = FakeNotionHttp(schema={"date": {"type": "title"}})  # missing pilot props
    t = dt.RealNotionTransport(token="tok", database_id="fwd_db", http=http, prefer_source="forward_validation_fallback")
    with pytest.raises(dt.NotionSchemaIncompatible):
        t.query(database_id="fwd_db", idempotency_key=f"{PILOT}:{DATE}")
    # No write call was attempted.
    assert not any(m == "POST" and p == "/pages" for m, p in http.calls)
    assert not any(m == "PATCH" for m, p in http.calls)


def test_no_automatic_schema_modification():
    http = FakeNotionHttp()
    t = dt.RealNotionTransport(token="tok", database_id="db", http=http)
    t.query(database_id="db", idempotency_key=f"{PILOT}:{DATE}")
    t.upsert(database_id="db", page_id=None, properties={"Date": DATE, "Pilot ID": PILOT})
    # Database schema is only read (GET), never modified (no POST/PATCH /databases).
    assert not any(p.startswith("/databases/") and m in ("POST", "PATCH") and not p.endswith("/query")
                   for m, p in http.calls)


def test_compatible_schema_permits_upsert():
    http = FakeNotionHttp()
    sync = notion_adapter(http)
    res = sync.upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.status == ns.SYNC_PASS and res.network_attempted is True


# ---------------------------------------------------------------------------
# 17-21. Notion query/create/update + idempotency + Discord once
# ---------------------------------------------------------------------------


def test_fake_notion_query_existing_page():
    http = FakeNotionHttp(existing_page="EXIST")
    t = dt.RealNotionTransport(token="tok", database_id="db", http=http)
    assert t.query(database_id="db", idempotency_key=f"{PILOT}:{DATE}") == {"page_id": "EXIST"}


def test_fake_notion_update_existing_page():
    http = FakeNotionHttp(existing_page="EXIST")
    sync = notion_adapter(http)
    res = sync.upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.status == ns.SYNC_PASS and res.page_action == "updated"
    assert any(m == "PATCH" and p == "/pages/EXIST" for m, p in http.calls)


def test_fake_notion_create_when_absent():
    http = FakeNotionHttp(existing_page=None)
    sync = notion_adapter(http)
    res = sync.upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.status == ns.SYNC_PASS and res.page_action == "created"
    assert any(m == "POST" and p == "/pages" for m, p in http.calls)


def test_notion_idempotency_key_unchanged():
    payload = ns.build_notion_payload(PILOT, {"date": DATE})
    assert payload["idempotency_key"] == f"{PILOT}:{DATE}"


def test_fake_discord_sends_exactly_once():
    http = FakeDiscordHttp(status=204)
    notify = discord_adapter(http)
    res = notify.notify("msg")
    assert res.status == dn.NOTIFY_PASS and len(http.posts) == 1


# ---------------------------------------------------------------------------
# 22-29. Duplicate protection / network_attempted / sanitization
# ---------------------------------------------------------------------------


def test_discord_pass_not_resent_by_later_reconcile(tmp_path):
    http = FakeDiscordHttp(status=204)
    dry_run(tmp_path, discord_notify=discord_adapter(http), allow_d=True)
    assert len(http.posts) == 1  # PASS during dry-run
    # Reconcile with a transport that would FAIL if used; PASS must not be retried.
    http2 = FakeDiscordHttp(status=500)
    reconcile(tmp_path, discord_notify=discord_adapter(http2), allow_d=True)
    assert len(http2.posts) == 0


def test_discord_fail_may_retry_once_per_reconcile(tmp_path):
    dry_run(tmp_path)  # discord SKIPPED
    http = FakeDiscordHttp(status=204)
    reconcile(tmp_path, discord_notify=discord_adapter(http), allow_d=True)
    assert len(http.posts) == 1


def test_real_transport_call_marks_network_attempted_true():
    http = FakeNotionHttp()
    res = notion_adapter(http).upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.network_attempted is True


def test_missing_transport_marks_network_attempted_false():
    res = ns.NotionDailySync(allow_network=True, transport=None).upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.network_attempted is False and res.status == ns.SYNC_FAIL and res.detail == "CREDENTIAL_MISSING"
    dres = dn.DiscordDailyNotify(allow_network=True, transport=None).notify("m")
    assert dres.network_attempted is False and dres.detail == "CREDENTIAL_MISSING"


def test_http_errors_sanitized():
    http = FakeNotionHttp(fail_on="/query")
    res = notion_adapter(http).upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.status == ns.SYNC_FAIL and "tok" not in res.detail


def test_token_webhook_db_absent_from_serialized_outputs(tmp_path):
    nhttp = FakeNotionHttp(fail_on="/query")
    n = dt.build_notion_transport(allow_network=True, http=nhttp,
                                  env={"NOTION_TOKEN": "SECRET_TOKEN", "NOTION_PILOT_DATABASE_ID": "SECRET_DB"})[0]
    dhttp = FakeDiscordHttp(status=500)
    d = dt.build_discord_transport(allow_network=True, http=dhttp, env={"MONITOR_DISCORD_WEBHOOK_URL": "http://secret"})[0]
    dry_run(tmp_path, notion_sync=ns.NotionDailySync(allow_network=True, transport=n,
                                                     env={"NOTION_TOKEN": "SECRET_TOKEN", "NOTION_PILOT_DATABASE_ID": "SECRET_DB"}),
            discord_notify=dn.DiscordDailyNotify(allow_network=True, transport=d, env={"MONITOR_DISCORD_WEBHOOK_URL": "http://secret"}),
            allow_n=True, allow_d=True)
    j = jr.DailyRunJournal(PILOT, DATE, str(tmp_path))
    blobs = "\n".join([
        (j.dir / jr.NOTION_PAYLOAD_FILENAME).read_text(encoding="utf-8"),
        (j.dir / jr.DISCORD_SUMMARY_FILENAME).read_text(encoding="utf-8"),
        (j.dir / jr.RUN_RESULT_FILENAME).read_text(encoding="utf-8"),
        (j.dir / jr.RUN_JOURNAL_FILENAME).read_text(encoding="utf-8"),
    ])
    assert "SECRET_TOKEN" not in blobs and "http://secret" not in blobs and "SECRET_DB" not in blobs


# ---------------------------------------------------------------------------
# 30-40. Reconcile finalization
# ---------------------------------------------------------------------------


def _row(tmp_path):
    wb = load_workbook(tmp_path / PILOT / "demo_strategy_pilot_results.xlsx")
    dp = wb["Daily Performance"]
    hdr = [c.value for c in dp[1]]
    return {hdr[i]: dp.cell(row=2, column=i + 1).value for i in range(len(hdr))}


def _preview(tmp_path):
    j = jr.DailyRunJournal(PILOT, DATE, str(tmp_path))
    return json.loads((j.dir / jr.NOTION_PAYLOAD_FILENAME).read_text(encoding="utf-8"))["properties"]


def test_reconcile_updates_status_ledger(tmp_path):
    dry_run(tmp_path)
    from src import demo_strategy_pilot_output_status as osm
    n_events_before = len(osm.OutputStatusStore(PILOT, str(tmp_path)).read_events())
    reconcile(tmp_path, notion_sync=notion_adapter(FakeNotionHttp()), allow_n=True)
    n_events_after = len(osm.OutputStatusStore(PILOT, str(tmp_path)).read_events())
    assert n_events_after > n_events_before


def test_reconcile_regenerates_notion_preview_after_fail(tmp_path):
    dry_run(tmp_path)  # SKIPPED
    reconcile(tmp_path, notion_sync=ns.NotionDailySync(allow_network=True, transport=None), allow_n=True)
    assert _preview(tmp_path)["Notion Sync Status"] == "FAIL"


def test_reconcile_regenerates_discord_preview(tmp_path):
    dry_run(tmp_path)
    reconcile(tmp_path, discord_notify=discord_adapter(FakeDiscordHttp(204)), allow_d=True)
    j = jr.DailyRunJournal(PILOT, DATE, str(tmp_path))
    summary = (j.dir / jr.DISCORD_SUMMARY_FILENAME).read_text(encoding="utf-8")
    assert "DRY-RUN／尚未授權自動下單" in summary


def test_reconcile_regenerates_previews_after_pass(tmp_path):
    dry_run(tmp_path)
    reconcile(tmp_path, notion_sync=notion_adapter(FakeNotionHttp()),
              discord_notify=discord_adapter(FakeDiscordHttp(204)), allow_n=True, allow_d=True)
    props = _preview(tmp_path)
    assert props["Notion Sync Status"] == "PASS" and props["Discord Notify Status"] == "PASS"


def test_excel_matches_final_fail_statuses(tmp_path):
    dry_run(tmp_path)
    reconcile(tmp_path, notion_sync=ns.NotionDailySync(allow_network=True, transport=None), allow_n=True)
    assert _row(tmp_path)["Notion Sync"] == "FAIL"


def test_excel_matches_final_pass_statuses(tmp_path):
    dry_run(tmp_path)
    reconcile(tmp_path, notion_sync=notion_adapter(FakeNotionHttp()),
              discord_notify=discord_adapter(FakeDiscordHttp(204)), allow_n=True, allow_d=True)
    row = _row(tmp_path)
    assert row["Notion Sync"] == "PASS" and row["Discord Notify"] == "PASS" and row["Excel Export"] == "OK"


def test_daily_record_count_remains_one(tmp_path):
    dry_run(tmp_path)
    reconcile(tmp_path, notion_sync=notion_adapter(FakeNotionHttp()), allow_n=True)
    assert len(PilotStore(PILOT, str(tmp_path)).read_daily()) == 1


def test_status_event_count_advances_only_on_change(tmp_path):
    from src import demo_strategy_pilot_output_status as osm
    dry_run(tmp_path)
    n0 = len(osm.OutputStatusStore(PILOT, str(tmp_path)).read_events())
    # Reconcile with no flags: nothing changes -> no new event.
    reconcile(tmp_path)
    n1 = len(osm.OutputStatusStore(PILOT, str(tmp_path)).read_events())
    assert n1 == n0


def test_no_trade_record(tmp_path):
    dry_run(tmp_path)
    reconcile(tmp_path, notion_sync=notion_adapter(FakeNotionHttp()), allow_n=True)
    assert PilotStore(PILOT, str(tmp_path)).read_trades() == []


def test_reconcile_never_loads_forward_source(tmp_path, monkeypatch):
    dry_run(tmp_path)
    monkeypatch.setattr(cli.fs, "load_primary_forward_strategy_result",
                        lambda **k: (_ for _ in ()).throw(AssertionError("source must not load")))
    r = reconcile(tmp_path)
    assert r.status in (rr.STATUS_RECONCILED, rr.STATUS_PARTIAL_OUTPUT_FAILURE)


# ---------------------------------------------------------------------------
# 41-44. Safety scans / JSON
# ---------------------------------------------------------------------------


def test_no_bybit_import_or_order_endpoint():
    src = (ROOT / "src/demo_strategy_pilot_delivery_transport.py").read_text(encoding="utf-8")
    assert "/v5/order" not in src and "order/create" not in src and "api-demo.bybit.com" not in src
    for ln in src.splitlines():
        s = ln.strip()
        if s.startswith(("import ", "from ")):
            assert "executors.bybit" not in s and "pybit" not in s


def test_no_live_executor_import():
    src = (ROOT / "src/demo_strategy_pilot_delivery_transport.py").read_text(encoding="utf-8")
    assert "BybitExecutor" not in src
    for ln in src.splitlines():
        s = ln.strip()
        if s.startswith(("import ", "from ")):
            assert "src.risk" not in s and not s.startswith("import main")


def test_no_retry_loop():
    src = (ROOT / "src/demo_strategy_pilot_delivery_transport.py").read_text(encoding="utf-8")
    assert "while True" not in src
    for token in ("import tenacity", "import backoff", "@retry"):
        assert token not in src


def test_json_only_stdout_valid(monkeypatch, capsys, tmp_path):
    _spy(monkeypatch)
    cli.main(["--mode", "plan", "--pilot-id", PILOT, "--date", DATE, "--json-only"])
    payload = json.loads(capsys.readouterr().out)
    assert payload["task_id"] == "TASK-014BR" and payload["order_execution_authorized"] is False


# ---------------------------------------------------------------------------
# 45-49. Docs / outputs / untouched
# ---------------------------------------------------------------------------


def test_documentation_updated():
    for rel in ("README.md", "docs/research/commands/NEXT_ACTION.md",
                "docs/research/commands/COMMAND_LOG.md"):
        text = (ROOT / rel).read_text(encoding="utf-8")
        assert "TASK-014BU" in text


def test_runtime_outputs_under_outputs_dir():
    from src.demo_strategy_pilot_store import CANONICAL_PILOT_ROOT
    rel = CANONICAL_PILOT_ROOT.resolve().relative_to(ROOT.resolve())
    assert str(rel).replace("\\", "/") == "outputs/demo_trading/pilot"


def test_delivery_status_tokens_defined():
    for tok in ("NETWORK_NOT_ALLOWED", "CREDENTIAL_MISSING", "TRANSPORT_CONSTRUCTION_FAILED",
                "NOTION_DATABASE_SCHEMA_INCOMPATIBLE", "HTTP_DELIVERY_FAILED"):
        assert hasattr(dt, tok)


def test_order_execution_remains_unauthorized():
    assert rr.ORDER_EXECUTION_AUTHORIZED is False
    assert rr.REASON_NOT_AUTHORIZED == "TASK-014BR_IS_DRY_RUN_REPORTING_WIRING_ONLY"


# ---------------------------------------------------------------------------
# TASK-014BU_FIX -- Pilot/date identity + full payload schema validation
# ---------------------------------------------------------------------------


def _payload():
    return ns.build_notion_payload(PILOT, {"date": DATE})["properties"]


def _transport(http):
    return dt.RealNotionTransport(token="tok", database_id="db", http=http)


def test_existing_page_lookup_matches_pilot_and_date():
    http = FakeNotionHttp()  # no Idempotency Key prop -> AND(Pilot ID, Date)
    _transport(http).query(database_id="db", idempotency_key=f"{PILOT}:{DATE}", properties=_payload())
    flt = http.query_filter
    assert "and" in flt
    props = {c["property"]: c for c in flt["and"]}
    assert set(props) == {"Pilot ID", "Date"}
    assert props["Pilot ID"]["rich_text"]["equals"] == PILOT
    assert props["Date"]["date"]["equals"] == DATE


def test_date_only_matching_is_impossible():
    http = FakeNotionHttp()
    _transport(http).query(database_id="db", idempotency_key=f"{PILOT}:{DATE}", properties=_payload())
    flt = http.query_filter
    # The filter is never a bare Date-only equality.
    assert not (set(flt.keys()) == {"property"} and flt.get("property") == "Date")
    assert "and" in flt  # both Pilot ID and Date are required


def test_two_pilots_same_date_remain_distinct():
    http = FakeNotionHttp()
    _transport(http).query(database_id="db", idempotency_key=f"PILOT_A:{DATE}", properties=_payload())
    fa = {c["property"]: c["rich_text"]["equals"] for c in http.query_filter["and"] if "rich_text" in c}
    http2 = FakeNotionHttp()
    _transport(http2).query(database_id="db", idempotency_key=f"PILOT_B:{DATE}", properties=_payload())
    fb = {c["property"]: c["rich_text"]["equals"] for c in http2.query_filter["and"] if "rich_text" in c}
    assert fa["Pilot ID"] == "PILOT_A" and fb["Pilot ID"] == "PILOT_B"


def test_exact_idempotency_key_preserved():
    assert ns.build_notion_payload(PILOT, {"date": DATE})["idempotency_key"] == f"{PILOT}:{DATE}"


def test_idempotency_key_property_lookup_when_available():
    schema = dict(COMPAT_SCHEMA)
    schema[dt.IDEMPOTENCY_KEY_PROPERTY] = {"type": "rich_text"}
    http = FakeNotionHttp(schema=schema)
    _transport(http).query(database_id="db", idempotency_key=f"{PILOT}:{DATE}", properties=_payload())
    flt = http.query_filter
    assert flt["property"] == dt.IDEMPOTENCY_KEY_PROPERTY
    assert flt["rich_text"]["equals"] == f"{PILOT}:{DATE}"


def test_pilot_id_date_fallback_lookup_without_key_property():
    http = FakeNotionHttp()  # COMPAT_SCHEMA has no Idempotency Key property
    _transport(http).query(database_id="db", idempotency_key=f"{PILOT}:{DATE}", properties=_payload())
    assert "and" in http.query_filter


def test_zero_matches_creates_one_page():
    http = FakeNotionHttp(results=[])
    sync = notion_adapter(http)
    res = sync.upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.status == ns.SYNC_PASS and res.page_action == "created"
    assert sum(1 for m, p in http.calls if m == "POST" and p == "/pages") == 1


def test_one_match_updates_exactly_that_page():
    http = FakeNotionHttp(results=[{"id": "PAGE-1"}])
    sync = notion_adapter(http)
    res = sync.upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.status == ns.SYNC_PASS and res.page_action == "updated"
    assert any(m == "PATCH" and p == "/pages/PAGE-1" for m, p in http.calls)
    assert not any(m == "POST" and p == "/pages" for m, p in http.calls)


def test_multiple_matches_fail_closed():
    http = FakeNotionHttp(results=[{"id": "A"}, {"id": "B"}])
    sync = notion_adapter(http)
    res = sync.upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.status == ns.SYNC_FAIL and res.detail == "NOTION_DUPLICATE_IDENTITY_CONFLICT"


def test_duplicate_conflict_performs_no_write():
    http = FakeNotionHttp(results=[{"id": "A"}, {"id": "B"}])
    notion_adapter(http).upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert not http.wrote_page()


def test_full_finalized_payload_property_set_validated():
    # Schema that has only the old weak five-field subset -> now incompatible.
    weak = {p: {"type": "rich_text"} for p in dt.REQUIRED_PILOT_SCHEMA_PROPS}
    weak["Date"] = {"type": "date"}
    http = FakeNotionHttp(schema=weak)
    res = notion_adapter(http).upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.status == ns.SYNC_FAIL and res.detail == "NOTION_DATABASE_SCHEMA_INCOMPATIBLE"
    assert not http.wrote_page()


def test_missing_pnl_property_fails_before_query_or_write():
    schema = dict(COMPAT_SCHEMA)
    del schema["Realized PnL USDT"]
    http = FakeNotionHttp(schema=schema)
    res = notion_adapter(http).upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.status == ns.SYNC_FAIL and res.detail == "NOTION_DATABASE_SCHEMA_INCOMPATIBLE"
    assert not any(p.endswith("/query") for m, p in http.calls)
    assert not http.wrote_page()


def test_missing_fingerprint_property_fails():
    schema = dict(COMPAT_SCHEMA)
    del schema["Plan Fingerprint"]
    http = FakeNotionHttp(schema=schema)
    res = notion_adapter(http).upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.status == ns.SYNC_FAIL and res.detail == "NOTION_DATABASE_SCHEMA_INCOMPATIBLE"


def test_incompatible_numeric_property_type_fails():
    schema = dict(COMPAT_SCHEMA)
    schema["Realized PnL USDT"] = {"type": "date"}  # numeric field cannot be a date
    http = FakeNotionHttp(schema=schema)
    res = notion_adapter(http).upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.status == ns.SYNC_FAIL and res.detail == "NOTION_DATABASE_SCHEMA_INCOMPATIBLE"


def test_incompatible_date_title_richtext_property_type_fails():
    schema = dict(COMPAT_SCHEMA)
    schema["Date"] = {"type": "checkbox"}  # Date cannot be a checkbox
    http = FakeNotionHttp(schema=schema)
    res = notion_adapter(http).upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.status == ns.SYNC_FAIL and res.detail == "NOTION_DATABASE_SCHEMA_INCOMPATIBLE"


def test_no_property_silently_discarded():
    # Every payload property maps to a schema property (no silent drop): the
    # number of built properties equals the number of payload properties.
    http = FakeNotionHttp(results=[])
    payload = _payload()
    _transport(http).upsert(database_id="db", page_id=None, properties=payload)
    create_body = [b for (m, p, b) in http.bodies if m == "POST" and p == "/pages"][0]
    assert len(create_body["properties"]) == len(payload)


def test_no_partial_write_on_schema_fail():
    schema = dict(COMPAT_SCHEMA)
    del schema["Notes"]
    http = FakeNotionHttp(schema=schema)
    notion_adapter(http).upsert(ns.build_notion_payload(PILOT, {"date": DATE}))
    assert not http.wrote_page()


def test_dedicated_and_fallback_db_both_full_validation():
    # Dedicated DB selected.
    db, source = dt.select_notion_database({"NOTION_PILOT_DATABASE_ID": "pilot"})
    assert source == "pilot"
    # Both paths run the same _validate_payload_schema before any write.
    weak = {p: {"type": "rich_text"} for p in dt.REQUIRED_PILOT_SCHEMA_PROPS}
    weak["Date"] = {"type": "date"}
    for source_name in ("pilot", "forward_validation_fallback"):
        http = FakeNotionHttp(schema=weak)
        t = dt.RealNotionTransport(token="tok", database_id="db", http=http, prefer_source=source_name)
        with pytest.raises(dt.NotionSchemaIncompatible):
            t.upsert(database_id="db", page_id=None, properties=_payload())
        assert not http.wrote_page()


def test_schema_failure_detail_sanitized():
    schema = dict(COMPAT_SCHEMA)
    del schema["Realized PnL USDT"]
    http = FakeNotionHttp(schema=schema)
    sync = dt.build_notion_transport(allow_network=True, http=http,
                                     env={"NOTION_TOKEN": "SECRET_TOKEN", "NOTION_PILOT_DATABASE_ID": "SECRET_DB"})[0]
    res = ns.NotionDailySync(allow_network=True, transport=sync,
                             env={"NOTION_TOKEN": "SECRET_TOKEN", "NOTION_PILOT_DATABASE_ID": "SECRET_DB"}).upsert(
        ns.build_notion_payload(PILOT, {"date": DATE}))
    assert res.status == ns.SYNC_FAIL
    assert "SECRET_TOKEN" not in res.detail and "SECRET_DB" not in res.detail
