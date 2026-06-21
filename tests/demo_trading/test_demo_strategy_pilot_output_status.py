"""Tests for TASK-014BT output-delivery status finalization.

All tests are strictly offline: temp roots, injected fake transports, no HTTP,
no orders.
"""

from __future__ import annotations

import dataclasses
import importlib
import json
import pathlib
import sys
from decimal import Decimal

import pytest

ROOT = pathlib.Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from src import demo_strategy_pilot_daily_journal as jr
from src import demo_strategy_pilot_daily_runner as rr
from src import demo_strategy_pilot_discord_notify as dn
from src import demo_strategy_pilot_notion_sync as ns
from src import demo_strategy_pilot_output_status as osm
from src.demo_strategy_pilot_reporting import PilotConfig
from src.demo_strategy_pilot_store import PilotStore

DATE = "2026-05-18"


def cfg(**over):
    base = dict(pilot_id="P1", start_date=DATE, strategy_name=rr.EXPECTED_STRATEGY_NAME,
                initial_equity_usdt=Decimal("10000"), notion_enabled=True, discord_enabled=True)
    base.update(over)
    return PilotConfig(**base)


def sr():
    return {"data_date": DATE, "data_status": "cache_fallback",
            "signals": [{"symbol": "SOLUSDT", "side": "long"}],
            "forward_summary": {"strategy": rr.EXPECTED_STRATEGY_NAME},
            "run_key": "prev3y_crypto", "market_data_date": DATE}


class FakeNotionTransport:
    def __init__(self, *, raise_on=None):
        self.raise_on = raise_on
        self.upsert_calls = 0

    def query(self, *, database_id, idempotency_key, headers, properties=None):
        if self.raise_on == "query":
            raise RuntimeError("notion query failed")
        return {}

    def upsert(self, *, database_id, page_id, properties, headers):
        self.upsert_calls += 1
        if self.raise_on == "upsert":
            raise RuntimeError("notion upsert failed")
        return {"page_id": "p"}


class FakeDiscordTransport:
    def __init__(self, *, raise_on=False):
        self.raise_on = raise_on
        self.posts = []

    def post(self, *, webhook_url, content):
        self.posts.append(content)
        if self.raise_on:
            raise RuntimeError("discord post failed")
        return True


def run(tmp_path, *, notion_sync=None, discord_notify=None, allow_notion=False, allow_discord=False):
    return rr.run_daily(mode="dry_run", pilot_id="P1", date=DATE, config=cfg(), strategy_result=sr(),
                        output_root=str(tmp_path), notion_sync=notion_sync, discord_notify=discord_notify,
                        allow_notion_network=allow_notion, allow_discord_network=allow_discord,
                        snapshot_date="20260518")


def status_rec(**over):
    base = dict(pilot_id="P1", date=DATE, excel_status="OK", notion_status="SKIPPED",
                discord_status="SKIPPED", excel_detail="", notion_detail="", discord_detail="",
                updated_at_utc="2026-05-18T00:00:00Z", plan_fingerprint="pf", input_fingerprint="if",
                daily_core_fingerprint="cf")
    base.update(over)
    return osm.OutputStatusRecord(**base)


# ---------------------------------------------------------------------------
# 1-10. Model + store + immutable core
# ---------------------------------------------------------------------------


def test_frozen_output_status_model():
    assert dataclasses.is_dataclass(osm.OutputStatusRecord)
    r = status_rec()
    with pytest.raises(dataclasses.FrozenInstanceError):
        r.excel_status = "FAIL"  # type: ignore[misc]


def test_allowed_status_validation():
    assert osm.ALLOWED_STATUSES == {"PENDING", "OK", "PASS", "FAIL", "SKIPPED"}
    with pytest.raises(osm.InvalidStatusError):
        status_rec(excel_status="WEIRD")


def test_append_only_status_event(tmp_path):
    st = osm.OutputStatusStore("P1", str(tmp_path))
    _, a1 = st.record_status(status_rec(notion_status="SKIPPED"))
    _, a2 = st.record_status(status_rec(notion_status="PASS"))
    assert a1 and a2
    assert len(st.read_events()) == 2


def test_atomic_latest_status_write(tmp_path):
    st = osm.OutputStatusStore("P1", str(tmp_path))
    st.record_status(status_rec())
    assert not (st.dir / (osm.LATEST_OUTPUT_STATUS_FILENAME + ".tmp")).exists()
    assert st.read_latest()["excel_status"] == "OK"


def test_identical_status_event_idempotent(tmp_path):
    st = osm.OutputStatusStore("P1", str(tmp_path))
    _, a1 = st.record_status(status_rec())
    _, a2 = st.record_status(status_rec())
    assert a1 is True and a2 is False
    assert len(st.read_events()) == 1


def test_malformed_status_ledger_refusal(tmp_path):
    st = osm.OutputStatusStore("P1", str(tmp_path))
    st.record_status(status_rec())
    with open(st.events_path, "a", encoding="utf-8") as fh:
        fh.write("{bad json\n")
    with pytest.raises(osm.MalformedStatusLedgerError):
        st.read_events()


def test_immutable_daily_fingerprint_deterministic():
    rec = {"date": DATE, "signal_count": 50, "order_count": 0, "filled_count": 0,
           "closed_trade_count": 0, "realized_pnl_usdt": "0", "trading_fees_usdt": "0",
           "funding_pnl_usdt": "0", "daily_net_pnl_usdt": "0", "cumulative_net_pnl_usdt": "0",
           "daily_return_pct": "0", "cumulative_return_pct": "0", "max_drawdown_pct": "0",
           "current_position_symbol": "", "current_position_side": "", "current_position_qty": "0"}
    a = osm.compute_daily_core_fingerprint(pilot_id="P1", daily_record=rec, input_fingerprint="i", plan_fingerprint="p")
    b = osm.compute_daily_core_fingerprint(pilot_id="P1", daily_record=rec, input_fingerprint="i", plan_fingerprint="p")
    assert a == b


def test_trading_field_mutation_refused():
    rec = {"date": DATE, "signal_count": 50}
    base = osm.compute_daily_core_fingerprint(pilot_id="P1", daily_record={**rec, "order_count": 0},
                                              input_fingerprint="i", plan_fingerprint="p")
    changed = osm.compute_daily_core_fingerprint(pilot_id="P1", daily_record={**rec, "order_count": 1},
                                                 input_fingerprint="i", plan_fingerprint="p")
    assert base != changed


def test_pnl_and_position_mutation_refused():
    rec = {"date": DATE, "signal_count": 1, "realized_pnl_usdt": "0", "current_position_qty": "0"}
    a = osm.compute_daily_core_fingerprint(pilot_id="P1", daily_record=rec, input_fingerprint="i", plan_fingerprint="p")
    b = osm.compute_daily_core_fingerprint(pilot_id="P1", daily_record={**rec, "realized_pnl_usdt": "5"},
                                           input_fingerprint="i", plan_fingerprint="p")
    c = osm.compute_daily_core_fingerprint(pilot_id="P1", daily_record={**rec, "current_position_qty": "0.1"},
                                           input_fingerprint="i", plan_fingerprint="p")
    assert a != b and a != c


def test_assert_immutable_core_conflict(tmp_path):
    st = osm.OutputStatusStore("P1", str(tmp_path))
    st.record_status(status_rec(daily_core_fingerprint="CORE_A"))
    st.assert_immutable_core_unchanged(date=DATE, expected_core_fp="CORE_A")  # ok
    with pytest.raises(osm.ImmutableDailyCoreConflict):
        st.assert_immutable_core_unchanged(date=DATE, expected_core_fp="CORE_B")


# ---------------------------------------------------------------------------
# 11-13. Daily record single / immutable / no trade
# ---------------------------------------------------------------------------


def test_initial_daily_record_appended_once(tmp_path):
    run(tmp_path)
    assert len(PilotStore("P1", str(tmp_path)).read_daily()) == 1


def test_no_second_daily_record_during_finalization(tmp_path):
    run(tmp_path)
    # finalization rebuilds Excel + ledger but must not append another record.
    assert len(PilotStore("P1", str(tmp_path)).read_daily()) == 1


def test_no_trade_record_appended(tmp_path):
    run(tmp_path)
    assert PilotStore("P1", str(tmp_path)).read_trades() == []


# ---------------------------------------------------------------------------
# 14-20. No-network finalized statuses (Excel)
# ---------------------------------------------------------------------------


def _excel_status_row(tmp_path):
    wb = load_workbook(tmp_path / "P1" / "demo_strategy_pilot_results.xlsx")
    dp = wb["Daily Performance"]
    hdr = [c.value for c in dp[1]]
    return {hdr[i]: dp.cell(row=2, column=i + 1).value for i in range(len(hdr))}, wb


def test_no_network_final_excel_status_ok(tmp_path):
    run(tmp_path)
    row, _ = _excel_status_row(tmp_path)
    assert row["Excel Export"] == "OK"


def test_no_network_final_notion_status_skipped(tmp_path):
    run(tmp_path)
    row, _ = _excel_status_row(tmp_path)
    assert row["Notion Sync"] == "SKIPPED"


def test_no_network_final_discord_status_skipped(tmp_path):
    run(tmp_path)
    row, _ = _excel_status_row(tmp_path)
    assert row["Discord Notify"] == "SKIPPED"


def test_excel_daily_performance_shows_final_statuses(tmp_path):
    run(tmp_path)
    row, _ = _excel_status_row(tmp_path)
    assert (row["Excel Export"], row["Notion Sync"], row["Discord Notify"]) == ("OK", "SKIPPED", "SKIPPED")


def test_workbook_contains_exactly_one_date_row(tmp_path):
    run(tmp_path)
    wb = load_workbook(tmp_path / "P1" / "demo_strategy_pilot_results.xlsx")
    dates = [c.value for c in wb["Daily Performance"]["A"]][1:]
    assert [d for d in dates if d] == [DATE]


def test_workbook_reopens(tmp_path):
    run(tmp_path)
    wb = load_workbook(tmp_path / "P1" / "demo_strategy_pilot_results.xlsx")
    assert wb.sheetnames == [
        "Pilot Summary", "Daily Performance", "Trades",
        "Execution Quality", "Forward Comparison", "Audit Log"]


def test_snapshot_shows_final_statuses(tmp_path):
    run(tmp_path)
    snap = tmp_path / "P1" / "snapshots" / "demo_strategy_pilot_results_20260518.xlsx"
    wb = load_workbook(snap)
    dp = wb["Daily Performance"]
    hdr = [c.value for c in dp[1]]
    row = {hdr[i]: dp.cell(row=2, column=i + 1).value for i in range(len(hdr))}
    assert row["Excel Export"] == "OK" and row["Notion Sync"] == "SKIPPED"


# ---------------------------------------------------------------------------
# 21-24. Notion/Discord finalized previews
# ---------------------------------------------------------------------------


def _journal(tmp_path):
    return jr.DailyRunJournal("P1", DATE, str(tmp_path))


def test_final_notion_preview_shows_final_statuses(tmp_path):
    run(tmp_path)
    payload = json.loads((_journal(tmp_path).dir / jr.NOTION_PAYLOAD_FILENAME).read_text(encoding="utf-8"))
    props = payload["properties"]
    assert props["Excel Export Status"] == "OK"
    assert props["Notion Sync Status"] == "SKIPPED"
    assert props["Discord Notify Status"] == "SKIPPED"


def test_notion_idempotency_key_unchanged(tmp_path):
    run(tmp_path)
    payload = json.loads((_journal(tmp_path).dir / jr.NOTION_PAYLOAD_FILENAME).read_text(encoding="utf-8"))
    assert payload["idempotency_key"] == f"P1:{DATE}"


def test_final_discord_preview_dry_run_warning_and_statuses(tmp_path):
    run(tmp_path)
    summary = (_journal(tmp_path).dir / jr.DISCORD_SUMMARY_FILENAME).read_text(encoding="utf-8")
    assert "DRY-RUN／尚未授權自動下單" in summary
    assert "Excel 匯出狀態：OK" in summary
    assert "Notion 同步狀態：SKIPPED" in summary


# ---------------------------------------------------------------------------
# 25-28. Fake network advances status
# ---------------------------------------------------------------------------


def test_fake_notion_pass_advances_status(tmp_path):
    sync = ns.NotionDailySync(allow_network=True, transport=FakeNotionTransport(),
                              env={"NOTION_TOKEN": "t", "NOTION_PILOT_DATABASE_ID": "db"})
    result = run(tmp_path, notion_sync=sync, allow_notion=True)
    assert result.notion["status"] == "PASS"
    row, _ = _excel_status_row(tmp_path)
    assert row["Notion Sync"] == "PASS"


def test_fake_notion_fail_records_fail(tmp_path):
    sync = ns.NotionDailySync(allow_network=True, transport=FakeNotionTransport(raise_on="upsert"),
                              env={"NOTION_TOKEN": "t", "NOTION_PILOT_DATABASE_ID": "db"})
    result = run(tmp_path, notion_sync=sync, allow_notion=True)
    assert result.notion["status"] == "FAIL" and result.exit_code == rr.EXIT_PARTIAL_OUTPUT
    row, _ = _excel_status_row(tmp_path)
    assert row["Notion Sync"] == "FAIL"
    assert len(PilotStore("P1", str(tmp_path)).read_daily()) == 1


def test_fake_discord_pass_advances_status(tmp_path):
    notify = dn.DiscordDailyNotify(allow_network=True, transport=FakeDiscordTransport(),
                                   env={"MONITOR_DISCORD_WEBHOOK_URL": "http://h"})
    result = run(tmp_path, discord_notify=notify, allow_discord=True)
    assert result.discord["status"] == "PASS"
    row, _ = _excel_status_row(tmp_path)
    assert row["Discord Notify"] == "PASS"


def test_fake_discord_fail_records_fail(tmp_path):
    notify = dn.DiscordDailyNotify(allow_network=True, transport=FakeDiscordTransport(raise_on=True),
                                   env={"MONITOR_DISCORD_WEBHOOK_URL": "http://h"})
    result = run(tmp_path, discord_notify=notify, allow_discord=True)
    assert result.discord["status"] == "FAIL"
    row, _ = _excel_status_row(tmp_path)
    assert row["Discord Notify"] == "FAIL"


# ---------------------------------------------------------------------------
# 29-35. Reconcile / idempotency / conflict
# ---------------------------------------------------------------------------


def _reconcile(tmp_path, **kw):
    return rr.run_daily(mode="reconcile_outputs", pilot_id="P1", date=DATE, config=cfg(),
                        strategy_result=None, output_root=str(tmp_path), snapshot_date="20260518", **kw)


def test_reconcile_retries_only_fail_or_skipped(tmp_path):
    run(tmp_path)  # notion SKIPPED
    sync = ns.NotionDailySync(allow_network=True, transport=FakeNotionTransport(),
                              env={"NOTION_TOKEN": "t", "NOTION_PILOT_DATABASE_ID": "db"})
    result = _reconcile(tmp_path, notion_sync=sync, allow_notion_network=True)
    assert result.notion["status"] == "PASS"
    row, _ = _excel_status_row(tmp_path)
    assert row["Notion Sync"] == "PASS"


def test_reconcile_leaves_pass_untouched(tmp_path):
    sync = ns.NotionDailySync(allow_network=True, transport=FakeNotionTransport(),
                              env={"NOTION_TOKEN": "t", "NOTION_PILOT_DATABASE_ID": "db"})
    run(tmp_path, notion_sync=sync, allow_notion=True)  # notion PASS
    # Reconcile with a transport that would FAIL if used; PASS must not be retried.
    sync2 = ns.NotionDailySync(allow_network=True, transport=FakeNotionTransport(raise_on="upsert"),
                               env={"NOTION_TOKEN": "t", "NOTION_PILOT_DATABASE_ID": "db"})
    result = _reconcile(tmp_path, notion_sync=sync2, allow_notion_network=True)
    assert result.notion["status"] == "PASS"


def test_reconcile_never_loads_forward_source(tmp_path, monkeypatch):
    run(tmp_path)
    import scripts.run_demo_strategy_pilot_daily as runcli
    monkeypatch.setattr(runcli.fs, "load_primary_forward_strategy_result",
                        lambda **k: (_ for _ in ()).throw(AssertionError("source must not load")))
    # Runner reconcile takes strategy_result=None and never calls the adapter.
    result = _reconcile(tmp_path)
    assert result.status in (rr.STATUS_RECONCILED, rr.STATUS_PARTIAL_OUTPUT_FAILURE)


def test_reconcile_never_adds_daily_or_trade(tmp_path):
    run(tmp_path)
    _reconcile(tmp_path)
    store = PilotStore("P1", str(tmp_path))
    assert len(store.read_daily()) == 1 and store.read_trades() == []


def test_identical_full_dry_run_remains_idempotent(tmp_path):
    run(tmp_path)
    r2 = run(tmp_path)
    assert r2.status == rr.STATUS_ALREADY_COMMITTED_IDEMPOTENT


def test_input_plan_conflict_still_fails_closed(tmp_path):
    run(tmp_path)
    other = dict(sr(), signals=[{"symbol": "BTCUSDT", "side": "long"}])
    r2 = rr.run_daily(mode="dry_run", pilot_id="P1", date=DATE, config=cfg(), strategy_result=other,
                      output_root=str(tmp_path), snapshot_date="20260518")
    assert r2.status == rr.STATUS_DAILY_PLAN_CONFLICT and r2.exit_code == rr.EXIT_CONFLICT


# ---------------------------------------------------------------------------
# 36-43. Safety / docs / outputs
# ---------------------------------------------------------------------------


NEW_FILES = [ROOT / "src/demo_strategy_pilot_output_status.py"]


def test_no_bybit_order_imports_or_endpoint():
    src = (ROOT / "src/demo_strategy_pilot_output_status.py").read_text(encoding="utf-8")
    assert "/v5/order" not in src and "order/create" not in src and "api-demo.bybit.com" not in src
    for ln in src.splitlines():
        s = ln.strip()
        if s.startswith(("import ", "from ")):
            assert "requests" not in s and "executors.bybit" not in s


def test_no_live_executor_import():
    src = (ROOT / "src/demo_strategy_pilot_output_status.py").read_text(encoding="utf-8")
    assert "BybitExecutor" not in src
    for ln in src.splitlines():
        s = ln.strip()
        if s.startswith(("import ", "from ")):
            assert "src.risk" not in s and not s.startswith("import main")


def test_no_credentials_in_outputs(tmp_path):
    sync = ns.NotionDailySync(allow_network=True, transport=FakeNotionTransport(raise_on="query"),
                              env={"NOTION_TOKEN": "SECRET_TOK", "NOTION_PILOT_DATABASE_ID": "db"})
    notify = dn.DiscordDailyNotify(allow_network=True, transport=FakeDiscordTransport(raise_on=True),
                                   env={"MONITOR_DISCORD_WEBHOOK_URL": "http://secret-hook"})
    run(tmp_path, notion_sync=sync, discord_notify=notify, allow_notion=True, allow_discord=True)
    j = _journal(tmp_path)
    blobs = [
        (j.dir / jr.NOTION_PAYLOAD_FILENAME).read_text(encoding="utf-8"),
        (j.dir / jr.DISCORD_SUMMARY_FILENAME).read_text(encoding="utf-8"),
        (j.dir / jr.RUN_RESULT_FILENAME).read_text(encoding="utf-8"),
        (osm.OutputStatusStore("P1", str(tmp_path)).latest_path).read_text(encoding="utf-8"),
    ]
    for b in blobs:
        assert "SECRET_TOK" not in b and "secret-hook" not in b


def test_status_event_no_secret(tmp_path):
    st = osm.OutputStatusStore("P1", str(tmp_path))
    st.record_status(status_rec(notion_detail="ok", discord_detail="ok"))
    assert "TOKEN" not in st.events_path.read_text(encoding="utf-8")


def test_documentation_updated():
    for rel in ("README.md", "docs/research/commands/NEXT_ACTION.md",
                "docs/research/commands/COMMAND_LOG.md"):
        text = (ROOT / rel).read_text(encoding="utf-8")
        assert "TASK-014BT" in text


def test_runtime_outputs_under_outputs_dir():
    from src.demo_strategy_pilot_store import CANONICAL_PILOT_ROOT
    rel = CANONICAL_PILOT_ROOT.resolve().relative_to(ROOT.resolve())
    assert str(rel).replace("\\", "/") == "outputs/demo_trading/pilot"


def test_excel_finalization_failure_recorded_without_duplicate(tmp_path, monkeypatch):
    # First Excel build OK, second (finalization) build fails -> recorded FAIL,
    # no second daily record.
    import scripts.build_demo_strategy_pilot_workbook as wbmod
    calls = {"n": 0}
    real = wbmod.build_workbook

    def flaky(pilot_id, output_root=None, **kw):
        calls["n"] += 1
        if calls["n"] >= 2:
            raise RuntimeError("finalization rebuild failed")
        return real(pilot_id, output_root, **kw)

    result = rr.run_daily(mode="dry_run", pilot_id="P1", date=DATE, config=cfg(), strategy_result=sr(),
                          output_root=str(tmp_path), workbook_builder=flaky, snapshot_date="20260518")
    assert result.excel["status"] == "FAIL"
    assert len(PilotStore("P1", str(tmp_path)).read_daily()) == 1


def test_output_status_in_run_result(tmp_path):
    run(tmp_path)
    rd = json.loads((_journal(tmp_path).dir / jr.RUN_RESULT_FILENAME).read_text(encoding="utf-8"))
    assert rd["output_status"]["excel_status"] == "OK"
    assert rd["output_status"]["notion_status"] == "SKIPPED"
